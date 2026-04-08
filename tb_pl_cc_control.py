import json
import os
import hashlib
import re
import sys
from collections import Counter, defaultdict
import math
from datetime import datetime
import time
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from PySide6.QtCore import Qt, QTimer, QThread, Signal
from PySide6.QtGui import QColor, QFont
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QHeaderView,
)

BASE_DIR = r"C:\TB-PL-CC"
NOTES_JSON = os.path.join(BASE_DIR, "notes.json")
RESPONSIBLES_JSON = os.path.join(BASE_DIR, "responsibles.json")
USERS_JSON = os.path.join(BASE_DIR, "users.json")
MUAVIN_EXPORT_DEFAULT = os.path.join(BASE_DIR, "Muavin_Analiz_Raporu.xlsx")

DEFAULT_SECTION_PERMISSIONS = {
    "dashboard": True,
    "analysis": True,
    "analysis_tb": True,
    "analysis_plcc": True,
    "analysis_balance": True,
    "analysis_income": True,
    "notes": True,
    "responsibles": True,
    "muavin": True,
    "muavin_user_based": True,
    "muavin_tax_based": True,
    "muavin_account_content": True,
    "muavin_finding": True,
    "muavin_text": True,
    "muavin_period": True,
    "muavin_doctype": True,
    "muavin_risk_user": True,
    "muavin_user": True,
    "muavin_contra": True,
    "muavin_drilldown": True,
    "muavin_document_lines": True,
    "muavin_account_doc_relation": True,
    "muavin_late7": True,
    "muavin_doccheck": True,
    "muavin_dupref": True,
    "muavin_duprefdetail": True,
    "muavin_taxref": True,
    "muavin_taxvendor": True,
    "muavin_tax": True,
    "muavin_cost": True,
    "regular_ft": True,
    "user_management": True,
}

MUAVIN_SECTION_PERMISSION_MAP = {
    "finding": "muavin_finding",
    "text": "muavin_text",
    "period": "muavin_period",
    "doctype": "muavin_doctype",
    "risk_user": "muavin_risk_user",
    "user": "muavin_user",
    "contra": "muavin_contra",
    "drilldown": "muavin_drilldown",
    "document_lines": "muavin_document_lines",
    "account_doc_relation": "muavin_account_doc_relation",
    "late7": "muavin_late7",
    "doccheck": "muavin_doccheck",
    "dupref": "muavin_dupref",
    "duprefdetail": "muavin_duprefdetail",
    "taxref": "muavin_taxref",
    "taxvendor": "muavin_taxvendor",
    "tax": "muavin_tax",
    "cost": "muavin_cost",
}

OPENING_PERIOD_TR = "Açılış"
OPENING_PERIOD_EN = "Opening"

ALL_PERIODS = [
    {"tr": "01-Şubat", "en": "01-February"},
    {"tr": "02-Mart", "en": "02-March"},
    {"tr": "03-Nisan", "en": "03-April"},
    {"tr": "04-Mayıs", "en": "04-May"},
    {"tr": "05-Haziran", "en": "05-June"},
    {"tr": "06-Temmuz", "en": "06-July"},
    {"tr": "07-Ağustos", "en": "07-August"},
    {"tr": "08-Eylül", "en": "08-September"},
    {"tr": "09-Ekim", "en": "09-October"},
    {"tr": "10-Kasım", "en": "10-November"},
    {"tr": "11-Aralık", "en": "11-December"},
    {"tr": "12-Ocak", "en": "12-January"},
]

PERIOD_ORDER = {p["tr"]: i + 1 for i, p in enumerate(ALL_PERIODS)}

HESAP_YONU_MAP = {
    "100": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "101": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "102": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "103": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Alacak"},
    "108": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "110": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "111": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "112": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "118": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "119": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Alacak"},
    "120": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "121": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "122": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Alacak"},
    "124": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Alacak"},
    "126": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "127": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "128": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "129": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Alacak"},
    "131": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "132": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "133": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "135": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "136": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "137": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Alacak"},
    "138": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "139": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Alacak"},
    "150": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "151": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "152": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "153": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "157": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "158": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Alacak"},
    "159": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "170": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "178": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "179": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "180": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "181": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "190": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "191": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "192": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "193": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "195": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "196": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "197": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "198": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Borç"},
    "199": {"hesapSinifi": "Dönen Varlıklar", "beklenenYon": "Alacak"},
    "220": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "221": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "222": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "224": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "226": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "229": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "231": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "232": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "233": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "235": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "236": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "237": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "239": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "240": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "241": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "242": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "243": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "244": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "245": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "246": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "247": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "248": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "249": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "250": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "251": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "252": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "253": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "254": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "255": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "256": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "257": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "258": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "259": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "260": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "261": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "262": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "263": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "264": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "267": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "268": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "269": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "271": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "272": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "277": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "278": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "279": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "280": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "281": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "291": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "292": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "293": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "294": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "295": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "297": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Borç"},
    "298": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "299": {"hesapSinifi": "Duran Varlıklar", "beklenenYon": "Alacak"},
    "300": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "301": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "302": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Borç"},
    "303": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "304": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "305": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "306": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "308": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Borç"},
    "309": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "320": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "321": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "322": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Borç"},
    "326": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "329": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "331": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "332": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "333": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "335": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "336": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "337": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Borç"},
    "340": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "349": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "350": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "358": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "360": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "361": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "368": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "369": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "370": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "371": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Borç"},
    "372": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "373": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "379": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "380": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "381": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "391": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "392": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "393": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "397": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "399": {"hesapSinifi": "Kısa Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "400": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "401": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "402": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Borç"},
    "405": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "407": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "408": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Borç"},
    "409": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "420": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "421": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "422": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Borç"},
    "426": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "429": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "431": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "432": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "433": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "436": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "437": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Borç"},
    "438": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "440": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "449": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "472": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "479": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "480": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "481": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "492": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "493": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "499": {"hesapSinifi": "Uzun Vadeli Yabancı Kaynaklar", "beklenenYon": "Alacak"},
    "500": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "501": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Borç"},
    "502": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "503": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Borç"},
    "520": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "521": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "522": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "523": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "525": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "526": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "529": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "540": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "541": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "542": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "548": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "549": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "570": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "580": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Borç"},
    "590": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Alacak"},
    "591": {"hesapSinifi": "Öz Kaynaklar", "beklenenYon": "Borç"},
    "600": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "601": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "602": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "610": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "611": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "612": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "620": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "621": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "622": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "623": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "630": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "631": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "632": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "640": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "641": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "642": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "643": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "644": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "645": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "646": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "647": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "648": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "649": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "653": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "654": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "655": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "656": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "657": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "658": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "659": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "660": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "661": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "671": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "679": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Alacak"},
    "680": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "681": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "689": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "691": {"hesapSinifi": "Gelir Tablosu", "beklenenYon": "Borç"},
    "710": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "720": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "730": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "740": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "750": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "760": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "770": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "780": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "790": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "791": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "792": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "793": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "794": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "795": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "796": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
    "797": {"hesapSinifi": "Maliyet Hesapları", "beklenenYon": "Borç"},
}

TRANSLATIONS = {
    "tr": {
        "appTitle": "TB & PL-CC Control",
        "appDesc": "TB ve PL-CC analizlerini, dönemsel değişimleri, not eşleştirmelerini ve VUK yön kontrolünü tek ekranda yöneten premium kontrol arayüzü.",
        "start": "Analizi Başlat",
        "loadedTb": "Yüklenen TB",
        "loadedPlcc": "Yüklenen PL-CC",
        "definedNotes": "Tanımlı Not",
        "findings": "Bulgular",
        "panel": "Kontrol Paneli",
        "panelDesc": "Program akışı, dosya yönetimi ve hızlı işlemler",
        "uploadTb": "TB Yükle",
        "uploadPlcc": "PL-CC Yükle",
        "noteDefinitions": "Not Tanımları",
        "responsibleAssignments": "Sorumlu Tayin",
        "muavinAnalysis": "Muavin Analiz",
        "processingStatus": "İşlem Durumu",
        "processingText": "TB ve PL-CC dosyalarını yükledikten sonra Analizi Başlat ile hesaplamaları çalıştırabilirsiniz.",
        "dashboard": "Dashboard",
        "analysis": "Analiz",
        "dashboardTitle": "Dosya ve Dönem Yönetimi",
        "dashboardDesc": "Yüklenen dosyaları ve karşılaştırılacak dönemleri bu ekrandan yönetebilirsiniz.",
        "tbFile": "TB Dosyası",
        "plccFile": "PL-CC Dosyası",
        "currentPeriod": "Son Dönem",
        "previousPeriod": "Karşılaştırma Dönemi",
        "tbReady": "TB Dosyası Yüklü",
        "plccReady": "PL-CC Dosyası Yüklü",
        "excelStyle": "Excel başlık stili: Premium Mavi / Beyaz",
        "analysisTitle": "Analiz Ekranı",
        "analysisDesc": "TB ve PL-CC arasında geçiş yaparak tek ekranda detay analiz yapabilirsiniz.",
        "searchTb": "TB içinde ara",
        "searchPlcc": "PL-CC içinde ara",
        "periodCompare": "Dönem karşılaştırması",
        "rowDensity": "Satır Aralığı",
        "compact": "Sık",
        "normal": "Normal",
        "wide": "Geniş",
        "exportExcel": "Excel Çıktısı",
        "autoFitColumns": "Kolonları Otomatik Sığdır",
        "toggleInfo": "Tümünü göster / sadece bulgular filtresi ile tabloyu yönetebilirsiniz.",
        "showAll": "Tümünü Göster",
        "findingsOnly": "Sadece Bulgular",
        "allAccounts": "Tüm Hesaplar",
        "balanceSheet": "Bilanço",
        "incomeStatement": "Gelir Tablosu",
        "tb": "TB",
        "plcc": "PL-CC",
        "account": "Hesap",
        "accountName": "Hesap Adı",
        "note": "Not",
        "responsible": "Sorumlu",
        "costCenter": "Masraf Yeri",
        "description": "Açıklama",
        "subtotal": "Alt Toplam",
        "noteDefTitle": "Not Tanımlama Ekranı",
        "noteDefDesc": "Hesap zorunlu, masraf yeri opsiyonel. Tanımlanan not ilgili raporlara otomatik bağlanır.",
        "accountCode": "Hesap Kodu",
        "optionalCostCenter": "Masraf Yeri (Opsiyonel)",
        "notePlaceholder": "Bu hesap için gösterilecek açıklama / uyarı notu",
        "save": "Kaydet",
        "clear": "Temizle",
        "matchInfo": "Eşleşme mantığı: Birden fazla kural varsa notlar | ile birleştirilir.",
        "savedNotes": "Kayıtlı Notlar",
        "savedNotesDesc": "Program içinde yönetilen açıklama ve uyarı listesi",
        "responsibleAssignTitle": "Sorumlu Tayin Ekranı",
        "responsibleAssignDesc": "Excel’den iki kolon halinde Hesap ve Sorumlu verisini yapıştırın. Program toplu şekilde kaydeder ve TB / PL-CC tarafında otomatik eşleştirir.",
        "pasteArea": "Excel Yapıştırma Alanı",
        "pasteResponsiblePlaceholder": "Hesap\tSorumlu\n100\tAyşe Yılmaz\n320\tMehmet Kaya",
        "bulkSave": "Toplu Kaydet",
        "clearAll": "Temizle",
        "savedResponsibles": "Kayıtlı Sorumlular",
        "savedResponsiblesDesc": "Hesap bazında atanan sorumlu listesi",
        "pasteFormatInfo": "Format: Her satırda Hesap ve Sorumlu olacak şekilde Excel’den doğrudan yapıştırabilirsiniz.",
        "notesControlInfo": "Satır aralığı değiştirilebilir.",
        "general": "Genel",
        "findingsTb": "TB Bulguları",
        "findingsPlcc": "PL-CC Bulguları",
        "compareChange": "% Değişim",
        "warningCount": "adet ters çalışma uyarısı",
        "normalState": "Normal",
        "reverseState": "Ters duruyor",
        "status": "Durum",
        "premiumView": "Premium görünüm aktif",
        "analysisReady": "Analiz hazır",
        "notLoadedTb": "TB yüklenmedi",
        "notLoadedPlcc": "PL-CC yüklenmedi",
        "opening": "Açılış",
        "rowsLoaded": "satır yüklendi",
        "total": "Toplam",
        "muavinAnalysis": "Muavin Analiz",
        "regularFtAnalysis": "Düzenli Gelen Ft Analiz",
        "loadData": "Yükle",
        "runAnalysis": "Başlat",
        "showRiskyOnly": "Sadece Riskli Kayıtlar",
        "searchOutput": "Çıktıda Ara",
        "currency": "Döviz Türü",
        "userName": "Kullanıcı Adı",
        "vendorCode": "Satıcı",
        "vendorName": "Satıcı Adı",
        "ebaStatus": "EBA Durum",
        "all": "Tümü",
        "noData": "Analiz verisi yok",
        "readyNoData": "Hazır veri yok",
        "vendors": "Satıcı",
        "period": "Dönem",
        "riskyRecords": "Riskli Kayıt",
        "ebaRecords": "EBA Kayıt",
        "filteredUniqueVendors": "Filtreli benzersiz satıcı",
        "previous": "Önceki",
        "averageScore": "Ortalama skor",
        "currentPeriodEbaTotal": "Son dönem EBA adet toplamı",
        "regularFtDesc": "FAGGL, EBA ve ZFI052 raporlarını birlikte analiz ederek satıcı bazında dönemsel fatura adet/tutar, son dönem EBA durumu ve ZFI052 muhatap bilgisini üretir.",
        "info": "Bilgi",
        "warning": "Uyarı",
        "error": "Hata",
        "fileCreated": "Dosya '{filename}' olarak oluşturuldu.",
        "noExportData": "Dışa aktarılacak veri yok.",
    },
    "en": {
        "appTitle": "TB & PL-CC Control",
        "appDesc": "Premium control interface that manages TB and PL-CC analyses, period changes, note matching, and VUK direction control on a single screen.",
        "start": "Start Analysis",
        "loadedTb": "Loaded TB",
        "loadedPlcc": "Loaded PL-CC",
        "definedNotes": "Defined Notes",
        "findings": "Findings",
        "panel": "Control Panel",
        "panelDesc": "Program flow, file management, and quick actions",
        "uploadTb": "Upload TB",
        "uploadPlcc": "Upload PL-CC",
        "noteDefinitions": "Note Definitions",
        "responsibleAssignments": "Responsible Assignment",
        "muavinAnalysis": "Subledger Analysis",
        "processingStatus": "Processing Status",
        "processingText": "After loading TB and PL-CC files, run calculations with Start Analysis.",
        "dashboard": "Dashboard",
        "analysis": "Analysis",
        "dashboardTitle": "File and Period Management",
        "dashboardDesc": "Manage loaded files and comparison periods from this screen.",
        "tbFile": "TB File",
        "plccFile": "PL-CC File",
        "currentPeriod": "Current Period",
        "previousPeriod": "Comparison Period",
        "tbReady": "TB File Loaded",
        "plccReady": "PL-CC File Loaded",
        "excelStyle": "Excel header style: Premium Blue / White",
        "analysisTitle": "Analysis Screen",
        "analysisDesc": "You can perform detailed analysis on the same screen by switching between TB and PL-CC.",
        "searchTb": "Search in TB",
        "searchPlcc": "Search in PL-CC",
        "periodCompare": "Period comparison",
        "rowDensity": "Row Spacing",
        "compact": "Compact",
        "normal": "Normal",
        "wide": "Wide",
        "exportExcel": "Excel Output",
        "autoFitColumns": "Auto Fit Columns",
        "toggleInfo": "Use all/findings filter to manage the table.",
        "showAll": "Show All",
        "findingsOnly": "Findings Only",
        "allAccounts": "All Accounts",
        "balanceSheet": "Balance Sheet",
        "incomeStatement": "Income Statement",
        "tb": "TB",
        "plcc": "PL-CC",
        "account": "Account",
        "accountName": "Account Name",
        "note": "Note",
        "responsible": "Responsible",
        "costCenter": "Cost Center",
        "description": "Description",
        "subtotal": "Subtotal",
        "noteDefTitle": "Note Definition Screen",
        "noteDefDesc": "Account is required, cost center is optional. Defined notes are automatically matched to related reports.",
        "accountCode": "Account Code",
        "optionalCostCenter": "Cost Center (Optional)",
        "notePlaceholder": "Explanation / warning note to display for this account",
        "save": "Save",
        "clear": "Clear",
        "matchInfo": "If multiple rules match, notes are joined with |.",
        "savedNotes": "Saved Notes",
        "savedNotesDesc": "List of explanations and warnings managed within the program",
        "responsibleAssignTitle": "Responsible Assignment Screen",
        "responsibleAssignDesc": "Paste two Excel columns as Account and Responsible. The program saves them in bulk and matches them automatically on TB / PL-CC side.",
        "pasteArea": "Excel Paste Area",
        "pasteResponsiblePlaceholder": "Account\tResponsible\n100\tAyşe Yılmaz\n320\tMehmet Kaya",
        "bulkSave": "Save in Bulk",
        "clearAll": "Clear",
        "savedResponsibles": "Saved Responsibles",
        "savedResponsiblesDesc": "Responsible list assigned by account",
        "pasteFormatInfo": "Format: You can paste directly from Excel with Account and Responsible on each row.",
        "notesControlInfo": "Row density can be changed.",
        "general": "General",
        "findingsTb": "TB Findings",
        "findingsPlcc": "PL-CC Findings",
        "compareChange": "% Change",
        "warningCount": "reverse-direction warnings",
        "normalState": "Normal",
        "reverseState": "Reverse direction",
        "status": "Status",
        "premiumView": "Premium view enabled",
        "analysisReady": "Analysis ready",
        "notLoadedTb": "TB not loaded",
        "notLoadedPlcc": "PL-CC not loaded",
        "opening": "Opening",
        "rowsLoaded": "rows loaded",
        "total": "Total",
        "muavinAnalysis": "Ledger Analysis",
        "regularFtAnalysis": "Recurring Invoice Analysis",
        "userManagement": "User Management",
        "loadData": "Load",
        "runAnalysis": "Run Analysis",
        "showRiskyOnly": "Show Risky Records Only",
        "searchOutput": "Search Output",
        "currency": "Currency",
        "userName": "User Name",
        "vendorCode": "Vendor Code",
        "vendorName": "Vendor Name",
        "ebaStatus": "EBA Status",
        "all": "All",
        "noData": "No analysis data",
        "readyNoData": "No data ready",
        "vendors": "Vendors",
        "period": "Period",
        "riskyRecords": "Risky Records",
        "ebaRecords": "EBA Records",
        "filteredUniqueVendors": "Filtered unique vendors",
        "previous": "Previous",
        "averageScore": "Average score",
        "currentPeriodEbaTotal": "Current-period total EBA count",
        "regularFtDesc": "Analyzes FAGGL, EBA, and ZFI052 reports together and produces vendor-based periodic invoice count/amount, current-period EBA status, and ZFI052 responsible information.",
        "info": "Info",
        "warning": "Warning",
        "error": "Error",
        "fileCreated": "File '{filename}' has been created.",
        "noExportData": "There is no data to export.",
    },
}


def ensure_base_dir():
    os.makedirs(BASE_DIR, exist_ok=True)


def load_notes() -> List[Dict]:
    ensure_base_dir()
    if not os.path.exists(NOTES_JSON):
        return []
    try:
        with open(NOTES_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_notes(notes: List[Dict]):
    ensure_base_dir()
    with open(NOTES_JSON, "w", encoding="utf-8") as f:
        json.dump(notes, f, ensure_ascii=False, indent=2)




def load_responsibles() -> List[Dict]:
    ensure_base_dir()
    if not os.path.exists(RESPONSIBLES_JSON):
        return []
    try:
        with open(RESPONSIBLES_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_responsibles(items: List[Dict]):
    ensure_base_dir()
    with open(RESPONSIBLES_JSON, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def hash_password(value: str) -> str:
    return hashlib.sha256(str(value).encode("utf-8")).hexdigest()


def normalize_user_permissions(permissions: Optional[Dict]) -> Dict[str, bool]:
    base = dict(DEFAULT_SECTION_PERMISSIONS)
    if isinstance(permissions, dict):
        for key in base.keys():
            if key in permissions:
                base[key] = bool(permissions.get(key))
    return base


def load_users() -> List[Dict]:
    ensure_base_dir()
    default_admin = [{"username": "admin", "password_hash": hash_password("admin"), "is_admin": True, "permissions": normalize_user_permissions(None)}]
    if not os.path.exists(USERS_JSON):
        with open(USERS_JSON, "w", encoding="utf-8") as f:
            json.dump(default_admin, f, ensure_ascii=False, indent=2)
        return default_admin
    try:
        with open(USERS_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list) or not data:
            raise ValueError("invalid users file")
        normalized = []
        for item in data:
            username = str(item.get("username", "")).strip()
            if not username:
                continue
            password_hash = str(item.get("password_hash", "")).strip()
            if not password_hash and item.get("password"):
                password_hash = hash_password(item.get("password"))
            normalized.append({
                "username": username,
                "password_hash": password_hash,
                "is_admin": bool(item.get("is_admin", username.lower() == "admin")),
                "permissions": normalize_user_permissions(item.get("permissions")),
            })
        if not any(x["username"].lower() == "admin" for x in normalized):
            normalized.insert(0, {"username": "admin", "password_hash": hash_password("admin"), "is_admin": True, "permissions": normalize_user_permissions(None)})
        save_users(normalized)
        return normalized
    except Exception:
        with open(USERS_JSON, "w", encoding="utf-8") as f:
            json.dump(default_admin, f, ensure_ascii=False, indent=2)
        return default_admin


def save_users(items: List[Dict]):
    ensure_base_dir()
    cleaned = []
    seen = set()
    for item in items:
        username = str(item.get("username", "")).strip()
        if not username:
            continue
        key = username.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append({
            "username": username,
            "password_hash": str(item.get("password_hash", "")).strip(),
            "is_admin": bool(item.get("is_admin", key == "admin")),
            "permissions": normalize_user_permissions(item.get("permissions")),
        })
    if not any(x["username"].lower() == "admin" for x in cleaned):
        cleaned.insert(0, {"username": "admin", "password_hash": hash_password("admin"), "is_admin": True, "permissions": normalize_user_permissions(None)})
    with open(USERS_JSON, "w", encoding="utf-8") as f:
        json.dump(cleaned, f, ensure_ascii=False, indent=2)


def verify_user_credentials(username: str, password: str) -> Optional[Dict]:
    username = str(username).strip()
    password_hash = hash_password(password)
    for user in load_users():
        if str(user.get("username", "")).strip().lower() == username.lower() and str(user.get("password_hash", "")).strip() == password_hash:
            return user
    return None


def join_responsibles(parts: List[str]) -> str:
    cleaned = []
    seen = set()
    for p in parts:
        p = str(p).strip()
        if not p:
            continue
        if p not in seen:
            cleaned.append(p)
            seen.add(p)
    return " | ".join(cleaned)



def collect_matching_responsibles(
    items: List[Dict],
    hesap_value: str,
    ana_hesap_value: str = ""
) -> str:
    hedef_hesap = normalize_hesap_prefix(hesap_value)
    hedef_ana_hesap = str(ana_hesap_value).strip()

    matched = []
    for item in items:
        kayit_hesap = normalize_hesap_prefix(item.get("hesap", ""))
        kayit_ana_hesap = str(item.get("anaHesap", "")).strip()

        if kayit_ana_hesap:
            if kayit_ana_hesap != hedef_ana_hesap:
                continue
        else:
            if kayit_hesap != hedef_hesap:
                continue

        matched.append(str(item.get("sorumlu", "")).strip())

    return join_responsibles(matched)

def normalize_col_name(value) -> str:
    return str(value).strip().lower().replace("ı", "i")


def normalize_hesap_prefix(value: str) -> str:
    s = str(value).strip()
    if not s:
        return ""
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) >= 3:
        return digits[:3]
    return s[:3]


def period_key_to_label(value) -> str:
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""

    try:
        n = int(float(s))
        if n == 0:
            return OPENING_PERIOD_TR
        if 1 <= n <= 12:
            return ALL_PERIODS[n - 1]["tr"]
    except Exception:
        pass

    if s == OPENING_PERIOD_TR or s == OPENING_PERIOD_EN:
        return OPENING_PERIOD_TR

    if s in PERIOD_ORDER:
        return s

    for p in ALL_PERIODS:
        if s == p["en"]:
            return p["tr"]

    return s


def sort_periods(periods: List[str]) -> List[str]:
    uniq = []
    seen = set()
    for p in periods:
        if p and p not in seen:
            uniq.append(p)
            seen.add(p)

    def _sort_key(x: str):
        if x == OPENING_PERIOD_TR:
            return (0, 0)
        return (1, PERIOD_ORDER.get(x, 999))

    return sorted(uniq, key=_sort_key)


def safe_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return 0.0

    s = str(value).strip()
    if not s or s.lower() == "nan":
        return 0.0

    s = s.replace("\xa0", "").replace(" ", "")

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")

    try:
        return float(s)
    except Exception:
        return 0.0


def format_number(value) -> str:
    v = safe_float(value)
    txt = f"{v:,.2f}"
    return txt.replace(",", "X").replace(".", ",").replace("X", ".")


def calc_change_percent(current: float, previous: float) -> str:
    current = safe_float(current)
    previous = safe_float(previous)
    if abs(previous) < 1e-12:
        if abs(current) < 1e-12:
            return "0,00%"
        return "Yeni"
    pct = ((current - previous) / abs(previous)) * 100
    return f"{pct:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")


def parse_percent_text(value: str) -> Optional[float]:
    s = str(value).strip()
    if not s or s == "-" or s.lower() == "yeni":
        return None
    if s.endswith("%"):
        s = s[:-1].strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def is_critical_variance(change_text: str, threshold: float = 10.0) -> bool:
    pct = parse_percent_text(change_text)
    return pct is not None and abs(pct) > threshold


def critical_variance_status(change_text: str, threshold: float = 10.0) -> str:
    return f"Kritik Sapma (>%{int(threshold)})" if is_critical_variance(change_text, threshold) else "Normal"


def get_direction_from_amount(value: float) -> str:
    v = safe_float(value)
    if v > 0:
        return "Borç"
    if v < 0:
        return "Alacak"
    return "-"


def get_control_status(expected: str, actual: str) -> str:
    if not expected or expected == "-" or not actual or actual == "-":
        return "Normal"
    return "Ters duruyor" if expected != actual else "Normal"


def masraf_sort_key(value: str):
    s = str(value or "").strip()
    if s.isdigit():
        return (0, int(s))
    return (1, s.lower())


def classify_financial_statement(hesap_prefix: str) -> str:
    hp = normalize_hesap_prefix(hesap_prefix)
    if not hp:
        return "Diğer"
    first = hp[:1]
    if first in ["1", "2", "3", "4", "5"]:
        return "Bilanço"
    if first in ["6", "7"]:
        return "Gelir Tablosu"
    if first == "8":
        return "Maliyet / Nazım"
    if first == "9":
        return "Nazım"
    return "Diğer"


def parse_muavin_period(value) -> str:
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    s = s.replace(".", "/").replace("-", "/")
    m = re.search(r"(\d{4})\s*/\s*(\d{1,2})", s)
    if m:
        year = m.group(1)
        month = int(m.group(2))
        return f"{year}/{month:02d}"
    m = re.search(r"(\d{4})(\d{2})", s)
    if m:
        year = m.group(1)
        month = int(m.group(2))
        if 1 <= month <= 13:
            return f"{year}/{month:02d}"
    return s


def normalize_text_value(value) -> str:
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    return s

def tokenize_text_for_nlp(value: str) -> List[str]:
    s = normalize_text_value(value).lower()
    if not s:
        return []
    s = re.sub(r"[^a-z0-9çğıöşü\s]", " ", s, flags=re.IGNORECASE)
    parts = [p.strip() for p in s.split() if p.strip()]
    stop_words = {
        "ve", "ile", "icin", "için", "bir", "bu", "da", "de", "the", "for", "to", "is", "are",
        "mi", "mu", "mü", "ya", "veya", "ile", "olarak", "olan", "olanlar", "ait", "gore", "göre",
        "no", "belge", "fatura", "kayit", "kayıt", "hesap", "satir", "satır", "kod", "adi", "adı"
    }
    return [p for p in parts if len(p) > 1 and p not in stop_words]


def build_text_group_nlp(value: str) -> Tuple[str, int]:
    tokens = tokenize_text_for_nlp(value)
    if not tokens:
        raw = normalize_text_value(value)
        return (raw[:80] if raw else "-", 0)

    top = Counter(tokens).most_common(3)
    grup = " ".join([str(t[0]) for t in top]).strip()
    skor = int(sum(int(t[1]) for t in top))
    return (grup if grup else "-", skor)


def classify_muavin_text_pattern(value) -> str:
    s = normalize_text_value(value).lower()
    if not s:
        return "-"
    rules = [
        ("odeme", "Ödeme Kaydı"),
        ("ödeme", "Ödeme Kaydı"),
        ("virman", "Virman"),
        ("mahsup", "Mahsup"),
        ("fatura", "Fatura"),
        ("iade", "İade"),
        ("kur fark", "Kur Farkı"),
        ("kira", "Kira"),
        ("banka", "Banka"),
        ("personel", "Personel"),
    ]
    hits=[]
    for key,label in rules:
        if key in s and label not in hits:
            hits.append(label)
    return " / ".join(hits) if hits else (normalize_text_value(value)[:80] or "-")


def build_late_7day_note(document_date, posting_date, created_date):
    belge_dt = parse_date_flexible(document_date)
    kayit_dt = parse_date_flexible(posting_date)
    giris_dt = parse_date_flexible(created_date)
    if pd.isna(giris_dt):
        return False, "", 0, pd.NaT, pd.NaT
    notes=[]
    delays=[]
    belge_deadline = next_month_seventh(belge_dt) if not pd.isna(belge_dt) else pd.NaT
    kayit_deadline = next_month_seventh(kayit_dt) if not pd.isna(kayit_dt) else pd.NaT
    giris_day = giris_dt.normalize() if hasattr(giris_dt, "normalize") else pd.Timestamp(giris_dt).normalize()
    if not pd.isna(belge_deadline) and giris_day > belge_deadline:
        d = int((giris_day - belge_deadline).days)
        delays.append(d)
        notes.append(f"Belge tarihine göre 7 gün aşıldı (+{d} gün)")
    if not pd.isna(kayit_deadline) and giris_day > kayit_deadline:
        d = int((giris_day - kayit_deadline).days)
        delays.append(d)
        notes.append(f"Kayıt tarihine göre 7 gün aşıldı (+{d} gün)")
    return bool(notes), " | ".join(notes), max(delays) if delays else 0, belge_deadline, kayit_deadline


def safe_unique_join(values, sep=" | ", limit: int = 12) -> str:
    cleaned = []
    seen = set()
    for v in values:
        s = normalize_text_value(v)
        if not s:
            continue
        if s not in seen:
            cleaned.append(s)
            seen.add(s)
    if len(cleaned) > limit:
        return sep.join(cleaned[:limit]) + f" (+{len(cleaned)-limit})"
    return sep.join(cleaned)


def text_search_match(value: str, query: str, mode: str = "contains") -> bool:
    source = normalize_text_value(value).lower().strip()
    q = normalize_text_value(query).lower().strip()
    if not q:
        return True
    if mode == "exact":
        return source == q
    return q in source


def row_text_search_match(values, query: str, mode: str = "contains") -> bool:
    q = normalize_text_value(query).lower().strip()
    if not q:
        return True
    normalized_values = [normalize_text_value(v).lower().strip() for v in values]
    if mode == "exact":
        return any(v == q for v in normalized_values if v)
    return q in " ".join(normalized_values)


def df_text_search_mask(df: pd.DataFrame, columns: List[str], query: str, mode: str = "contains") -> pd.Series:
    q = normalize_text_value(query).lower().strip()
    if df is None or df.empty:
        return pd.Series(dtype=bool)
    if not q:
        return pd.Series(True, index=df.index)
    mask = pd.Series(False, index=df.index)
    for col in columns:
        if col not in df.columns:
            continue
        series = df[col].astype(str).str.lower().str.strip().fillna("")
        if mode == "exact":
            mask = mask | (series == q)
        else:
            mask = mask | series.str.contains(re.escape(q), na=False)
    return mask

def trend_alarm_text(current: float, previous: float, threshold: float = 20.0) -> str:
    if abs(previous) < 1e-12:
        if abs(current) < 1e-12:
            return "Normal"
        return "Yeni Hareket"
    pct = ((current - previous) / abs(previous)) * 100.0
    if pct >= threshold:
        return f"Artış Alarmı (%{pct:,.1f})".replace(",", "X").replace(".", ",").replace("X", ".")
    if pct <= -threshold:
        return f"Azalış Alarmı (%{pct:,.1f})".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"Normal (%{pct:,.1f})".replace(",", "X").replace(".", ",").replace("X", ".")

def safe_zscore(value: float, mean: float, std: float) -> float:
    if std <= 1e-12:
        return 0.0
    return (value - mean) / std

def score_band(score: float) -> str:
    if score >= 80:
        return "Çok Yüksek"
    if score >= 60:
        return "Yüksek"
    if score >= 40:
        return "Orta"
    return "Düşük"

def risk_badge_from_text(value: str) -> Optional[str]:
    s = str(value or "").lower()
    if any(x in s for x in ["çok yüksek", "kritik", "alarm", "incele", "anormallik", "riskli", "eşleşmedi", "net bakiye var", "açık belge"]):
        return "danger"
    if any(x in s for x in ["yüksek", "orta", "yeni hareket", "kontrol", "uyarı"]):
        return "warn"
    if any(x in s for x in ["düşük", "normal", "netleşmiş", "tamam"]):
        return "success"
    return None

def numeric_risk_badge(value) -> Optional[str]:
    try:
        v = float(str(value).replace(".", "").replace(",", "."))
    except Exception:
        return None
    if v >= 80:
        return "danger"
    if v >= 40:
        return "warn"
    return "success"

def make_muavin_findings(df: pd.DataFrame) -> List[str]:
    findings = []
    if df is None or df.empty:
        return ["Seçili filtrelerde muavin verisi bulunamadı."]

    findings.extend([
        "Risk Skoru Rehberi: 0-19 Düşük, 20-39 Orta, 40-59 Yüksek, 60+ Kritik.",
        "Skor bileşenleri: mükerrer kayıt, 7 gün sonrası kayıt, belge ilişkisel test, vergi riski, önemlilik, kullanıcı paterni ve trend sapması.",
    ])

    total_rows = len(df)
    belge_sayisi = int(df["belge_anahtar"].replace("", pd.NA).dropna().nunique())
    ters_rows = int(df["has_ters_kayit"].sum())
    denkl_oran = float(df["has_denklestirme"].mean() * 100.0) if total_rows else 0.0
    findings.append(f"Seçili hesapta {total_rows} satır ve {belge_sayisi} benzersiz belge bulunmaktadır.")
    findings.append(f"Denkleştirme oranı %{denkl_oran:,.1f} seviyesindedir.".replace(",", "X").replace(".", ",").replace("X", "."))
    if ters_rows:
        ters_tutar = float(df.loc[df["has_ters_kayit"], "up_tutar"].sum())
        findings.append(f"Ters kayıt içeren {ters_rows} satır vardır; toplam ters kayıt tutarı {format_number(ters_tutar)} TL seviyesindedir.")
    else:
        findings.append("Seçili kapsamda ters kayıt içeren satır bulunmamaktadır.")

    if "risk_skoru" in df.columns:
        top_user = df.groupby("kullanici")["risk_skoru"].mean().sort_values(ascending=False)
        top_user = top_user[top_user.index != ""]
        if not top_user.empty:
            findings.append(f"En yüksek ortalama risk skoru kullanıcı bazında {top_user.index[0]} üzerinde yoğunlaşmaktadır.")

    if "anomali_skoru" in df.columns:
        top_contra = df.groupby(["karsi_hesap", "karsi_hesap_tanimi"])["anomali_skoru"].mean().sort_values(ascending=False)
        if len(top_contra):
            (kodu, tanim), skor = top_contra.index[0], float(top_contra.iloc[0])
            acik = f"{kodu} - {tanim}".strip(" -")
            findings.append(f"Satıcı/cari tarafında en yüksek anormallik göstergesi {acik} için {skor:,.1f} puandır.".replace(",", "X").replace(".", ",").replace("X", "."))

    if "cost_change_alarm" in df.columns:
        alarm_df = df[df["cost_change_alarm"].astype(str).str.contains("Alarm|Yeni Hareket", na=False)]
        if not alarm_df.empty:
            top_cost = alarm_df.groupby(["masraf_yeri", "masraf_yeri_tanimi"]).size().sort_values(ascending=False)
            (cc, cc_text) = top_cost.index[0]
            findings.append(f"Masraf yeri tarafında hareket alarmı en belirgin şekilde {cc} {cc_text} üzerinde görülmektedir.".strip())

    top_docs = df.groupby("belge_turu").size().sort_values(ascending=False)
    if len(top_docs):
        findings.append(f"Belge yoğunluğu en çok {top_docs.index[0]} belge türünde toplanmaktadır.")

    return findings

def read_excel_flexible(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    engine = "openpyxl" if ext == ".xlsx" else None
    try:
        return pd.read_excel(path, engine=engine)
    except Exception:
        return pd.read_excel(path)


def read_excel_headers_only(path: str) -> List[str]:
    ext = os.path.splitext(path)[1].lower()
    engine = "openpyxl" if ext == ".xlsx" else None
    try:
        df = pd.read_excel(path, engine=engine, nrows=0)
    except Exception:
        df = pd.read_excel(path, nrows=0)
    return [str(c).strip() for c in df.columns]


def read_excel_selected_columns(path: str, usecols: Optional[List[str]] = None) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    engine = "openpyxl" if ext == ".xlsx" else None
    try:
        return pd.read_excel(path, engine=engine, usecols=usecols)
    except Exception:
        return pd.read_excel(path, usecols=usecols)






def load_single_excel_payload(path: str, file_label: str = "", language: str = "tr", progress_cb=None) -> Dict:
    label = file_label or os.path.basename(path)
    _safe_progress(progress_cb, 10, (f"{label} okunuyor..." if language == "tr" else f"Reading {label}..."))
    df = read_excel_flexible(path)
    _safe_progress(progress_cb, 90, (f"{label} hazırlandı." if language == "tr" else f"{label} is ready."))
    return {"path": path, "df": df, "file_label": label}


MAX_EXPORT_ROWS_PER_FILE = 300000


def split_export_path(path: str, part_no: int) -> str:
    base, ext = os.path.splitext(path)
    ext = ext or ".xlsx"
    return f"{base}_part-{part_no}{ext}"


def _normalize_sheet_payload_df(sheet: Dict) -> pd.DataFrame:
    df = sheet.get("df", pd.DataFrame())
    if not isinstance(df, pd.DataFrame):
        df = pd.DataFrame(df)
    return df.copy()


def write_single_sheet_excel(path: str, headers: List[str], rows: List[List[str]], sheet_name: str = "Rapor", progress_cb=None) -> Dict:
    _safe_progress(progress_cb, 15, "Excel yazımı başlatılıyor...")
    df = pd.DataFrame(rows, columns=headers)
    row_count = len(df.index)
    if row_count <= MAX_EXPORT_ROWS_PER_FILE:
        _safe_progress(progress_cb, 60, "Sayfa oluşturuluyor...")
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=str(sheet_name)[:31] or "Rapor", index=False)
        _safe_progress(progress_cb, 100, "Excel yazımı tamamlandı.")
        return {"path": path, "paths": [path], "part_count": 1, "row_count": row_count}

    part_count = int(math.ceil(row_count / MAX_EXPORT_ROWS_PER_FILE))
    paths = []
    for part_no in range(1, part_count + 1):
        start_idx = (part_no - 1) * MAX_EXPORT_ROWS_PER_FILE
        end_idx = min(part_no * MAX_EXPORT_ROWS_PER_FILE, row_count)
        part_df = df.iloc[start_idx:end_idx].copy()
        part_path = split_export_path(path, part_no)
        with pd.ExcelWriter(part_path, engine="openpyxl") as writer:
            part_df.to_excel(writer, sheet_name=str(sheet_name)[:31] or "Rapor", index=False)
        paths.append(part_path)
        _safe_progress(progress_cb, min(98, 15 + int((part_no / part_count) * 80)), f"Excel parça {part_no}/{part_count} hazırlanıyor...")
    _safe_progress(progress_cb, 100, "Excel parçalı yazımı tamamlandı.")
    return {"path": paths[0], "paths": paths, "part_count": part_count, "row_count": row_count}


def write_multi_sheet_excel(path: str, sheets: List[Dict], progress_cb=None) -> Dict:
    _safe_progress(progress_cb, 10, "Excel çoklu sayfa yazımı başlatılıyor...")
    normalized_sheets = []
    oversized_max_rows = 0
    for idx, sheet in enumerate(sheets, start=1):
        name = str(sheet.get("sheet_name", f"Sheet{idx}"))[:31] or f"Sheet{idx}"
        df = _normalize_sheet_payload_df(sheet)
        normalized_sheets.append({"sheet_name": name, "df": df})
        oversized_max_rows = max(oversized_max_rows, len(df.index))

    if oversized_max_rows <= MAX_EXPORT_ROWS_PER_FILE:
        total = max(len(normalized_sheets), 1)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for idx, sheet in enumerate(normalized_sheets, start=1):
                sheet["df"].to_excel(writer, sheet_name=sheet["sheet_name"], index=False)
                _safe_progress(progress_cb, min(95, 10 + int((idx / total) * 80)), f"{sheet['sheet_name']} yazılıyor...")
        _safe_progress(progress_cb, 100, "Excel çoklu sayfa yazımı tamamlandı.")
        return {"path": path, "paths": [path], "part_count": 1, "row_count": oversized_max_rows}

    part_count = int(math.ceil(oversized_max_rows / MAX_EXPORT_ROWS_PER_FILE))
    paths = []
    for part_no in range(1, part_count + 1):
        part_path = split_export_path(path, part_no)
        start_idx = (part_no - 1) * MAX_EXPORT_ROWS_PER_FILE
        end_idx = part_no * MAX_EXPORT_ROWS_PER_FILE
        with pd.ExcelWriter(part_path, engine="openpyxl") as writer:
            for idx, sheet in enumerate(normalized_sheets, start=1):
                df = sheet["df"]
                if len(df.index) > MAX_EXPORT_ROWS_PER_FILE:
                    export_df = df.iloc[start_idx:end_idx].copy()
                else:
                    export_df = df.copy()
                export_df.to_excel(writer, sheet_name=sheet["sheet_name"], index=False)
        paths.append(part_path)
        _safe_progress(progress_cb, min(98, 10 + int((part_no / part_count) * 85)), f"Excel parça {part_no}/{part_count} hazırlanıyor...")
    _safe_progress(progress_cb, 100, "Excel çoklu sayfa parçalı yazımı tamamlandı.")
    return {"path": paths[0], "paths": paths, "part_count": part_count, "row_count": oversized_max_rows}


def write_financial_statement_excel_payload(path: str, title: str, headers: List[str], rows: List[List[str]], subtotal_row_indexes=None, progress_cb=None) -> Dict:
    _safe_progress(progress_cb, 20, "Finansal tablo Excel çıktısı hazırlanıyor...")
    export_financial_statement_excel(path, title, headers, rows, subtotal_row_indexes or [])
    _safe_progress(progress_cb, 100, "Finansal tablo Excel çıktısı tamamlandı.")
    return {"path": path}

class BackgroundWorker(QThread):
    progress = Signal(int, str)
    succeeded = Signal(object)
    failed = Signal(str)

    def __init__(self, fn, *args, **kwargs):
        super().__init__()
        self.fn = fn
        self.args = args
        self.kwargs = kwargs

    def run(self):
        try:
            result = self.fn(*self.args, progress_cb=self._emit_progress, **self.kwargs)
            self.succeeded.emit(result)
        except Exception as e:
            self.failed.emit(str(e))

    def _emit_progress(self, value: int, message: str):
        self.progress.emit(int(value), str(message))


def _safe_progress(progress_cb, value: int, message: str):
    if progress_cb is not None:
        progress_cb(int(value), str(message))


def load_muavin_headers_payload(paths: List[str], language: str = "tr", progress_cb=None) -> Dict:
    all_cols = []
    total = max(len(paths), 1)
    for idx, p in enumerate(paths, start=1):
        msg = (
            f"Muavin başlıkları okunuyor... ({idx}/{total})"
            if language == "tr" else
            f"Reading subledger headers... ({idx}/{total})"
        )
        _safe_progress(progress_cb, min(45, 8 + int((idx / total) * 30)), msg)
        all_cols.extend(read_excel_headers_only(p))
    available_columns = list(dict.fromkeys([str(c).strip() for c in all_cols if str(c).strip()]))
    if len(paths) == 1:
        display_path = paths[0]
    else:
        basenames = [os.path.basename(p) for p in paths[:3]]
        tail = f" +{len(paths) - 3}" if len(paths) > 3 else ""
        display_path = " | ".join(basenames) + tail
    return {
        "file_paths": list(paths),
        "available_columns": available_columns,
        "display_path": display_path,
    }


def build_muavin_analysis_payload_from_files(
    file_paths: List[str],
    mapping: Dict[str, str],
    language: str = "tr",
    progress_cb=None,
) -> Dict:
    selected_cols = [str(v).strip() for v in mapping.values() if str(v).strip()]
    selected_cols = list(dict.fromkeys(selected_cols))
    loaded_frames = []
    total = max(len(file_paths), 1)
    for idx, file_path in enumerate(file_paths, start=1):
        msg = (
            f"Muavin kaynak dosyaları yükleniyor... ({idx}/{total})"
            if language == "tr" else
            f"Loading subledger source files... ({idx}/{total})"
        )
        _safe_progress(progress_cb, min(35, 8 + int((idx / total) * 20)), msg)
        df_part = read_excel_selected_columns(file_path, usecols=selected_cols if selected_cols else None)
        df_part["_source_file"] = os.path.basename(file_path)
        loaded_frames.append(df_part)
    raw_df = pd.concat(loaded_frames, ignore_index=True) if loaded_frames else pd.DataFrame()
    return build_muavin_analysis_payload(raw_df=raw_df, mapping=mapping, language=language, progress_cb=progress_cb)


def build_muavin_analysis_payload(raw_df: pd.DataFrame, mapping: Dict[str, str], language: str = "tr", progress_cb=None) -> Dict:
    if raw_df is None or raw_df.empty:
        return {
            "raw_df": pd.DataFrame(),
            "clean_df": pd.DataFrame(),
            "accounts": [],
            "periods": [],
            "doc_types": [],
            "users": [],
            "contras": [],
            "cost_centers": [],
        }

    _safe_progress(progress_cb, 12, "Muavin verisi kopyalanıyor ve kolonlar hazırlanıyor..." if language == "tr" else "Copying subledger data and preparing columns...")
    df = raw_df.copy()
    rename_map = {}
    wanted = {k: v for k, v in mapping.items() if v}
    required_labels = {
        "yilay": "Yıl/ay" if language == "tr" else "Year/Month",
        "ana_hesap": "Ana hesap" if language == "tr" else "Main account",
        "ana_hesap_adi": "DK hesabı uzun metni" if language == "tr" else "G/L account long text",
    }
    missing = [required_labels[k] for k in ["yilay", "ana_hesap", "ana_hesap_adi"] if not wanted.get(k)]
    if missing:
        raise ValueError(("Muavin dosyasında bulunamayan kolonlar: " if language == "tr" else "Missing required subledger columns: ") + ", ".join(missing))
    for new, old in wanted.items():
        if old and str(old).strip() in df.columns:
            rename_map[str(old).strip()] = new
    df = df.rename(columns=rename_map)

    _safe_progress(progress_cb, 22, "Muavin kolonları normalize ediliyor..." if language == "tr" else "Normalizing subledger columns...")
    for col in ["ana_hesap","ana_hesap_adi","referans","belge_numarasi","belge_turu","karsi_hesap_tanimi","islem_kodu","belge_pb","karsi_hesap","denklestirme","metin","ters_kayit","kullanici","vergi_gostergesi","masraf_yeri","masraf_yeri_tanimi"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].map(normalize_text_value)
    for col in ["up_tutar","belge_pb_tutar"]:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = df[col].map(safe_float)

    _safe_progress(progress_cb, 32, "Temel muavin göstergeleri hazırlanıyor..." if language == "tr" else "Preparing core subledger indicators...")
    df["donem"] = df["yilay"].map(parse_muavin_period)
    df["hesap_prefix"] = df["ana_hesap"].map(normalize_hesap_prefix)
    df["finansal_tablo"] = df["hesap_prefix"].map(classify_financial_statement)
    df["has_denklestirme"] = df["denklestirme"].astype(str).str.strip() != ""
    df["has_ters_kayit"] = df["ters_kayit"].astype(str).str.strip() != ""
    df["belge_anahtar"] = df.apply(lambda x: normalize_text_value(x["belge_numarasi"]) or normalize_text_value(x["referans"]), axis=1)

    df = df[(df["ana_hesap"].astype(str).str.strip() != "") | (df["ana_hesap_adi"].astype(str).str.strip() != "")]
    df = df[df["donem"].astype(str).str.strip() != ""].copy()
    df = df[~((df["belge_anahtar"] == "") & (df["belge_turu"] == "") & (df["karsi_hesap"] == "") & (df["karsi_hesap_tanimi"] == "") & (df["metin"] == "") & (df["up_tutar"].abs() < 1e-12))].copy()

    df["duplicate_ref_contra_key"] = df.apply(
        lambda x: (
            f"{normalize_text_value(x['referans'])}||{normalize_text_value(x['karsi_hesap']) or normalize_text_value(x['karsi_hesap_tanimi'])}"
            if normalize_text_value(x["referans"])
            else ""
        ),
        axis=1
    )

    open_ref_df = df[
        (df["referans"].astype(str).str.strip() != "")
        & (~df["has_ters_kayit"])
        & (
            (df["karsi_hesap"].astype(str).str.strip() != "")
            | (df["karsi_hesap_tanimi"].astype(str).str.strip() != "")
        )
    ].copy()
    if not open_ref_df.empty:
        ref_contra_doc_counts = open_ref_df.groupby("duplicate_ref_contra_key")["belge_anahtar"].nunique().to_dict()
        df["duplicate_open_reference"] = df["duplicate_ref_contra_key"].map(
            lambda x: ref_contra_doc_counts.get(str(x).strip(), 0) > 1 if str(x).strip() else False
        )
        df["duplicate_open_reference_count"] = df["duplicate_ref_contra_key"].map(
            lambda x: int(ref_contra_doc_counts.get(str(x).strip(), 0)) if str(x).strip() else 0
        )
    else:
        df["duplicate_open_reference"] = False
        df["duplicate_open_reference_count"] = 0

    _safe_progress(progress_cb, 45, "Kullanıcı risk skorları hesaplanıyor..." if language == "tr" else "Calculating user risk scores...")
    grp_ud = df.groupby(["belge_turu", "kullanici"]).agg(
        satir=("ana_hesap", "size"),
        belge_sayisi=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
        tutar=("up_tutar", lambda s: float(s.abs().sum())),
        ters_orani=("has_ters_kayit", "mean"),
        denkl_orani=("has_denklestirme", "mean"),
        cari_sayisi=("karsi_hesap", lambda s: s.replace("", pd.NA).dropna().nunique()),
    ).reset_index()
    if not grp_ud.empty:
        grp_ud["risk_skoru"] = grp_ud["satir"] * 0.8 + grp_ud["belge_sayisi"] * 1.2 + grp_ud["tutar"].rank(pct=True) * 35 + grp_ud["ters_orani"] * 30 + (1 - grp_ud["denkl_orani"]) * 10 + grp_ud["cari_sayisi"].rank(pct=True) * 15
        risk_map = {(r["belge_turu"], r["kullanici"]): float(r["risk_skoru"]) for _, r in grp_ud.iterrows()}
        df["risk_skoru"] = df.apply(lambda x: risk_map.get((x["belge_turu"], x["kullanici"]), 0.0), axis=1)
    else:
        df["risk_skoru"] = 0.0

    _safe_progress(progress_cb, 58, "Satıcı / cari yoğunluk ve anormallik skorları hesaplanıyor..." if language == "tr" else "Calculating vendor/account concentration and anomaly scores...")
    grp_contra = df.groupby(["karsi_hesap", "karsi_hesap_tanimi"]).agg(
        satir=("ana_hesap", "size"),
        belge_sayisi=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
        tutar=("up_tutar", lambda s: float(s.abs().sum())),
        ters_orani=("has_ters_kayit", "mean"),
        vergi_sayisi=("vergi_gostergesi", lambda s: s.replace("", pd.NA).dropna().nunique()),
        belge_turu_sayisi=("belge_turu", lambda s: s.replace("", pd.NA).dropna().nunique()),
    ).reset_index()
    if not grp_contra.empty:
        mt, st = grp_contra["tutar"].mean(), grp_contra["tutar"].std(ddof=0)
        ms, ss = grp_contra["satir"].mean(), grp_contra["satir"].std(ddof=0)
        grp_contra["yogunluk_skoru"] = grp_contra["tutar"].rank(pct=True) * 50 + grp_contra["satir"].rank(pct=True) * 30 + grp_contra["belge_sayisi"].rank(pct=True) * 20
        grp_contra["anomali_skoru"] = grp_contra.apply(lambda r: max(0.0, safe_zscore(r["tutar"], mt, st)) * 30 + max(0.0, safe_zscore(r["satir"], ms, ss)) * 20 + r["ters_orani"] * 30 + (1 if r["vergi_sayisi"] >= 3 else 0) * 10 + (1 if r["belge_turu_sayisi"] >= 4 else 0) * 10, axis=1)
        contra_map = {(r["karsi_hesap"], r["karsi_hesap_tanimi"]): (float(r["yogunluk_skoru"]), float(r["anomali_skoru"])) for _, r in grp_contra.iterrows()}
        df["yogunluk_skoru"] = df.apply(lambda x: contra_map.get((x["karsi_hesap"], x["karsi_hesap_tanimi"]), (0.0, 0.0))[0], axis=1)
        df["anomali_skoru"] = df.apply(lambda x: contra_map.get((x["karsi_hesap"], x["karsi_hesap_tanimi"]), (0.0, 0.0))[1], axis=1)
    else:
        df["yogunluk_skoru"] = 0.0
        df["anomali_skoru"] = 0.0

    _safe_progress(progress_cb, 68, "Masraf yeri trendleri hazırlanıyor..." if language == "tr" else "Preparing cost center trends...")
    cc = df.groupby(["masraf_yeri", "masraf_yeri_tanimi", "donem"])["up_tutar"].sum().reset_index()
    if not cc.empty:
        cc = cc.sort_values(["masraf_yeri", "donem"])
        cc["prev_tutar"] = cc.groupby(["masraf_yeri", "masraf_yeri_tanimi"])["up_tutar"].shift(1).fillna(0.0)
        cc["degisim_yuzde"] = cc.apply(lambda r: 0.0 if abs(float(r["prev_tutar"])) < 1e-12 and abs(float(r["up_tutar"])) < 1e-12 else (999.0 if abs(float(r["prev_tutar"])) < 1e-12 else ((float(r["up_tutar"]) - float(r["prev_tutar"])) / abs(float(r["prev_tutar"])) * 100.0)), axis=1)
        cc["cost_change_alarm"] = cc.apply(lambda r: trend_alarm_text(float(r["up_tutar"]), float(r["prev_tutar"])), axis=1)
        cc_map = {(r["masraf_yeri"], r["masraf_yeri_tanimi"], r["donem"]): (float(r["prev_tutar"]), float(r["degisim_yuzde"]), r["cost_change_alarm"]) for _, r in cc.iterrows()}
        df["prev_cost_tutar"] = df.apply(lambda x: cc_map.get((x["masraf_yeri"], x["masraf_yeri_tanimi"], x["donem"]), (0.0, 0.0, "Normal"))[0], axis=1)
        df["cost_degisim_yuzde"] = df.apply(lambda x: cc_map.get((x["masraf_yeri"], x["masraf_yeri_tanimi"], x["donem"]), (0.0, 0.0, "Normal"))[1], axis=1)
        df["cost_change_alarm"] = df.apply(lambda x: cc_map.get((x["masraf_yeri"], x["masraf_yeri_tanimi"], x["donem"]), (0.0, 0.0, "Normal"))[2], axis=1)
    else:
        df["prev_cost_tutar"] = 0.0
        df["cost_degisim_yuzde"] = 0.0
        df["cost_change_alarm"] = "Normal"

    _safe_progress(progress_cb, 78, "Belge ilişkisel kontrolleri hazırlanıyor..." if language == "tr" else "Preparing document relation controls...")
    doc_grp = df.groupby("belge_anahtar").agg(
        doc_satir=("ana_hesap", "size"),
        doc_net_tutar=("up_tutar", "sum"),
        doc_abs_tutar=("up_tutar", lambda s: float(s.abs().sum())),
        doc_unique_accounts=("ana_hesap", lambda s: s.replace("", pd.NA).dropna().nunique()),
        doc_has_denkl=("has_denklestirme", "max"),
        doc_has_ters=("has_ters_kayit", "max"),
        doc_duplicate_ref=("duplicate_open_reference", "max"),
        doc_duplicate_ref_count=("duplicate_open_reference_count", "max"),
    ).reset_index()

    def _doc_status(r):
        anahtar = normalize_text_value(r["belge_anahtar"])
        if not anahtar:
            return "Belge anahtarı yok" if language == "tr" else "No document key"
        if r.get("doc_duplicate_ref"):
            return (f"Ters kayıt boşken referans + karşıt hesap mükerrer ({int(r.get('doc_duplicate_ref_count', 0))} belge)" if language == "tr"
                    else f"Duplicate reference + contra account while reversal is empty ({int(r.get('doc_duplicate_ref_count', 0))} docs)")
        if r["doc_has_denkl"] and r["doc_has_ters"]:
            return "Denkleştirme + ters kayıt birlikte" if language == "tr" else "Clearing + reversal together"
        if abs(float(r["doc_net_tutar"])) <= 1.0 and int(r["doc_unique_accounts"]) >= 2:
            return "Belge netleşmiş" if language == "tr" else "Document netted"
        if r["doc_has_denkl"] and abs(float(r["doc_net_tutar"])) > 1.0:
            return "Denkleşmiş fakat net bakiye var" if language == "tr" else "Cleared but net balance remains"
        if r["doc_has_ters"] and abs(float(r["doc_net_tutar"])) > 1.0:
            return "Ters kayıt var, bakiye kontrolü gerekli" if language == "tr" else "Reversal exists, balance control needed"
        if int(r["doc_satir"]) >= 2 and not r["doc_has_denkl"]:
            return "Açık belge / ilişki kontrol" if language == "tr" else "Open document / relation control"
        return "Normal"

    if not doc_grp.empty:
        doc_grp["doc_relation_status"] = doc_grp.apply(_doc_status, axis=1)
        doc_map = {r["belge_anahtar"]: (int(r["doc_satir"]), float(r["doc_net_tutar"]), float(r["doc_abs_tutar"]), int(r["doc_unique_accounts"]), bool(r["doc_has_denkl"]), bool(r["doc_has_ters"]), bool(r["doc_duplicate_ref"]), int(r["doc_duplicate_ref_count"]), r["doc_relation_status"]) for _, r in doc_grp.iterrows()}
        df["doc_satir"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[0])
        df["doc_net_tutar"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[1])
        df["doc_abs_tutar"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[2])
        df["doc_unique_accounts"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[3])
        df["doc_duplicate_ref"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[6])
        df["doc_duplicate_ref_count"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[7])
        df["doc_relation_status"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[8])
    else:
        df["doc_satir"] = 0
        df["doc_net_tutar"] = 0.0
        df["doc_abs_tutar"] = 0.0
        df["doc_unique_accounts"] = 0
        df["doc_duplicate_ref"] = False
        df["doc_duplicate_ref_count"] = 0
        df["doc_relation_status"] = "Normal"

    df["risk_flag"] = (
        (df["risk_skoru"] >= 60)
        | (df["anomali_skoru"] >= 60)
        | (df["cost_change_alarm"].astype(str).str.contains("Alarm|Yeni Hareket", na=False))
        | (df["doc_relation_status"].astype(str) != "Normal")
    )

    _safe_progress(progress_cb, 92, "Muavin filtre listeleri hazırlanıyor..." if language == "tr" else "Preparing subledger filter lists...")
    cost_centers = sorted([
        f"{r['masraf_yeri']} - {r['masraf_yeri_tanimi']}".strip(" -")
        for _, r in df[["masraf_yeri","masraf_yeri_tanimi"]].drop_duplicates().iterrows()
        if normalize_text_value(r["masraf_yeri"]) or normalize_text_value(r["masraf_yeri_tanimi"])
    ])
    return {
        "raw_df": raw_df,
        "clean_df": df,
        "accounts": sorted(df["ana_hesap"].dropna().astype(str).unique().tolist()),
        "periods": sorted(df["donem"].dropna().astype(str).unique().tolist()),
        "doc_types": sorted([x for x in df["belge_turu"].dropna().astype(str).unique().tolist() if x]),
        "users": sorted([x for x in df["kullanici"].dropna().astype(str).unique().tolist() if x]),
        "contras": sorted([x for x in df["karsi_hesap_tanimi"].dropna().astype(str).unique().tolist() if x]),
        "cost_centers": cost_centers,
    }


def build_versioned_filename(operation_name: str, ext: str = ".xlsx", base_dir: Optional[str] = None) -> str:
    ensure_base_dir()
    target_dir = base_dir or BASE_DIR
    os.makedirs(target_dir, exist_ok=True)

    cleaned = re.sub(r"[^A-Za-z0-9ÇĞİÖŞÜçğıöşü_ -]", "", str(operation_name)).strip()
    cleaned = cleaned.replace(" ", "_") or "Rapor"
    date_txt = datetime.now().strftime("%d%m%Y")

    version = 1
    while True:
        filename = f"{cleaned}_{date_txt}_v{version}{ext}"
        full_path = os.path.join(target_dir, filename)
        if not os.path.exists(full_path):
            return full_path
        version += 1


def open_file_after_save(path: str):
    try:
        normalized_path = os.path.normpath(path)
        if sys.platform.startswith("win"):
            os.startfile(normalized_path)
        elif sys.platform == "darwin":
            os.system(f'open "{normalized_path}"')
        else:
            os.system(f'xdg-open "{normalized_path}" >/dev/null 2>&1 &')
    except Exception:
        pass


class ExcelPasteTableWidget(QTableWidget):
    def __init__(self, rows: int = 30, column_headers=None, stretch_last=True, parent=None):
        headers = column_headers if column_headers else ["Hesap", "Sorumlu"]
        col_count = len(headers)
        super().__init__(rows, col_count, parent)
        self.setHorizontalHeaderLabels(headers)
        self.verticalHeader().setVisible(False)
        self.setSelectionBehavior(QTableWidget.SelectItems)
        self.setSelectionMode(QTableWidget.ContiguousSelection)
        self.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed | QTableWidget.AnyKeyPressed)
        for i in range(col_count):
            if i == col_count - 1 and stretch_last:
                self.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)
            else:
                self.horizontalHeader().setSectionResizeMode(i, QHeaderView.ResizeToContents)
        self.setWordWrap(False)

    def keyPressEvent(self, event):
        if (event.modifiers() & Qt.ControlModifier) and event.key() == Qt.Key_V:
            self.paste_from_clipboard()
            return
        super().keyPressEvent(event)

    def paste_from_clipboard(self):
        text = QApplication.clipboard().text()
        if not text.strip():
            return

        start_row = max(self.currentRow(), 0)
        start_col = max(self.currentColumn(), 0)
        if start_col >= self.columnCount():
            start_col = 0

        lines = [line for line in text.splitlines() if line.strip()]
        required_rows = start_row + len(lines)
        if required_rows > self.rowCount():
            self.setRowCount(required_rows + 5)

        for row_offset, line in enumerate(lines):
            parts = [p.strip() for p in line.split("\t")]
            if len(parts) <= 1:
                parts = [p.strip() for p in re.split(r";|,", line)]
            if len(parts) < 2:
                continue

            values = parts[:self.columnCount()]
            if len(parts) > self.columnCount():
                fixed = parts[:self.columnCount() - 1]
                fixed.append("\t".join(parts[self.columnCount() - 1:]).strip())
                values = fixed

            for col_offset, value in enumerate(values):
                col = start_col + col_offset
                if col >= self.columnCount():
                    break
                self.setItem(start_row + row_offset, col, QTableWidgetItem(value))

    def clear_data(self, keep_rows: int = 30):
        self.clearContents()
        self.setRowCount(keep_rows)


class NumericTableWidgetItem(QTableWidgetItem):
    def __init__(self, text: str, sort_value=None):
        super().__init__(text)
        self.sort_value = sort_value

    def __lt__(self, other):
        if isinstance(other, NumericTableWidgetItem):
            a = self.sort_value
            b = other.sort_value
            if a is not None and b is not None:
                return a < b
        return super().__lt__(other)


def parse_sort_value(text: str):
    s = str(text).strip()
    if not s or s == "-":
        return None
    if s == "Yeni":
        return float("inf")
    if s.endswith("%"):
        num = s[:-1].strip().replace(".", "").replace(",", ".")
        try:
            return float(num)
        except Exception:
            return None
    num = s.replace(".", "").replace(",", ".")
    try:
        return float(num)
    except Exception:
        return None



def ensure_muavin_derived_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if df.empty:
        for col in [
            "belge_tarihi", "kayit_tarihi", "giris_tarihi",
            "belge_tarihi_dt", "kayit_tarihi_dt", "giris_tarihi_dt",
            "late_7day_flag", "late_7day_note", "late_7day_delay_days", "late_7day_deadline_dt",
        ]:
            if col not in df.columns:
                df[col] = pd.Series(dtype="object")
        return df

    for col in ["belge_tarihi", "kayit_tarihi", "giris_tarihi"]:
        if col not in df.columns:
            df[col] = ""

    if "belge_tarihi_dt" not in df.columns:
        df["belge_tarihi_dt"] = df["belge_tarihi"].map(parse_date_flexible)
    if "kayit_tarihi_dt" not in df.columns:
        df["kayit_tarihi_dt"] = df["kayit_tarihi"].map(parse_date_flexible)
    if "giris_tarihi_dt" not in df.columns:
        df["giris_tarihi_dt"] = df["giris_tarihi"].map(parse_date_flexible)

    need_late_cols = any(
        col not in df.columns
        for col in ["late_7day_flag", "late_7day_note", "late_7day_delay_days", "late_7day_deadline_dt"]
    )
    if need_late_cols:
        late_results = df.apply(
            lambda x: build_late_7day_note(x.get("belge_tarihi_dt"), x.get("kayit_tarihi_dt"), x.get("giris_tarihi_dt")),
            axis=1
        )
        df["late_7day_flag"] = late_results.map(lambda t: bool(t[0]))
        df["late_7day_note"] = late_results.map(lambda t: t[1])
        df["late_7day_delay_days"] = late_results.map(lambda t: int(t[2]))
        df["late_7day_belge_deadline_dt"] = late_results.map(lambda t: t[3])
        df["late_7day_deadline_dt"] = late_results.map(lambda t: t[4])

    return df


AUDIT_MATERIALITY_THRESHOLD = 50000.0

def muavin_risk_level(score: float) -> str:
    score = safe_float(score)
    if score >= 60:
        return "Kritik"
    if score >= 40:
        return "Yüksek"
    if score >= 20:
        return "Orta"
    return "Düşük"

def muavin_risk_level_en(score: float) -> str:
    score = safe_float(score)
    if score >= 60:
        return "Critical"
    if score >= 40:
        return "High"
    if score >= 20:
        return "Medium"
    return "Low"

def join_unique_notes(parts) -> str:
    vals=[]
    seen=set()
    for p in parts:
        s=normalize_text_value(p)
        if not s or s in seen:
            continue
        vals.append(s); seen.add(s)
    return " | ".join(vals)

def build_muavin_audit_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if df.empty:
        for col in [
            "duplicate_type", "tax_risk_flag", "materiality_flag", "user_risk_flag",
            "audit_risk_score", "audit_risk_level", "audit_risk_level_en",
            "audit_risk_explanation", "combined_risk_flag", "relation_expected",
            "relation_missing", "relation_status_detail", "fraud_risk_flag"
        ]:
            if col not in df.columns:
                df[col] = pd.Series(dtype="object")
        return df

    work = df.copy()
    work["abs_up_tutar"] = work["up_tutar"].abs()
    work["tax_risk_flag"] = False
    work.loc[work["hesap_prefix"].isin(["191", "360", "391"]) & (work["vergi_gostergesi"].astype(str).str.strip() == ""), "tax_risk_flag"] = True
    work["materiality_flag"] = work["abs_up_tutar"] >= AUDIT_MATERIALITY_THRESHOLD

    exact_counts = work[(work["referans"].astype(str).str.strip() != "")].groupby(["referans", "karsi_hesap", "abs_up_tutar"])["belge_anahtar"].transform("size")
    fuzzy_counts = work[(work["karsi_hesap"].astype(str).str.strip() != "")].groupby(["karsi_hesap", "abs_up_tutar", "donem"])["belge_anahtar"].transform("size")
    split_counts = work[(work["referans"].astype(str).str.strip() != "")].groupby(["referans", "karsi_hesap"])["belge_anahtar"].transform("nunique")
    work["duplicate_type"] = ""
    work.loc[exact_counts.reindex(work.index).fillna(0) > 1, "duplicate_type"] = "Exact duplicate"
    work.loc[(work["duplicate_type"] == "") & (split_counts.reindex(work.index).fillna(0) > 1), "duplicate_type"] = "Split duplicate"
    work.loc[(work["duplicate_type"] == "") & (fuzzy_counts.reindex(work.index).fillna(0) > 1), "duplicate_type"] = "Fuzzy duplicate"
    work.loc[(work["duplicate_type"] == "") & (work.get("duplicate_open_reference", False) == True), "duplicate_type"] = "Reference duplicate"

    user_risk = work.groupby("kullanici").agg(risk_ort=("risk_skoru", "mean"), riskli_oran=("risk_flag", "mean"), satir=("kullanici", "size")).reset_index()
    if not user_risk.empty:
        user_risk["user_risk_flag"] = (user_risk["risk_ort"] >= 60) | (user_risk["riskli_oran"] >= 0.30)
        user_map = dict(zip(user_risk["kullanici"], user_risk["user_risk_flag"]))
        work["user_risk_flag"] = work["kullanici"].map(lambda x: bool(user_map.get(x, False)))
    else:
        work["user_risk_flag"] = False

    doc_accounts = work.groupby("belge_anahtar")["hesap_prefix"].agg(lambda s: sorted({str(x).strip() for x in s if str(x).strip()})).to_dict()
    rel_expected = {}
    rel_missing = {}
    rel_detail = {}
    for belge, accs in doc_accounts.items():
        expected=[]; missing=[]
        aset=set(accs)
        if "320" in aset or "321" in aset:
            expected.append("191/360 veya 191/391 vergi ilişkisi")
            if not ({"191","360","391"} & aset):
                missing.append("191/360/391")
        if "191" in aset and not ({"320","321","360","391"} & aset):
            expected.append("320/360/391 karşılığı")
            missing.append("320/360/391")
        if "360" in aset and "191" not in aset:
            expected.append("191 karşılığı")
            missing.append("191")
        rel_expected[belge] = ", ".join(expected)
        rel_missing[belge] = ", ".join(sorted(set(missing)))
        rel_detail[belge] = "İlişki eksik" if missing else ("Beklenen ilişki seti oluşmuş" if expected else "Standart ilişki beklentisi tanımsız")
    work["relation_expected"] = work["belge_anahtar"].map(lambda x: rel_expected.get(x, ""))
    work["relation_missing"] = work["belge_anahtar"].map(lambda x: rel_missing.get(x, ""))
    work["relation_status_detail"] = work["belge_anahtar"].map(lambda x: rel_detail.get(x, "Standart ilişki beklentisi tanımsız"))

    scores=[]; levels=[]; levels_en=[]; exps=[]; comb=[]; fraud=[]
    for _, r in work.iterrows():
        score=0; notes=[]
        dup_type = normalize_text_value(r.get("duplicate_type", ""))
        if dup_type:
            score += 40; notes.append(f"Mükerrer kayıt riski ({dup_type})")
        if bool(r.get("late_7day_flag", False)):
            score += 30; notes.append("7 gün sonrası kayıt")
        if normalize_text_value(r.get("doc_relation_status", "")) not in ["", "Normal", "Belge netleşmiş"]:
            score += 25; notes.append(f"Belge ilişkisel test: {r.get('doc_relation_status')}")
        if bool(r.get("tax_risk_flag", False)):
            score += 20; notes.append("Vergi göstergesi / vergi hesap riski")
        if bool(r.get("materiality_flag", False)):
            score += 20; notes.append(f"Önemlilik eşiği aşıldı (>{format_number(AUDIT_MATERIALITY_THRESHOLD)})")
        if bool(r.get("user_risk_flag", False)):
            score += 15; notes.append("Riskli kullanıcı paterni")
        if normalize_text_value(r.get("cost_change_alarm", "")).startswith("Artış Alarmı") or normalize_text_value(r.get("cost_change_alarm", "")).startswith("Azalış Alarmı"):
            score += 10; notes.append(f"Trend sapması: {r.get('cost_change_alarm')}")
        if normalize_text_value(r.get("relation_missing", "")):
            score += 15; notes.append(f"Eksik hesap ilişkisi: {r.get('relation_missing')}")
        score = min(score, 100)
        if bool(r.get("late_7day_flag", False)) and dup_type:
            comb.append("Cut-off + Duplicate")
        elif bool(r.get("tax_risk_flag", False)) and dup_type:
            comb.append("Tax + Duplicate")
        elif bool(r.get("user_risk_flag", False)) and dup_type:
            comb.append("User + Duplicate")
        elif bool(r.get("late_7day_flag", False)) and bool(r.get("tax_risk_flag", False)):
            comb.append("Cut-off + Tax")
        else:
            comb.append("")
        fraud.append(bool(dup_type) and bool(r.get("user_risk_flag", False)))
        scores.append(score)
        levels.append(muavin_risk_level(score))
        levels_en.append(muavin_risk_level_en(score))
        exps.append(join_unique_notes(notes) or "Kontrol bulgusu yok")
    work["audit_risk_score"] = scores
    work["audit_risk_level"] = levels
    work["audit_risk_level_en"] = levels_en
    work["audit_risk_explanation"] = exps
    work["combined_risk_flag"] = comb
    work["fraud_risk_flag"] = fraud
    return work

def set_table_item(table: QTableWidget, row: int, col: int, text: str, align=None, badge=None):
    sort_value = parse_sort_value(text)
    item = NumericTableWidgetItem(text, sort_value) if sort_value is not None else QTableWidgetItem(text)
    if align is not None:
        item.setTextAlignment(align)

    if badge == "danger":
        item.setBackground(QColor("#FEF2F2"))
        item.setForeground(QColor("#B91C1C"))
    elif badge == "success":
        item.setBackground(QColor("#ECFDF5"))
        item.setForeground(QColor("#047857"))
    elif badge == "warn":
        item.setBackground(QColor("#FEF3C7"))
        item.setForeground(QColor("#92400E"))
    elif badge == "subtotal":
        item.setBackground(QColor("#EFF6FF"))
        item.setForeground(QColor("#1E3A8A"))

    table.setItem(row, col, item)


def join_notes(parts: List[str]) -> str:
    cleaned = []
    seen = set()
    for p in parts:
        p = str(p).strip()
        if not p:
            continue
        if p not in seen:
            cleaned.append(p)
            seen.add(p)
    return " | ".join(cleaned)



def collect_matching_notes(
    notes: List[Dict],
    hesap_value: str,
    ana_hesap_value: str = "",
    masraf_yeri: str = ""
) -> Tuple[str, str]:
    hedef_hesap = normalize_hesap_prefix(hesap_value)
    hedef_ana_hesap = str(ana_hesap_value).strip()
    hedef_masraf = str(masraf_yeri).strip()

    tr_parts = []
    en_parts = []

    for n in notes:
        kayit_hesap = normalize_hesap_prefix(n.get("hesap", ""))
        kayit_ana_hesap = str(n.get("anaHesap", "")).strip()
        kayit_masraf = str(n.get("masrafYeri", "")).strip()

        if kayit_ana_hesap:
            if kayit_ana_hesap != hedef_ana_hesap:
                continue
        else:
            if kayit_hesap != hedef_hesap:
                continue

        if hedef_masraf:
            if kayit_masraf not in ["", hedef_masraf]:
                continue

        tr_parts.append(str(n.get("not", "")).strip())
        en_parts.append(str(n.get("noteEn", n.get("not", ""))).strip())

    return join_notes(tr_parts), join_notes(en_parts)


BALANCE_GROUP_ORDER = [
    "Dönen Varlıklar",
    "Duran Varlıklar",
    "AKTİF TOPLAM",
    "Kısa Vadeli Yabancı Kaynaklar",
    "Uzun Vadeli Yabancı Kaynaklar",
    "Öz Kaynaklar",
    "PASİF TOPLAM",
    "AKTİF - PASİF FARKI",
]

INCOME_STATEMENT_TEMPLATE = [
    {"kod": "60", "label": "Brüt Satışlar", "sign": "+", "desc": "600-602 toplamı", "specs": [600, 601, 602], "rowType": "group"},
    {"kod": "600", "label": "Yurtiçi Satışlar", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [600], "rowType": "detail"},
    {"kod": "601", "label": "Yurtdışı Satışlar", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [601], "rowType": "detail"},
    {"kod": "602", "label": "Diğer Gelirler", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [602], "rowType": "detail"},
    {"kod": "61", "label": "Satış İndirimleri (-)", "sign": "-", "desc": "610-612 toplamı", "specs": [610, 611, 612], "rowType": "group"},
    {"kod": "610", "label": "Satıştan İadeler (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [610], "rowType": "detail"},
    {"kod": "611", "label": "Satış İskontoları (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [611], "rowType": "detail"},
    {"kod": "612", "label": "Diğer İndirimler (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [612], "rowType": "detail"},
    {"kod": "", "label": "Net Satışlar", "sign": "", "desc": "Brüt Satışlar - Satış İndirimleri", "formula": "net_sales", "rowType": "subtotal"},
    {"kod": "62", "label": "Satışların Maliyeti (-)", "sign": "-", "desc": "620-623 toplamı", "specs": [620, 621, 622, 623], "rowType": "group"},
    {"kod": "620", "label": "Satılan Mamuller Maliyeti (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [620], "rowType": "detail"},
    {"kod": "621", "label": "Satılan Ticari Mallar Maliyeti (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [621], "rowType": "detail"},
    {"kod": "622", "label": "Satılan Hizmet Maliyeti (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [622], "rowType": "detail"},
    {"kod": "623", "label": "Diğer Satışların Maliyeti (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [623], "rowType": "detail"},
    {"kod": "", "label": "Brüt Satış Kârı / Zararı", "sign": "", "desc": "Net Satışlar - Satışların Maliyeti", "formula": "gross_profit", "rowType": "subtotal"},
    {"kod": "63", "label": "Faaliyet Giderleri (-)", "sign": "-", "desc": "630-632 toplamı", "specs": [630, 631, 632], "rowType": "group"},
    {"kod": "630", "label": "Araştırma ve Geliştirme Giderleri (-)", "sign": "-", "desc": "7'li hesaplardan yansıyabilir / doğrudan olabilir", "specs": [630], "rowType": "detail"},
    {"kod": "631", "label": "Pazarlama, Satış ve Dağıtım Giderleri (-)", "sign": "-", "desc": "7'li hesaplardan yansıyabilir / doğrudan olabilir", "specs": [631], "rowType": "detail"},
    {"kod": "632", "label": "Genel Yönetim Giderleri (-)", "sign": "-", "desc": "7'li hesaplardan yansıyabilir / doğrudan olabilir", "specs": [632], "rowType": "detail"},
    {"kod": "64", "label": "Diğer Faaliyetlerden Olağan Gelir ve Kârlar", "sign": "+", "desc": "640-649 toplamı", "specs": [(640, 649)], "rowType": "group"},
    {"kod": "640", "label": "İştiraklerden Temettü Gelirleri", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [640], "rowType": "detail"},
    {"kod": "641", "label": "Bağlı Ortaklıklardan Temettü Gelirleri", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [641], "rowType": "detail"},
    {"kod": "642", "label": "Faiz Gelirleri", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [642], "rowType": "detail"},
    {"kod": "643", "label": "Komisyon Gelirleri", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [643], "rowType": "detail"},
    {"kod": "644", "label": "Konusu Kalmayan Karşılıklar", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [644], "rowType": "detail"},
    {"kod": "645", "label": "Menkul Kıymet Satış Kârları", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [645], "rowType": "detail"},
    {"kod": "646", "label": "Kambiyo Kârları", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [646], "rowType": "detail"},
    {"kod": "647", "label": "Reeskont Faiz Gelirleri", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [647], "rowType": "detail"},
    {"kod": "648", "label": "Enflasyon Düzeltmesi Kârları", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [648], "rowType": "detail"},
    {"kod": "649", "label": "Diğer Olağan Gelir ve Kârlar", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [649], "rowType": "detail"},
    {"kod": "65", "label": "Diğer Faaliyetlerden Olağan Gider ve Zararlar (-)", "sign": "-", "desc": "653-659 toplamı", "specs": [(653, 659)], "rowType": "group"},
    {"kod": "653", "label": "Komisyon Giderleri (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [653], "rowType": "detail"},
    {"kod": "654", "label": "Karşılık Giderleri (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [654], "rowType": "detail"},
    {"kod": "655", "label": "Menkul Kıymet Satış Zararları (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [655], "rowType": "detail"},
    {"kod": "656", "label": "Kambiyo Zararları (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [656], "rowType": "detail"},
    {"kod": "657", "label": "Reeskont Faiz Giderleri (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [657], "rowType": "detail"},
    {"kod": "658", "label": "Enflasyon Düzeltmesi Zararları (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [658], "rowType": "detail"},
    {"kod": "659", "label": "Diğer Olağan Gider ve Zararlar (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [659], "rowType": "detail"},
    {"kod": "66", "label": "Finansman Giderleri (-)", "sign": "-", "desc": "660-661 toplamı", "specs": [660, 661], "rowType": "group"},
    {"kod": "660", "label": "Kısa Vadeli Borçlanma Giderleri (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [660], "rowType": "detail"},
    {"kod": "661", "label": "Uzun Vadeli Borçlanma Giderleri (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [661], "rowType": "detail"},
    {"kod": "", "label": "Olağan Kâr / Zarar", "sign": "", "desc": "Brüt kâr/zarar - faaliyet giderleri + olağan gelirler - olağan giderler - finansman giderleri", "formula": "ordinary_profit", "rowType": "subtotal"},
    {"kod": "67", "label": "Olağandışı Gelir ve Kârlar", "sign": "+", "desc": "671 ve 679 toplamı", "specs": [671, 679], "rowType": "group"},
    {"kod": "671", "label": "Önceki Dönem Gelir ve Kârları", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [671], "rowType": "detail"},
    {"kod": "679", "label": "Diğer Olağandışı Gelir ve Kârlar", "sign": "+", "desc": "Alt hesaplardan toplanır", "specs": [679], "rowType": "detail"},
    {"kod": "68", "label": "Olağandışı Gider ve Zararlar (-)", "sign": "-", "desc": "680, 681 ve 689 toplamı", "specs": [680, 681, 689], "rowType": "group"},
    {"kod": "680", "label": "Çalışmayan Kısım Gider ve Zararları (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [680], "rowType": "detail"},
    {"kod": "681", "label": "Önceki Dönem Gider ve Zararları (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [681], "rowType": "detail"},
    {"kod": "689", "label": "Diğer Olağandışı Gider ve Zararlar (-)", "sign": "-", "desc": "Alt hesaplardan toplanır", "specs": [689], "rowType": "detail"},
    {"kod": "690", "label": "Dönem Kârı veya Zararı", "sign": "", "desc": "Olağan kâr/zarar + olağandışı gelirler - olağandışı giderler", "formula": "profit_before_tax", "specs": [690], "rowType": "subtotal"},
    {"kod": "691", "label": "Dönem Kârı Vergi ve Diğer Yasal Yükümlülük Karşılıkları (-)", "sign": "-", "desc": "Vergi karşılığını ilgili hesaptan çeker", "specs": [691], "rowType": "group"},
    {"kod": "692", "label": "Dönem Net Kârı veya Zararı", "sign": "", "desc": "690 - 691", "formula": "net_profit", "specs": [692], "rowType": "subtotal"},
]

def get_balance_group_by_account(hesap_prefix: str) -> str:
    first = str(hesap_prefix).strip()[:1]
    if first == "1":
        return "Dönen Varlıklar"
    if first == "2":
        return "Duran Varlıklar"
    if first == "3":
        return "Kısa Vadeli Yabancı Kaynaklar"
    if first == "4":
        return "Uzun Vadeli Yabancı Kaynaklar"
    if first == "5":
        return "Öz Kaynaklar"
    return ""

def convert_balance_value(hesap_prefix: str, amount: float) -> float:
    first = str(hesap_prefix).strip()[:1]
    return amount if first in ["1", "2"] else (-amount if first in ["3", "4", "5"] else 0.0)

def matches_account(hesap_prefix: str, specs: List):
    hesap_prefix = str(hesap_prefix).strip()
    if not hesap_prefix:
        return False
    try:
        kod = int(hesap_prefix)
    except Exception:
        return False

    for spec in specs:
        if isinstance(spec, tuple) and len(spec) == 2:
            if spec[0] <= kod <= spec[1]:
                return True
        else:
            if kod == int(spec):
                return True
    return False

def convert_income_value(hesap_prefix: str, amount: float) -> float:
    if matches_account(hesap_prefix, [600, 601, 602, (640, 649), 671, 679, 690, 692]):
        return -amount
    if matches_account(hesap_prefix, [610, 611, 612, 620, 621, 622, 623, 630, 631, 632, (653, 659), 660, 661, 680, 681, 689, 691]):
        return amount
    return 0.0



def normalize_balance_periods(periods: List[str]) -> List[str]:
    ordered = sort_periods(periods)
    if OPENING_PERIOD_TR in ordered:
        ordered = [OPENING_PERIOD_TR] + [p for p in ordered if p != OPENING_PERIOD_TR]
    return ordered

def build_balance_sheet_summary(rows: List[Dict], periods: List[str], current_period: Optional[str] = None, previous_period: Optional[str] = None) -> List[Dict]:
    periods = normalize_balance_periods(periods)
    base_groups = {
        "Dönen Varlıklar": {p: 0.0 for p in periods},
        "Duran Varlıklar": {p: 0.0 for p in periods},
        "Kısa Vadeli Yabancı Kaynaklar": {p: 0.0 for p in periods},
        "Uzun Vadeli Yabancı Kaynaklar": {p: 0.0 for p in periods},
        "Öz Kaynaklar": {p: 0.0 for p in periods},
    }

    for row in rows:
        hesap = row.get("hesap", "")
        group = get_balance_group_by_account(hesap)
        if not group:
            continue
        for p in periods:
            raw = safe_float(row.get("valuesNumeric", {}).get(p, 0.0))
            base_groups[group][p] += convert_balance_value(hesap, raw)

    aktif = {p: base_groups["Dönen Varlıklar"][p] + base_groups["Duran Varlıklar"][p] for p in periods}
    pasif = {
        p: base_groups["Kısa Vadeli Yabancı Kaynaklar"][p] +
           base_groups["Uzun Vadeli Yabancı Kaynaklar"][p] +
           base_groups["Öz Kaynaklar"][p]
        for p in periods
    }
    diff = {p: aktif[p] - pasif[p] for p in periods}

    ordered_rows = []
    row_map = {
        "Dönen Varlıklar": ("group", base_groups["Dönen Varlıklar"]),
        "Duran Varlıklar": ("group", base_groups["Duran Varlıklar"]),
        "AKTİF TOPLAM": ("subtotal", aktif),
        "Kısa Vadeli Yabancı Kaynaklar": ("group", base_groups["Kısa Vadeli Yabancı Kaynaklar"]),
        "Uzun Vadeli Yabancı Kaynaklar": ("group", base_groups["Uzun Vadeli Yabancı Kaynaklar"]),
        "Öz Kaynaklar": ("group", base_groups["Öz Kaynaklar"]),
        "PASİF TOPLAM": ("subtotal", pasif),
        "AKTİF - PASİF FARKI": ("danger", diff),
    }

    for label in BALANCE_GROUP_ORDER:
        row_type, values_num = row_map[label]
        vals = dict(values_num)
        vals["TOPLAM"] = sum(values_num.values())
        current_value = vals.get(current_period, 0.0) if current_period else 0.0
        previous_value = vals.get(previous_period, 0.0) if previous_period else 0.0
        degisim = calc_change_percent(current_value, previous_value) if current_period and previous_period else "-"
        ordered_rows.append({
            "label": label,
            "rowType": row_type,
            "valuesNumeric": vals,
            "values": {k: format_number(v) for k, v in vals.items()},
            "compareCurrent": format_number(current_value) if current_period else "0,00",
            "comparePrevious": format_number(previous_value) if previous_period else "0,00",
            "degisim": degisim,
            "criticalVariance": is_critical_variance(degisim),
            "varianceStatus": critical_variance_status(degisim),
        })

    income_summary = build_income_statement_summary(rows, periods, current_period=current_period, previous_period=previous_period)
    net_profit_values = {p: 0.0 for p in periods}
    for income_row in income_summary:
        if income_row.get("label") == "Dönem Net Kârı veya Zararı":
            net_profit_values = {p: safe_float(income_row.get("valuesNumeric", {}).get(p, 0.0)) for p in periods}
            break

    class8_values = {p: 0.0 for p in periods}
    class8_accounts_by_period = {p: [] for p in periods}
    for row in rows:
        hesap = str(row.get("hesap", "")).strip()
        if not hesap.startswith("8"):
            continue
        hesap_label = str(row.get("anaHesapTam") or row.get("hesap") or "").strip()
        hesap_name = str(row.get("hesapAdi") or row.get("hesapAdiEn") or "").strip()
        acct_text = f"{hesap_label} - {hesap_name}".strip(" -")
        for p in periods:
            raw = safe_float(row.get("valuesNumeric", {}).get(p, 0.0))
            class8_values[p] += raw
            if abs(raw) > 1e-9 and acct_text:
                class8_accounts_by_period[p].append(acct_text)

    residual_values = {p: diff[p] - net_profit_values.get(p, 0.0) for p in periods}

    def _append_extra_balance_row(label: str, values_num: Dict[str, float], row_type: str = "group", status_text: str = "Normal"):
        vals = dict(values_num)
        vals["TOPLAM"] = sum(values_num.values())
        current_value = vals.get(current_period, 0.0) if current_period else 0.0
        previous_value = vals.get(previous_period, 0.0) if previous_period else 0.0
        degisim = calc_change_percent(current_value, previous_value) if current_period and previous_period else "-"
        ordered_rows.append({
            "label": label,
            "rowType": row_type,
            "valuesNumeric": vals,
            "values": {k: format_number(v) for k, v in vals.items()},
            "compareCurrent": format_number(current_value) if current_period else "0,00",
            "comparePrevious": format_number(previous_value) if previous_period else "0,00",
            "degisim": degisim,
            "criticalVariance": is_critical_variance(degisim),
            "varianceStatus": status_text,
        })

    current_class8_accounts = safe_unique_join(class8_accounts_by_period.get(current_period, []) if current_period else [], limit=5)
    class8_status = "8'li hesap bakiyesi yok"
    if any(abs(v) > 1e-9 for v in class8_values.values()):
        class8_status = "8'li hesaplar bakiye veriyor"
        if current_class8_accounts:
            class8_status = f"8'li hesaplar bakiye veriyor: {current_class8_accounts}"

    current_residual = residual_values.get(current_period, 0.0) if current_period else 0.0
    current_class8 = class8_values.get(current_period, 0.0) if current_period else 0.0
    if abs(current_residual) <= 1.0:
        control_status = "Bilanço farkı ile net kâr uyumlu"
        control_row_type = "success"
    elif abs(current_residual - current_class8) <= 1.0:
        control_status = "Fark 8'li hesap bakiyesinden kaynaklı"
        if current_class8_accounts:
            control_status += f": {current_class8_accounts}"
        control_row_type = "warn"
    elif abs(current_class8) > 1.0:
        control_status = "8'li hesaplar incelenmeli"
        if current_class8_accounts:
            control_status += f": {current_class8_accounts}"
        control_row_type = "danger"
    else:
        control_status = "Fark incelenmeli"
        control_row_type = "danger"

    _append_extra_balance_row("DÖNEM NET KÂRI / ZARARI", net_profit_values, "subtotal", "Gelir tablosu dip toplamı")
    _append_extra_balance_row("8'Lİ HESAP BAKİYESİ", class8_values, "warn" if any(abs(v) > 1e-9 for v in class8_values.values()) else "success", class8_status)
    _append_extra_balance_row("BİLANÇO FARKI - NET KÂR KONTROLÜ", residual_values, control_row_type, control_status)

    return ordered_rows

def build_income_statement_summary(rows: List[Dict], periods: List[str], current_period: Optional[str] = None, previous_period: Optional[str] = None) -> List[Dict]:
    line_values = {}

    for spec in INCOME_STATEMENT_TEMPLATE:
        if "specs" not in spec:
            continue
        line_values[spec["label"]] = {p: 0.0 for p in periods}

    for row in rows:
        hesap = row.get("hesap", "")
        for spec in INCOME_STATEMENT_TEMPLATE:
            specs = spec.get("specs")
            if not specs:
                continue
            if matches_account(hesap, specs):
                for p in periods:
                    raw = safe_float(row.get("valuesNumeric", {}).get(p, 0.0))
                    line_values[spec["label"]][p] += convert_income_value(hesap, raw)

    calculated = {
        "Net Satışlar": {p: 0.0 for p in periods},
        "Brüt Satış Kârı / Zararı": {p: 0.0 for p in periods},
        "Olağan Kâr / Zarar": {p: 0.0 for p in periods},
        "Dönem Kârı veya Zararı": {p: 0.0 for p in periods},
        "Dönem Net Kârı veya Zararı": {p: 0.0 for p in periods},
    }

    for p in periods:
        calculated["Net Satışlar"][p] = line_values["Brüt Satışlar"][p] - line_values["Satış İndirimleri (-)"][p]
        calculated["Brüt Satış Kârı / Zararı"][p] = calculated["Net Satışlar"][p] - line_values["Satışların Maliyeti (-)"][p]
        calculated["Olağan Kâr / Zarar"][p] = (
            calculated["Brüt Satış Kârı / Zararı"][p]
            - line_values["Faaliyet Giderleri (-)"][p]
            + line_values["Diğer Faaliyetlerden Olağan Gelir ve Kârlar"][p]
            - line_values["Diğer Faaliyetlerden Olağan Gider ve Zararlar (-)"][p]
            - line_values["Finansman Giderleri (-)"][p]
        )
        calculated["Dönem Kârı veya Zararı"][p] = (
            calculated["Olağan Kâr / Zarar"][p]
            + line_values["Olağandışı Gelir ve Kârlar"][p]
            - line_values["Olağandışı Gider ve Zararlar (-)"][p]
        )
        calculated["Dönem Net Kârı veya Zararı"][p] = (
            calculated["Dönem Kârı veya Zararı"][p]
            - line_values["Dönem Kârı Vergi ve Diğer Yasal Yükümlülük Karşılıkları (-)"][p]
        )

    formula_map = {
        "net_sales": "Net Satışlar",
        "gross_profit": "Brüt Satış Kârı / Zararı",
        "ordinary_profit": "Olağan Kâr / Zarar",
        "profit_before_tax": "Dönem Kârı veya Zararı",
        "net_profit": "Dönem Net Kârı veya Zararı",
    }

    ordered_rows = []
    for spec in INCOME_STATEMENT_TEMPLATE:
        label = spec["label"]
        if "formula" in spec:
            values_num = calculated[formula_map[spec["formula"]]]
        else:
            values_num = line_values.get(label, {p: 0.0 for p in periods})

        vals = dict(values_num)
        vals["TOPLAM"] = sum(values_num.values())
        current_value = vals.get(current_period, 0.0) if current_period else 0.0
        previous_value = vals.get(previous_period, 0.0) if previous_period else 0.0
        degisim = calc_change_percent(current_value, previous_value) if current_period and previous_period else "-"

        ordered_rows.append({
            "kod": spec.get("kod", ""),
            "label": label,
            "sign": spec.get("sign", ""),
            "description": spec.get("desc", ""),
            "rowType": spec.get("rowType", "group"),
            "valuesNumeric": vals,
            "values": {k: format_number(v) for k, v in vals.items()},
            "compareCurrent": format_number(current_value) if current_period else "0,00",
            "comparePrevious": format_number(previous_value) if previous_period else "0,00",
            "degisim": degisim,
            "criticalVariance": is_critical_variance(degisim),
            "varianceStatus": critical_variance_status(degisim),
        })

    return ordered_rows

def compute_dashboard_financial_metrics(rows: List[Dict], current_period: str) -> Dict[str, float]:
    periods = [OPENING_PERIOD_TR] + [p["tr"] for p in ALL_PERIODS]
    if current_period not in periods:
        periods.append(current_period)
    metric_period = current_period if current_period else OPENING_PERIOD_TR
    balance_rows = build_balance_sheet_summary(rows, [metric_period], current_period=metric_period, previous_period=None)
    income_rows = build_income_statement_summary(rows, [current_period], current_period=current_period, previous_period=None)

    metrics = {"aktif_toplam": 0.0, "pasif_toplam": 0.0, "net_kar": 0.0}
    for row in balance_rows:
        if row["label"] == "AKTİF TOPLAM":
            metrics["aktif_toplam"] = safe_float(row["valuesNumeric"].get(current_period, 0.0))
        elif row["label"] == "PASİF TOPLAM":
            metrics["pasif_toplam"] = safe_float(row["valuesNumeric"].get(current_period, 0.0))
    for row in income_rows:
        if row["label"] == "Dönem Net Kârı veya Zararı":
            metrics["net_kar"] = safe_float(row["valuesNumeric"].get(current_period, 0.0))
            break
    return metrics

def export_financial_statement_excel(path: str, title: str, headers: List[str], rows: List[List[str]], subtotal_row_indexes=None):
    subtotal_row_indexes = set(subtotal_row_indexes or [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"

    title_fill = PatternFill("solid", fgColor="1D4ED8")
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    subtotal_fill = PatternFill("solid", fgColor="EFF6FF")
    thin = Side(style="thin", color="D1D5DB")

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = title_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    header_row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, row in enumerate(rows, start=header_row + 1):
        is_sub = (r_idx - (header_row + 1)) in subtotal_row_indexes
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="right" if c_idx > 1 else "left", vertical="center")
            if is_sub:
                cell.fill = subtotal_fill
                cell.font = Font(bold=True)

    for col_cells in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 4, 14), 36)

    wb.save(path)

FISCAL_MONTHS_TR = {
    1: "Şubat",
    2: "Mart",
    3: "Nisan",
    4: "Mayıs",
    5: "Haziran",
    6: "Temmuz",
    7: "Ağustos",
    8: "Eylül",
    9: "Ekim",
    10: "Kasım",
    11: "Aralık",
    12: "Ocak",
}

FISCAL_MONTHS_EN = {
    "Şubat": "February",
    "Mart": "March",
    "Nisan": "April",
    "Mayıs": "May",
    "Haziran": "June",
    "Temmuz": "July",
    "Ağustos": "August",
    "Eylül": "September",
    "Ekim": "October",
    "Kasım": "November",
    "Aralık": "December",
    "Ocak": "January",
}


def parse_fiscal_period_label(value) -> str:
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    s = s.replace(",", ".")
    m = re.search(r"(\d{1,3})\s*\.\s*(\d{4})", s)
    if not m:
        return s
    period_no = int(m.group(1))
    year = int(m.group(2))
    if period_no < 1 or period_no > 12:
        return s
    month_name = FISCAL_MONTHS_TR.get(period_no, "")
    display_year = year + 1 if period_no == 12 else year
    return f"{month_name} {display_year}".strip()


def regular_period_sort_key(label: str):
    s = str(label).strip()
    if not s:
        return (9999, 99)
    parts = s.split()
    if len(parts) < 2:
        return (9999, 99)
    month_name = " ".join(parts[:-1]).strip()
    try:
        year = int(parts[-1])
    except Exception:
        year = 9999
    month_order = {v: k for k, v in FISCAL_MONTHS_TR.items()}
    return (year, month_order.get(month_name, 99))


def parse_date_flexible(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value
    try:
        dt = pd.to_datetime(value, dayfirst=True, errors="coerce")
        return dt
    except Exception:
        return pd.NaT


def month_year_tr_from_date(value) -> str:
    dt = parse_date_flexible(value)
    if pd.isna(dt):
        return ""
    month_names = {
        1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan", 5: "Mayıs", 6: "Haziran",
        7: "Temmuz", 8: "Ağustos", 9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"
    }
    return f"{month_names.get(int(dt.month), '')} {int(dt.year)}".strip()


def format_date_display(value) -> str:
    dt = parse_date_flexible(value)
    if pd.isna(dt):
        return "-"
    return dt.strftime("%d.%m.%Y")


def next_month_seventh(value):
    dt = parse_date_flexible(value)
    if pd.isna(dt):
        return pd.NaT
    year = int(dt.year)
    month = int(dt.month)
    if month == 12:
        year += 1
        month = 1
    else:
        month += 1
    try:
        return pd.Timestamp(year=year, month=month, day=7)
    except Exception:
        return pd.NaT


def compute_late_7day_control(posting_date, created_date):
    is_late, note, delay_days, _belge_deadline, kayit_deadline = build_late_7day_note(pd.NaT, posting_date, created_date)
    return is_late, note, delay_days, kayit_deadline


def normalize_vendor_code(value) -> str:
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    if re.fullmatch(r"[0-9]+(?:\.0+)?", s):
        try:
            return str(int(float(s)))
        except Exception:
            return s
    return s


def normalize_invoice_no(value) -> str:
    s = str(value).strip().upper()
    if not s or s.lower() == "nan":
        return ""
    s = s.replace(" ", "")
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s


def safe_int(value) -> int:
    try:
        return int(value)
    except Exception:
        try:
            return int(float(value))
        except Exception:
            return 0


def calc_change_percent_text(current: float, previous: float) -> str:
    current = safe_float(current)
    previous = safe_float(previous)
    if abs(previous) < 1e-12:
        if abs(current) < 1e-12:
            return "0,00%"
        return "Yeni"
    pct = ((current - previous) / abs(previous)) * 100.0
    return f"{pct:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")


def evaluate_regular_risk(note_text: str, eba_count: int, zfi_muhatap: str, adet_change_text: str, tutar_change_text: str, eba_rule: str = "", zfi_rule: str = "") -> tuple:
    note = str(note_text or "")
    adet_pct = parse_percent_text(adet_change_text)
    tutar_pct = parse_percent_text(tutar_change_text)
    has_eba = safe_int(eba_count) > 0
    has_zfi = normalize_text_value(zfi_muhatap) != ""

    score = 0
    rules = []

    if "Son Dönem Ft.Yok" in note and not has_eba:
        score += 50
        rules.append("Son dönem yok / EBA yok")
    elif "Son Dönem Ft.Yok" in note and has_eba:
        score += 30
        rules.append("Son dönem yok / EBA var")

    if "Aralıklı Gelen" in note:
        score += 20
        rules.append("Aralıklı akış")

    if "Beklenen" in note:
        score += 15
        rules.append("Adet sapması")

    if adet_pct is not None and abs(adet_pct) >= 50:
        score += 15
        rules.append("Adet %50+ değişim")
    elif adet_pct is not None and abs(adet_pct) >= 20:
        score += 8
        rules.append("Adet %20+ değişim")

    if tutar_pct is not None and abs(tutar_pct) >= 50:
        score += 15
        rules.append("Tutar %50+ değişim")
    elif tutar_pct is not None and abs(tutar_pct) >= 30:
        score += 8
        rules.append("Tutar %30+ değişim")

    if has_eba and not has_zfi:
        score += 15
        rules.append("EBA var / ZFI052 yok")

    if "Yeni" in note:
        score += 5
        rules.append("Yeni kayıt")

    if score >= 60:
        status = "Kritik"
    elif score >= 35:
        status = "Yüksek"
    elif score >= 15:
        status = "Orta"
    else:
        status = "Normal"

    if not rules:
        rules = ["Stabil akış"]

    return int(score), status, " / ".join(rules)


def build_regular_risk_status(note_text: str, eba_count: int, zfi_muhatap: str, adet_change_text: str, tutar_change_text: str) -> str:
    return evaluate_regular_risk(note_text, eba_count, zfi_muhatap, adet_change_text, tutar_change_text)[1]


def build_regular_note(period_counts: list, period_amounts: list, eba_count: int, current_label: str, previous_label: str) -> str:
    notes = []
    if not period_counts:
        return ""
    current_count = safe_int(period_counts[-1])
    current_amount = safe_float(period_amounts[-1])
    prev_count = safe_int(period_counts[-2]) if len(period_counts) >= 2 else 0
    prev_amount = safe_float(period_amounts[-2]) if len(period_amounts) >= 2 else 0.0
    prior_counts = [safe_int(x) for x in period_counts[:-1]]
    prior_amounts = [safe_float(x) for x in period_amounts[:-1]]
    active_prior = [x for x in prior_counts if x > 0]
    last_three = prior_counts[-3:] if len(prior_counts) >= 3 else prior_counts

    if prior_counts and len(active_prior) == len(prior_counts) and current_count == 0 and safe_int(eba_count) == 0:
        notes.append("Son Dönem Ft.Yok")
    if last_three and all(x > 0 for x in last_three) and current_count == 0:
        notes.append("Son 3 döneme göre son dönem bekleniyordu")
    if not active_prior and current_count > 0:
        notes.append("Yeni")
    if prior_counts and 0 < len(active_prior) < len(prior_counts):
        if current_count == 0:
            notes.append("Aralıklı Gelen - Son Dönem Yok")
        else:
            notes.append("Aralıklı Gelen")
    if prev_count > current_count:
        notes.append(f"Beklenen {prev_count - current_count} adet eksik")
    elif current_count > prev_count and len(period_counts) >= 2 and prev_count > 0:
        notes.append(f"{current_label} adet artışı var")
    if abs(prev_amount) > 1e-12 and abs(current_amount - prev_amount) > 1.0:
        diff = current_amount - prev_amount
        if diff < 0:
            notes.append(f"Tutar {format_number(abs(diff))} azaldı")
        else:
            notes.append(f"Tutar {format_number(abs(diff))} arttı")
    if current_count == 0 and safe_int(eba_count) > 0:
        notes.append("EBA'da kayıt var")
    if not prior_counts and current_count == 0 and safe_int(eba_count) > 0:
        notes.append("Sadece EBA kaydı var")
    if prior_amounts and current_amount == 0 and any(abs(x) > 1e-12 for x in prior_amounts) and safe_int(eba_count) == 0:
        notes.append("Tutar akışı son dönemde kesildi")
    return join_notes(notes)

class TbPlCcControlWindow(QMainWindow):
    def __init__(self, current_user: str = "admin", language: str = "tr"):
        super().__init__()

        self.language = language if language in ["tr", "en"] else "tr"
        self.current_user = current_user
        self.user_records = load_users()
        self.active_worker: Optional[BackgroundWorker] = None
        self.busy_started_at: Optional[float] = None
        self.busy_last_message: str = ""
        self.busy_timer = QTimer(self)
        self.busy_timer.setInterval(250)
        self.busy_timer.timeout.connect(self.refresh_busy_elapsed)
        self.active_view = "tb"
        self.analysis_filter = "all"
        self.row_density = "normal"
        self.notes_row_density = "normal"
        self.tb_financial_filter = "all"
        self.analysis_search_mode = "contains"

        self.tb_file_path = ""
        self.plcc_file_path = ""

        self.tb_raw_df: Optional[pd.DataFrame] = None
        self.plcc_raw_df: Optional[pd.DataFrame] = None

        self.notes = load_notes()
        self.responsibles = load_responsibles()

        self.muavin_file_path = ""
        self.muavin_file_paths: List[str] = []
        self.muavin_available_columns: List[str] = []
        self.muavin_raw_df: Optional[pd.DataFrame] = None
        self.muavin_clean_df: Optional[pd.DataFrame] = None
        self.muavin_accounts: List[str] = []
        self.muavin_periods: List[str] = []
        self.muavin_doc_types: List[str] = []
        self.muavin_users: List[str] = []
        self.muavin_contras: List[str] = []
        self.muavin_cost_centers: List[str] = []
        self.muavin_selected_account = "Tümü"
        self.muavin_statement_filter = "Tümü"
        self.muavin_period_filter = "Tümü"
        self.muavin_doc_type_filter = "Tümü"
        self.muavin_user_filter = "Tümü"
        self.muavin_contra_filter = "Tümü"
        self.muavin_cost_filter = "Tümü"
        self.muavin_search_text = ""
        self.muavin_search_mode = "contains"
        self.muavin_risk_only = False
        self.muavin_cost_alarm_only = False
        self.muavin_selected_document = ""
        self.muavin_selected_reference = ""
        self.muavin_row_density = "normal"
        self.muavin_analysis_ready = False
        self.muavin_current_view = "user_based"
        self.muavin_selected_contra_name = ""
        self.muavin_column_mapping: Dict[str, str] = {}
        self.muavin_mapping_combos = {}
        self.muavin_section_states = {
            "finding": True,
            "text": True,
            "period": True,
            "doctype": True,
            "risk_user": True,
            "user": True,
            "contra": True,
            "drilldown": True,
            "document_lines": True,
            "account_doc_relation": True,
            "late7": True,
            "dupref": True,
            "duprefdetail": True,
            "doccheck": True,
            "taxref": True,
            "taxvendor": True,
            "tax": True,
            "cost": True,
            "docflow": True,
            "docmatrix": True,
            "userflow": True,
            "vendor_doc_relation": True,
        }

        self.regular_ft_faggl_path = ""
        self.regular_ft_eba_path = ""
        self.regular_ft_zfi052_path = ""
        self.regular_ft_faggl_df: Optional[pd.DataFrame] = None
        self.regular_ft_eba_df: Optional[pd.DataFrame] = None
        self.regular_ft_zfi052_df: Optional[pd.DataFrame] = None
        self.regular_ft_output_df: pd.DataFrame = pd.DataFrame()
        self.regular_ft_periods: List[str] = []
        self.regular_ft_current_period = ""
        self.regular_ft_previous_period = ""
        self.regular_ft_currency_filter = "Tümü"
        self.regular_ft_user_filter = "Tümü"
        self.regular_ft_vendor_filter = "Tümü"
        self.regular_ft_eba_status_filter = "Tümü"
        self.regular_ft_search_text = ""
        self.regular_ft_search_mode = "contains"
        self.regular_ft_risk_only = False
        self.regular_ft_row_density = "normal"
        self.regular_ft_analysis_ready = False
        self.regular_ft_base_output_df: pd.DataFrame = pd.DataFrame()
        self.regular_ft_user_period_map: Dict[Tuple[str, str, str], str] = {}
        self.regular_ft_eba_pending_period_dict: Dict[Tuple[str, str], Dict] = {}
        self.regular_ft_eba_invoice_period_dict: Dict[Tuple[str, str], List[str]] = {}
        self.regular_ft_faggl_invoice_keys_map: Dict[Tuple[str, str], List[str]] = {}
        self.regular_ft_zfi_invoice_map: Dict[Tuple[str, str], str] = {}
        self.regular_ft_zfi_period_map: Dict[Tuple[str, str], str] = {}
        self.regular_ft_zfi_vendor_map: Dict[str, str] = {}

        self.available_periods: List[str] = ["02-Mart", "03-Nisan"]
        self.current_period = "03-Nisan"
        self.previous_period = "02-Mart"

        self.tb_rows_cache: List[Dict] = []
        self.plcc_detail_cache: List[Dict] = []
        self.plcc_subtotal_cache: List[Dict] = []
        self.analysis_has_run = False
        self.is_busy = False

        self.setWindowTitle("TB & PL-CC Control")
        self.resize(1560, 940)
        self.setMinimumSize(1360, 840)

        self.apply_styles()
        self.build_ui()
        self.refresh_all()

    def t(self, key: str) -> str:
        return TRANSLATIONS[self.language].get(key, key)

    def regular_ft_all_label(self) -> str:
        return self.t("all")

    def regular_ft_period_display(self, period_label: str) -> str:
        s = str(period_label).strip()
        if self.language != "en" or not s:
            return s
        parts = s.split()
        if len(parts) < 2:
            return s
        month_name = " ".join(parts[:-1]).strip()
        year = parts[-1]
        return f"{FISCAL_MONTHS_EN.get(month_name, month_name)} {year}"

    def regular_ft_column_label(self, column_name: str) -> str:
        base_map = {
            "Satıcı": self.t("vendorCode"),
            "Satıcı Adı": self.t("vendorName"),
            "Döviz Türü": self.t("currency"),
            "Kullanıcı Adı": self.t("userName"),
            "EBA Adet": "EBA Count",
            "EBA Tutar": "EBA Amount",
            "EBA Son Durum": "EBA Status",
            "EBA Kural": "EBA Rule" if self.language == "en" else "EBA Kural",
            "ZFI052 Muhatap": "ZFI052 Responsible" if self.language == "en" else "ZFI052 Muhatap",
            "ZFI052 Kural": "ZFI052 Rule" if self.language == "en" else "ZFI052 Kural",
            "Adet % Değişim": "Count % Change" if self.language == "en" else "Adet % Değişim",
            "Tutar % Değişim": "Amount % Change" if self.language == "en" else "Tutar % Değişim",
            "Analiz Notu": "Analysis Note" if self.language == "en" else "Analiz Notu",
            "Risk Skoru": "Risk Score" if self.language == "en" else "Risk Skoru",
            "Risk Durumu": "Risk Level" if self.language == "en" else "Risk Durumu",
            "Risk Kuralı": "Risk Rule" if self.language == "en" else "Risk Kuralı",
            "Bilgi": "Info" if self.language == "en" else "Bilgi",
        }
        if column_name in base_map:
            return base_map[column_name]
        m = re.match(r"^(.*) (Adet|Tutar)$", str(column_name).strip())
        if m:
            p = self.regular_ft_period_display(m.group(1))
            suffix = "Count" if m.group(2) == "Adet" and self.language == "en" else "Amount" if m.group(2) == "Tutar" and self.language == "en" else m.group(2)
            return f"{p} {suffix}"
        return column_name

    def period_display_label(self, period_tr: str) -> str:
        if period_tr == OPENING_PERIOD_TR:
            return OPENING_PERIOD_EN if self.language == "en" else OPENING_PERIOD_TR
        found = next((x for x in ALL_PERIODS if x["tr"] == period_tr), None)
        if found:
            return found[self.language]
        return period_tr

    def apply_styles(self):
        self.setStyleSheet("""
        QMainWindow { background: #F3F7FC; }
        QWidget { font-family: "Segoe UI"; color: #0F172A; font-size: 13px; }

        #headerCard {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #0F172A, stop:0.5 #102A5C, stop:1 #1D4ED8);
            border: 1px solid #D9E3F0;
            border-radius: 28px;
        }
        #topBadge {
            background: rgba(255,255,255,0.12);
            border: 1px solid rgba(255,255,255,0.18);
            border-radius: 14px;
            color: white;
            padding: 6px 12px;
            font-size: 11px;
            font-weight: 600;
        }
        #headerTitle { color: white; font-size: 30px; font-weight: 700; }
        #headerDesc { color: #DCE7FF; font-size: 13px; }

        #langButton, #langButtonActive {
            border-radius: 12px;
            padding: 8px 14px;
            font-weight: 700;
            min-width: 54px;
        }
        #langButton {
            background: transparent;
            border: 1px solid rgba(255,255,255,0.15);
            color: white;
        }
        #langButtonActive {
            background: white;
            border: 1px solid white;
            color: #0F172A;
        }

        #primaryButton {
            background: white;
            color: #0F172A;
            border: 1px solid white;
            border-radius: 16px;
            padding: 10px 18px;
            font-weight: 700;
        }
        #primaryButton:hover {
            background: #F8FAFC;
        }
        #primaryButton:pressed {
            background: #E2E8F0;
            border: 1px solid #CBD5E1;
            padding-top: 11px;
            padding-left: 19px;
            padding-right: 17px;
            padding-bottom: 9px;
        }

        #summaryCard, #sidebarCard, #contentCard, #miniCard {
            background: rgba(255,255,255,0.96);
            border: 1px solid #E2E8F0;
            border-radius: 24px;
        }

        #miniCard { border-radius: 18px; }
        #summaryTitle, #miniCardTitle { color: #64748B; font-size: 12px; }
        #summaryValue { color: #0F172A; font-size: 16px; font-weight: 700; }
        #sectionTitle { color: #0F172A; font-size: 20px; font-weight: 700; }
        #sectionDesc, #softInfo { color: #64748B; font-size: 12px; }
        #fieldLabel { color: #334155; font-size: 12px; font-weight: 600; }

        #pillDefault, #pillWarn, #pillOk, #pillInfo, #dashboardPillOk, #dashboardPillInfo {
            border-radius: 12px;
            padding: 4px 8px;
            font-size: 11px;
            font-weight: 600;
            min-height: 28px;
            max-height: 34px;
        }
        #pillDefault { background: #F8FAFC; border: 1px solid #CBD5E1; color: #334155; }
        #pillWarn { background: #FFFBEB; border: 1px solid #FCD34D; color: #B45309; }
        #pillOk { background: #ECFDF5; border: 1px solid #A7F3D0; color: #047857; }
        #pillInfo { background: #EFF6FF; border: 1px solid #BFDBFE; color: #1D4ED8; }
        #dashboardPillOk { background: #ECFDF5; border: 1px solid #A7F3D0; color: #047857; }
        #dashboardPillInfo { background: #EFF6FF; border: 1px solid #BFDBFE; color: #1D4ED8; }

        QTabWidget::pane { border: none; }
        QTabBar::tab {
            background: white;
            border: 1px solid #CBD5E1;
            padding: 11px 18px;
            border-radius: 16px;
            margin-right: 8px;
            font-weight: 700;
            min-width: 130px;
        }
        QTabBar::tab:selected {
            background: #1D4ED8;
            color: white;
            border: 1px solid #1D4ED8;
        }

        QLineEdit, QComboBox, QTextEdit {
            background: white;
            border: 1px solid #CBD5E1;
            border-radius: 14px;
            padding: 8px 10px;
        }

        #secondaryButton, #badgeButton, #toggleButton, #toggleButtonActive, #filterSoft, #filterSoftActive {
            border-radius: 14px;
            padding: 10px 14px;
            font-weight: 700;
        }
        #secondaryButton {
            background: white;
            border: 1px solid #CBD5E1;
        }
        #badgeButton {
            background: #EFF6FF;
            border: 1px solid #BFDBFE;
            color: #1D4ED8;
            padding: 6px 12px;
            font-size: 11px;
        }
        #toggleButton {
            background: white;
            border: 1px solid #CBD5E1;
            color: #334155;
        }
        #toggleButtonActive {
            background: #1D4ED8;
            border: 1px solid #1D4ED8;
            color: white;
        }
        #filterSoft {
            background: white;
            border: 1px solid #CBD5E1;
            color: #334155;
        }
        #filterSoftActive {
            background: #EFF6FF;
            border: 1px solid #BFDBFE;
            color: #1D4ED8;
        }

        QProgressBar {
            background: #E2E8F0;
            border: none;
            border-radius: 6px;
            height: 12px;
        }
        QProgressBar::chunk {
            background: #1D4ED8;
            border-radius: 6px;
        }

        QTableWidget {
            background: white;
            border: 1px solid #CBD5E1;
            border-radius: 18px;
            gridline-color: #E2E8F0;
            selection-background-color: #DBEAFE;
            selection-color: #0F172A;
        }

        QHeaderView::section {
            background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #1D4ED8, stop:1 #2563EB);
            color: white;
            border: none;
            padding: 10px 8px;
            font-size: 12px;
            font-weight: 700;
        }
        """)

    def build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        outer = QVBoxLayout(central)
        outer.setContentsMargins(22, 18, 22, 18)
        outer.setSpacing(18)

        outer.addWidget(self.build_header())

        bottom = QHBoxLayout()
        bottom.setSpacing(18)
        outer.addLayout(bottom, 1)

        bottom.addWidget(self.build_sidebar(), 0)

        self.tabs = QTabWidget()
        bottom.addWidget(self.tabs, 1)

        self.dashboard_tab = self.build_dashboard_tab()
        self.analysis_tab = self.build_analysis_tab()
        self.notes_tab = self.build_notes_tab()
        self.responsibles_tab = self.build_responsibles_tab()
        self.muavin_tab = self.build_muavin_tab()
        self.regular_ft_tab = self.build_regular_ft_tab()
        self.users_tab = self.build_users_tab()

        self.tabs.addTab(self.dashboard_tab, "")
        self.tabs.addTab(self.analysis_tab, "")
        self.tabs.addTab(self.notes_tab, "")
        self.tabs.addTab(self.responsibles_tab, "")
        self.tabs.addTab(self.muavin_tab, "")
        self.tabs.addTab(self.regular_ft_tab, "")
        self.tabs.addTab(self.users_tab, "")

    def build_header(self):
        card = QFrame()
        card.setObjectName("headerCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(24, 22, 24, 22)
        layout.setSpacing(18)

        top = QHBoxLayout()
        layout.addLayout(top)

        left = QVBoxLayout()
        top.addLayout(left, 1)

        self.header_badge = QLabel()
        self.header_badge.setObjectName("topBadge")
        left.addWidget(self.header_badge, 0, Qt.AlignLeft)

        self.header_title = QLabel()
        self.header_title.setObjectName("headerTitle")
        left.addWidget(self.header_title)

        self.header_desc = QLabel()
        self.header_desc.setObjectName("headerDesc")
        self.header_desc.setWordWrap(True)
        left.addWidget(self.header_desc)

        right = QVBoxLayout()
        top.addLayout(right)

        lang_row = QHBoxLayout()
        right.addLayout(lang_row)

        self.btn_tr = QPushButton("TR")
        self.btn_en = QPushButton("EN")
        self.btn_tr.clicked.connect(lambda: self.set_language("tr"))
        self.btn_en.clicked.connect(lambda: self.set_language("en"))
        lang_row.addWidget(self.btn_tr)
        lang_row.addWidget(self.btn_en)

        self.start_button = QPushButton()
        self.start_button.setObjectName("primaryButton")
        self.start_button.clicked.connect(self.start_analysis)
        right.addWidget(self.start_button)

        summary = QHBoxLayout()
        summary.setSpacing(12)
        layout.addLayout(summary)

        self.summary_cards = []
        for _ in range(4):
            frame = QFrame()
            frame.setObjectName("summaryCard")
            fl = QVBoxLayout(frame)
            fl.setContentsMargins(18, 16, 18, 16)
            fl.setSpacing(8)

            title = QLabel()
            title.setObjectName("summaryTitle")
            value = QLabel()
            value.setObjectName("summaryValue")
            fl.addWidget(title)
            fl.addWidget(value)

            btn_row = QHBoxLayout()
            fl.addLayout(btn_row)

            btn1 = QPushButton()
            btn1.setObjectName("badgeButton")
            btn2 = QPushButton()
            btn2.setObjectName("badgeButton")
            btn1.hide()
            btn2.hide()
            btn_row.addWidget(btn1)
            btn_row.addWidget(btn2)

            self.summary_cards.append({
                "title": title,
                "value": value,
                "btn1": btn1,
                "btn2": btn2,
            })
            summary.addWidget(frame)

        self.summary_cards[3]["btn1"].clicked.connect(self.jump_to_tb_findings)
        self.summary_cards[3]["btn2"].clicked.connect(self.jump_to_plcc_findings)
        return card

    def build_sidebar(self):
        card = QFrame()
        card.setObjectName("sidebarCard")
        card.setFixedWidth(310)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(14)

        self.sidebar_title = QLabel()
        self.sidebar_title.setObjectName("sectionTitle")
        layout.addWidget(self.sidebar_title)

        self.sidebar_desc = QLabel()
        self.sidebar_desc.setObjectName("sectionDesc")
        self.sidebar_desc.setWordWrap(True)
        layout.addWidget(self.sidebar_desc)

        self.btn_upload_tb = QPushButton()
        self.btn_upload_tb.setObjectName("secondaryButton")
        self.btn_upload_tb.clicked.connect(self.load_tb_file)

        self.btn_upload_plcc = QPushButton()
        self.btn_upload_plcc.setObjectName("secondaryButton")
        self.btn_upload_plcc.clicked.connect(self.load_plcc_file)

        self.btn_note_defs = QPushButton()
        self.btn_note_defs.setObjectName("secondaryButton")
        self.btn_note_defs.clicked.connect(lambda: self.tabs.setCurrentIndex(2))

        self.btn_responsibles = QPushButton()
        self.btn_responsibles.setObjectName("secondaryButton")
        self.btn_responsibles.clicked.connect(lambda: self.tabs.setCurrentIndex(3))

        self.btn_muavin = QPushButton()
        self.btn_muavin.setObjectName("secondaryButton")
        self.btn_muavin.clicked.connect(lambda: self.tabs.setCurrentIndex(4))

        self.btn_regular_ft = QPushButton()
        self.btn_regular_ft.setObjectName("secondaryButton")
        self.btn_regular_ft.clicked.connect(lambda: self.tabs.setCurrentIndex(5))

        layout.addWidget(self.btn_upload_tb)
        layout.addWidget(self.btn_upload_plcc)
        layout.addWidget(self.btn_note_defs)
        layout.addWidget(self.btn_responsibles)
        layout.addWidget(self.btn_muavin)
        layout.addWidget(self.btn_regular_ft)

        proc_box = QFrame()
        proc_box.setObjectName("miniCard")
        proc_layout = QVBoxLayout(proc_box)
        proc_layout.setContentsMargins(14, 14, 14, 14)

        top = QHBoxLayout()
        proc_layout.addLayout(top)

        self.proc_label = QLabel()
        self.proc_label.setObjectName("miniCardTitle")
        top.addWidget(self.proc_label)
        top.addStretch()

        self.proc_value = QLabel("0%")
        self.proc_value.setObjectName("summaryValue")
        top.addWidget(self.proc_value)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        proc_layout.addWidget(self.progress)

        self.proc_text = QLabel()
        self.proc_text.setObjectName("softInfo")
        self.proc_text.setWordWrap(True)
        proc_layout.addWidget(self.proc_text)

        layout.addWidget(proc_box)
        layout.addStretch()
        return card

    def build_dashboard_tab(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        card = QFrame()
        card.setObjectName("contentCard")
        layout.addWidget(card)

        main = QVBoxLayout(card)
        main.setContentsMargins(18, 16, 18, 16)
        main.setSpacing(10)

        self.dashboard_title = QLabel()
        self.dashboard_title.setObjectName("sectionTitle")
        main.addWidget(self.dashboard_title)

        self.dashboard_desc = QLabel()
        self.dashboard_desc.setObjectName("sectionDesc")
        self.dashboard_desc.setWordWrap(True)
        self.dashboard_desc.setMaximumHeight(36)
        main.addWidget(self.dashboard_desc)

        form = QGridLayout()
        form.setContentsMargins(0, 2, 0, 2)
        form.setHorizontalSpacing(10)
        form.setVerticalSpacing(6)
        main.addLayout(form)

        self.lbl_tb_file = QLabel()
        self.lbl_tb_file.setObjectName("fieldLabel")
        self.lbl_plcc_file = QLabel()
        self.lbl_plcc_file.setObjectName("fieldLabel")
        self.lbl_curr_period = QLabel()
        self.lbl_curr_period.setObjectName("fieldLabel")
        self.lbl_prev_period = QLabel()
        self.lbl_prev_period.setObjectName("fieldLabel")

        self.tb_file_input = QLineEdit()
        self.tb_file_input.setReadOnly(True)
        self.tb_file_input.setMinimumHeight(40)
        self.tb_file_input.setMaximumHeight(40)

        self.plcc_file_input = QLineEdit()
        self.plcc_file_input.setReadOnly(True)
        self.plcc_file_input.setMinimumHeight(40)
        self.plcc_file_input.setMaximumHeight(40)

        self.cmb_current = QComboBox()
        self.cmb_previous = QComboBox()
        self.cmb_current.setMinimumHeight(40)
        self.cmb_current.setMaximumHeight(40)
        self.cmb_previous.setMinimumHeight(40)
        self.cmb_previous.setMaximumHeight(40)

        self.cmb_current.currentIndexChanged.connect(self.on_period_changed)
        self.cmb_previous.currentIndexChanged.connect(self.on_period_changed)

        form.addWidget(self.lbl_tb_file, 0, 0)
        form.addWidget(self.lbl_plcc_file, 0, 1)
        form.addWidget(self.tb_file_input, 1, 0)
        form.addWidget(self.plcc_file_input, 1, 1)
        form.addWidget(self.lbl_curr_period, 2, 0)
        form.addWidget(self.lbl_prev_period, 2, 1)
        form.addWidget(self.cmb_current, 3, 0)
        form.addWidget(self.cmb_previous, 3, 1)

        pill_row = QHBoxLayout()
        pill_row.setContentsMargins(0, 4, 0, 0)
        pill_row.setSpacing(6)
        main.addLayout(pill_row)

        self.dashboard_pills = []
        for obj in ["pillOk", "pillOk", "pillInfo"]:
            lbl = QLabel()
            lbl.setObjectName(obj)
            self.dashboard_pills.append(lbl)
            pill_row.addWidget(lbl)

        pill_row.addStretch()

        metrics_row = QHBoxLayout()
        metrics_row.setSpacing(10)
        main.addLayout(metrics_row)

        self.dashboard_metric_cards = []
        for _ in range(3):
            frame = QFrame()
            frame.setObjectName("miniCard")
            fl = QVBoxLayout(frame)
            fl.setContentsMargins(14, 14, 14, 14)
            fl.setSpacing(6)
            title = QLabel()
            title.setObjectName("miniCardTitle")
            value = QLabel()
            value.setObjectName("summaryValue")
            period = QLabel()
            period.setObjectName("softInfo")
            fl.addWidget(title)
            fl.addWidget(value)
            fl.addWidget(period)
            metrics_row.addWidget(frame)
            self.dashboard_metric_cards.append({"title": title, "value": value, "period": period})

        main.addStretch()

        return page

    def build_analysis_tab(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)

        card = QFrame()
        card.setObjectName("contentCard")
        layout.addWidget(card)

        main = QVBoxLayout(card)
        main.setContentsMargins(22, 22, 22, 22)
        main.setSpacing(14)

        self.analysis_title = QLabel()
        self.analysis_title.setObjectName("sectionTitle")
        main.addWidget(self.analysis_title)

        self.analysis_desc = QLabel()
        self.analysis_desc.setObjectName("sectionDesc")
        self.analysis_desc.setWordWrap(True)
        main.addWidget(self.analysis_desc)

        row1 = QHBoxLayout()
        main.addLayout(row1)

        self.search_input = QLineEdit()
        self.search_input.textChanged.connect(self.refresh_analysis_tables)
        row1.addWidget(self.search_input, 1)

        self.cmb_analysis_search_mode = QComboBox()
        self.cmb_analysis_search_mode.currentIndexChanged.connect(self.on_analysis_search_mode_changed)
        self.cmb_analysis_search_mode.setMinimumHeight(40)
        self.cmb_analysis_search_mode.setMinimumWidth(170)
        row1.addWidget(self.cmb_analysis_search_mode)

        self.period_info = QLabel()
        self.period_info.setObjectName("pillDefault")
        row1.addWidget(self.period_info)

        self.cmb_density = QComboBox()
        self.cmb_density.currentIndexChanged.connect(self.on_density_changed)
        row1.addWidget(self.cmb_density)

        self.btn_export = QPushButton()
        self.btn_export.setObjectName("secondaryButton")
        self.btn_export.clicked.connect(self.export_current_view)
        row1.addWidget(self.btn_export)

        self.btn_auto_fit = QPushButton()
        self.btn_auto_fit.setObjectName("secondaryButton")
        self.btn_auto_fit.clicked.connect(self.auto_fit_active_table)
        row1.addWidget(self.btn_auto_fit)

        row2 = QHBoxLayout()
        main.addLayout(row2)

        self.btn_view_tb = QPushButton()
        self.btn_view_tb.clicked.connect(lambda: self.set_active_view("tb"))
        self.btn_view_plcc = QPushButton()
        self.btn_view_plcc.clicked.connect(lambda: self.set_active_view("plcc"))
        self.btn_show_all = QPushButton()
        self.btn_show_all.clicked.connect(lambda: self.set_analysis_filter("all"))
        self.btn_findings = QPushButton()
        self.btn_findings.clicked.connect(lambda: self.set_analysis_filter("findings"))

        self.btn_tb_all = QPushButton()
        self.btn_tb_all.clicked.connect(lambda: self.set_tb_financial_filter("all"))
        self.btn_tb_balance = QPushButton()
        self.btn_tb_balance.clicked.connect(lambda: self.set_tb_financial_filter("balance"))
        self.btn_tb_income = QPushButton()
        self.btn_tb_income.clicked.connect(lambda: self.set_tb_financial_filter("income"))

        row2.addWidget(self.btn_view_tb)
        row2.addWidget(self.btn_view_plcc)
        row2.addWidget(self.btn_show_all)
        row2.addWidget(self.btn_findings)
        row2.addWidget(self.btn_tb_all)
        row2.addWidget(self.btn_tb_balance)
        row2.addWidget(self.btn_tb_income)
        row2.addStretch()

        self.toggle_info = QLabel()
        self.toggle_info.setObjectName("softInfo")
        row2.addWidget(self.toggle_info)

        self.tb_table = QTableWidget()
        self.plcc_table = QTableWidget()
        self.setup_table(self.tb_table)
        self.setup_table(self.plcc_table)

        main.addWidget(self.tb_table, 1)
        main.addWidget(self.plcc_table, 1)

        return page



    def build_regular_ft_tab(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        page_layout.addWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)

        outer = QVBoxLayout(container)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        card = QFrame()
        card.setObjectName("contentCard")
        outer.addWidget(card)

        main = QVBoxLayout(card)
        main.setContentsMargins(22, 22, 22, 22)
        main.setSpacing(14)

        self.regular_ft_title = QLabel("Düzenli Gelen Ft Analiz")
        self.regular_ft_title.setObjectName("sectionTitle")
        main.addWidget(self.regular_ft_title)

        self.regular_ft_desc = QLabel(
            "FAGGL, EBA ve ZFI052 raporlarını birlikte analiz ederek satıcı bazında dönemsel fatura adet/tutar, son dönem EBA durumu ve ZFI052 muhatap bilgisini üretir."
        )
        self.regular_ft_desc.setObjectName("sectionDesc")
        self.regular_ft_desc.setWordWrap(True)
        main.addWidget(self.regular_ft_desc)

        top = QHBoxLayout()
        main.addLayout(top)

        self.btn_regular_ft_load_faggl = QPushButton("FAGGL Yükle")
        self.btn_regular_ft_load_faggl.setObjectName("secondaryButton")
        self.btn_regular_ft_load_faggl.clicked.connect(self.load_regular_ft_faggl)
        top.addWidget(self.btn_regular_ft_load_faggl)

        self.btn_regular_ft_load_eba = QPushButton("EBA Yükle")
        self.btn_regular_ft_load_eba.setObjectName("secondaryButton")
        self.btn_regular_ft_load_eba.clicked.connect(self.load_regular_ft_eba)
        top.addWidget(self.btn_regular_ft_load_eba)

        self.btn_regular_ft_load_zfi052 = QPushButton("ZFI052 Yükle")
        self.btn_regular_ft_load_zfi052.setObjectName("secondaryButton")
        self.btn_regular_ft_load_zfi052.clicked.connect(self.load_regular_ft_zfi052)
        top.addWidget(self.btn_regular_ft_load_zfi052)

        self.btn_regular_ft_start = QPushButton("Başlat")
        self.btn_regular_ft_start.setObjectName("primaryButton")
        self.btn_regular_ft_start.clicked.connect(self.start_regular_ft_analysis)
        top.addWidget(self.btn_regular_ft_start)

        self.regular_ft_info_pill = QLabel("-")
        self.regular_ft_info_pill.setObjectName("pillInfo")
        top.addWidget(self.regular_ft_info_pill)
        top.addStretch()

        file_grid = QGridLayout()
        file_grid.setHorizontalSpacing(10)
        file_grid.setVerticalSpacing(8)
        main.addLayout(file_grid)

        self.regular_ft_faggl_path_input = QLineEdit()
        self.regular_ft_faggl_path_input.setReadOnly(True)
        self.regular_ft_eba_path_input = QLineEdit()
        self.regular_ft_eba_path_input.setReadOnly(True)
        self.regular_ft_zfi052_path_input = QLineEdit()
        self.regular_ft_zfi052_path_input.setReadOnly(True)

        self.lbl_regular_ft_faggl = QLabel("FAGGL")
        self.lbl_regular_ft_eba = QLabel("EBA")
        self.lbl_regular_ft_zfi052 = QLabel("ZFI052")
        file_grid.addWidget(self.lbl_regular_ft_faggl, 0, 0)
        file_grid.addWidget(self.lbl_regular_ft_eba, 0, 1)
        file_grid.addWidget(self.lbl_regular_ft_zfi052, 0, 2)
        file_grid.addWidget(self.regular_ft_faggl_path_input, 1, 0)
        file_grid.addWidget(self.regular_ft_eba_path_input, 1, 1)
        file_grid.addWidget(self.regular_ft_zfi052_path_input, 1, 2)

        filter_row = QHBoxLayout()
        main.addLayout(filter_row)

        self.btn_regular_ft_risk_only = QPushButton("Sadece Riskli Kayıtlar")
        self.btn_regular_ft_risk_only.setCheckable(True)
        self.btn_regular_ft_risk_only.clicked.connect(self.toggle_regular_ft_risk_only)
        filter_row.addWidget(self.btn_regular_ft_risk_only)

        self.btn_regular_ft_autofit = QPushButton("Kolonları Otomatik Sığdır")
        self.btn_regular_ft_autofit.setObjectName("secondaryButton")
        self.btn_regular_ft_autofit.clicked.connect(self.auto_fit_regular_ft_table)
        filter_row.addWidget(self.btn_regular_ft_autofit)

        self.cmb_regular_ft_density = QComboBox()
        self.cmb_regular_ft_density.currentIndexChanged.connect(self.on_regular_ft_density_changed)
        filter_row.addWidget(self.cmb_regular_ft_density)

        self.btn_regular_ft_export = QPushButton("Excel Çıktısı")
        self.btn_regular_ft_export.setObjectName("secondaryButton")
        self.btn_regular_ft_export.clicked.connect(self.export_regular_ft_analysis)
        filter_row.addWidget(self.btn_regular_ft_export)
        filter_row.addStretch()

        filters = QGridLayout()
        filters.setHorizontalSpacing(10)
        filters.setVerticalSpacing(8)
        main.addLayout(filters)

        self.cmb_regular_ft_currency = QComboBox()
        self.cmb_regular_ft_currency.currentIndexChanged.connect(self.refresh_regular_ft_table)
        self.cmb_regular_ft_user = QComboBox()
        self.cmb_regular_ft_user.currentIndexChanged.connect(self.refresh_regular_ft_table)
        self.cmb_regular_ft_vendor = QComboBox()
        self.cmb_regular_ft_vendor.currentIndexChanged.connect(self.refresh_regular_ft_table)
        self.cmb_regular_ft_eba_status = QComboBox()
        self.cmb_regular_ft_eba_status.currentIndexChanged.connect(self.refresh_regular_ft_table)
        self.input_regular_ft_search = QLineEdit()
        self.input_regular_ft_search.textChanged.connect(self.refresh_regular_ft_table)

        self.cmb_regular_ft_current_period = QComboBox()
        self.cmb_regular_ft_current_period.currentIndexChanged.connect(self.on_regular_ft_period_compare_changed)
        self.cmb_regular_ft_previous_period = QComboBox()
        self.cmb_regular_ft_previous_period.currentIndexChanged.connect(self.on_regular_ft_period_compare_changed)

        self.lbl_regular_ft_currency = QLabel("Döviz Türü")
        self.lbl_regular_ft_user = QLabel("Kullanıcı Adı")
        self.lbl_regular_ft_vendor = QLabel("Satıcı Adı")
        self.lbl_regular_ft_eba_status = QLabel("EBA Durum")
        self.lbl_regular_ft_search = QLabel("Çıktıda Ara")
        self.lbl_regular_ft_search_mode = QLabel("Arama Tipi")
        self.lbl_regular_ft_current_period = QLabel("Son Dönem")
        self.lbl_regular_ft_previous_period = QLabel("Karşılaştırma Dönemi")
        filters.addWidget(self.lbl_regular_ft_currency, 0, 0)
        filters.addWidget(self.lbl_regular_ft_user, 0, 1)
        filters.addWidget(self.lbl_regular_ft_vendor, 0, 2)
        filters.addWidget(self.lbl_regular_ft_eba_status, 0, 3)
        filters.addWidget(self.lbl_regular_ft_search, 0, 4)
        filters.addWidget(self.lbl_regular_ft_search_mode, 0, 5)
        filters.addWidget(self.cmb_regular_ft_currency, 1, 0)
        filters.addWidget(self.cmb_regular_ft_user, 1, 1)
        filters.addWidget(self.cmb_regular_ft_vendor, 1, 2)
        filters.addWidget(self.cmb_regular_ft_eba_status, 1, 3)
        filters.addWidget(self.input_regular_ft_search, 1, 4)
        self.cmb_regular_ft_search_mode = QComboBox()
        self.cmb_regular_ft_search_mode.currentIndexChanged.connect(self.on_regular_ft_search_mode_changed)
        self.cmb_regular_ft_search_mode.setMinimumHeight(40)
        filters.addWidget(self.cmb_regular_ft_search_mode, 1, 5)
        filters.addWidget(self.lbl_regular_ft_current_period, 2, 0)
        filters.addWidget(self.lbl_regular_ft_previous_period, 2, 1)
        filters.addWidget(self.cmb_regular_ft_current_period, 3, 0)
        filters.addWidget(self.cmb_regular_ft_previous_period, 3, 1)

        metric_row = QHBoxLayout()
        main.addLayout(metric_row)
        self.regular_ft_metric_cards = []
        for _ in range(4):
            frame = QFrame()
            frame.setObjectName("miniCard")
            fl = QVBoxLayout(frame)
            fl.setContentsMargins(14, 14, 14, 14)
            fl.setSpacing(6)
            title = QLabel("-")
            title.setObjectName("miniCardTitle")
            value = QLabel("0")
            value.setObjectName("summaryValue")
            sub = QLabel("-")
            sub.setObjectName("softInfo")
            sub.setWordWrap(True)
            fl.addWidget(title)
            fl.addWidget(value)
            fl.addWidget(sub)
            metric_row.addWidget(frame)
            self.regular_ft_metric_cards.append({"title": title, "value": value, "sub": sub})

        self.regular_ft_table = QTableWidget()
        self.setup_table(self.regular_ft_table)
        main.addWidget(self.regular_ft_table, 1)

        return page

    def build_notes_tab(self):
        page = QWidget()
        layout = QHBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)

        left = QFrame()
        left.setObjectName("contentCard")
        right = QFrame()
        right.setObjectName("contentCard")
        layout.addWidget(left, 1)
        layout.addWidget(right, 1)

        ll = QVBoxLayout(left)
        ll.setContentsMargins(22, 22, 22, 22)
        ll.setSpacing(12)

        self.notes_title = QLabel()
        self.notes_title.setObjectName("sectionTitle")
        ll.addWidget(self.notes_title)

        self.notes_desc = QLabel()
        self.notes_desc.setObjectName("sectionDesc")
        self.notes_desc.setWordWrap(True)
        ll.addWidget(self.notes_desc)

        self.note_paste_hint = QLabel()
        self.note_paste_hint.setObjectName("softInfo")
        self.note_paste_hint.setWordWrap(True)
        ll.addWidget(self.note_paste_hint)

        self.lbl_note_paste = QLabel()
        self.lbl_note_paste.setObjectName("fieldLabel")
        ll.addWidget(self.lbl_note_paste)

        self.input_note_paste = ExcelPasteTableWidget(
            30,
            column_headers=["Hesap", "Ana Hesap", "Masraf Yeri", "Not"],
            stretch_last=True,
        )
        self.input_note_paste.setMinimumHeight(280)
        ll.addWidget(self.input_note_paste)

        paste_btn_row = QHBoxLayout()
        ll.addLayout(paste_btn_row)

        self.btn_note_bulk_save = QPushButton()
        self.btn_note_bulk_save.setObjectName("secondaryButton")
        self.btn_note_bulk_save.clicked.connect(self.save_notes_from_paste)
        paste_btn_row.addWidget(self.btn_note_bulk_save)

        self.btn_note_paste_clear = QPushButton()
        self.btn_note_paste_clear.setObjectName("secondaryButton")
        self.btn_note_paste_clear.clicked.connect(lambda: self.input_note_paste.clear_data())
        paste_btn_row.addWidget(self.btn_note_paste_clear)
        paste_btn_row.addStretch()

        self.note_match_info = QLabel()
        self.note_match_info.setObjectName("softInfo")
        self.note_match_info.setWordWrap(True)
        ll.addWidget(self.note_match_info)
        ll.addStretch()

        rl = QVBoxLayout(right)
        rl.setContentsMargins(22, 22, 22, 22)

        self.saved_notes_title = QLabel()
        self.saved_notes_title.setObjectName("sectionTitle")
        rl.addWidget(self.saved_notes_title)

        self.saved_notes_desc = QLabel()
        self.saved_notes_desc.setObjectName("sectionDesc")
        self.saved_notes_desc.setWordWrap(True)
        rl.addWidget(self.saved_notes_desc)

        top = QHBoxLayout()
        rl.addLayout(top)

        self.notes_control_info = QLabel()
        self.notes_control_info.setObjectName("softInfo")
        top.addWidget(self.notes_control_info)
        top.addStretch()

        self.btn_notes_save_changes = QPushButton()
        self.btn_notes_save_changes.setObjectName("secondaryButton")
        self.btn_notes_save_changes.clicked.connect(self.save_notes_table_changes)
        top.addWidget(self.btn_notes_save_changes)

        self.btn_notes_delete_selected = QPushButton()
        self.btn_notes_delete_selected.setObjectName("secondaryButton")
        self.btn_notes_delete_selected.clicked.connect(self.delete_selected_note)
        top.addWidget(self.btn_notes_delete_selected)

        self.cmb_notes_density = QComboBox()
        self.cmb_notes_density.currentIndexChanged.connect(self.on_notes_density_changed)
        top.addWidget(self.cmb_notes_density)

        self.notes_table = QTableWidget()
        self.setup_table(self.notes_table)
        self.notes_table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed | QTableWidget.AnyKeyPressed)
        rl.addWidget(self.notes_table, 1)

        return page

    def build_responsibles_tab(self):
        page = QWidget()
        layout = QHBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)

        left = QFrame()
        left.setObjectName("contentCard")
        right = QFrame()
        right.setObjectName("contentCard")
        layout.addWidget(left, 1)
        layout.addWidget(right, 1)

        ll = QVBoxLayout(left)
        ll.setContentsMargins(22, 22, 22, 22)

        self.resp_title = QLabel()
        self.resp_title.setObjectName("sectionTitle")
        ll.addWidget(self.resp_title)

        self.resp_desc = QLabel()
        self.resp_desc.setObjectName("sectionDesc")
        self.resp_desc.setWordWrap(True)
        ll.addWidget(self.resp_desc)

        self.lbl_resp_paste = QLabel()
        self.lbl_resp_paste.setObjectName("fieldLabel")
        ll.addWidget(self.lbl_resp_paste)

        self.input_resp_paste = ExcelPasteTableWidget(
            30,
            column_headers=["Hesap", "Ana Hesap", "Sorumlu"],
            stretch_last=True
        )
        self.input_resp_paste.setMinimumHeight(320)
        ll.addWidget(self.input_resp_paste)

        btn_row = QHBoxLayout()
        ll.addLayout(btn_row)

        self.btn_resp_save = QPushButton()
        self.btn_resp_save.setObjectName("secondaryButton")
        self.btn_resp_save.clicked.connect(self.save_responsibles_from_paste)
        btn_row.addWidget(self.btn_resp_save)

        self.btn_resp_clear = QPushButton()
        self.btn_resp_clear.setObjectName("secondaryButton")
        self.btn_resp_clear.clicked.connect(lambda: self.input_resp_paste.clear_data())
        btn_row.addWidget(self.btn_resp_clear)
        btn_row.addStretch()

        self.resp_format_info = QLabel()
        self.resp_format_info.setObjectName("softInfo")
        self.resp_format_info.setWordWrap(True)
        ll.addWidget(self.resp_format_info)
        ll.addStretch()

        rl = QVBoxLayout(right)
        rl.setContentsMargins(22, 22, 22, 22)

        self.saved_resp_title = QLabel()
        self.saved_resp_title.setObjectName("sectionTitle")
        rl.addWidget(self.saved_resp_title)

        self.saved_resp_desc = QLabel()
        self.saved_resp_desc.setObjectName("sectionDesc")
        self.saved_resp_desc.setWordWrap(True)
        rl.addWidget(self.saved_resp_desc)

        resp_top = QHBoxLayout()
        rl.addLayout(resp_top)
        resp_top.addStretch()

        self.btn_resp_save_changes = QPushButton()
        self.btn_resp_save_changes.setObjectName("secondaryButton")
        self.btn_resp_save_changes.clicked.connect(self.save_responsibles_table_changes)
        resp_top.addWidget(self.btn_resp_save_changes)

        self.btn_resp_delete_selected = QPushButton()
        self.btn_resp_delete_selected.setObjectName("secondaryButton")
        self.btn_resp_delete_selected.clicked.connect(self.delete_selected_responsible)
        resp_top.addWidget(self.btn_resp_delete_selected)

        self.responsibles_table = QTableWidget()
        self.setup_table(self.responsibles_table)
        self.responsibles_table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed | QTableWidget.AnyKeyPressed)
        rl.addWidget(self.responsibles_table, 1)

        return page



    def build_users_tab(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        page_layout.addWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)

        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(16)

        left = QFrame()
        left.setObjectName("contentCard")
        right = QFrame()
        right.setObjectName("contentCard")
        layout.addWidget(left, 3)
        layout.addWidget(right, 2)

        ll = QVBoxLayout(left)
        ll.setContentsMargins(24, 24, 24, 24)
        ll.setSpacing(14)

        self.users_title = QLabel()
        self.users_title.setObjectName("sectionTitle")
        ll.addWidget(self.users_title)

        self.users_desc = QLabel()
        self.users_desc.setObjectName("sectionDesc")
        self.users_desc.setWordWrap(True)
        ll.addWidget(self.users_desc)

        self.users_current_info = QLabel()
        self.users_current_info.setObjectName("pillInfo")
        ll.addWidget(self.users_current_info, 0, Qt.AlignLeft)

        form_card = QFrame()
        form_card.setObjectName("miniCard")
        form_layout = QGridLayout(form_card)
        form_layout.setContentsMargins(16, 16, 16, 16)
        form_layout.setHorizontalSpacing(14)
        form_layout.setVerticalSpacing(10)
        ll.addWidget(form_card)

        self.lbl_user_username = QLabel()
        self.lbl_user_username.setObjectName("fieldLabel")
        form_layout.addWidget(self.lbl_user_username, 0, 0, 1, 2)
        self.input_user_username = QLineEdit()
        self.input_user_username.setMinimumHeight(42)
        form_layout.addWidget(self.input_user_username, 1, 0, 1, 2)

        self.lbl_user_password = QLabel()
        self.lbl_user_password.setObjectName("fieldLabel")
        form_layout.addWidget(self.lbl_user_password, 2, 0)
        self.lbl_user_password_confirm = QLabel()
        self.lbl_user_password_confirm.setObjectName("fieldLabel")
        form_layout.addWidget(self.lbl_user_password_confirm, 2, 1)

        password_left = QHBoxLayout()
        password_left.setSpacing(8)
        self.input_user_password = QLineEdit()
        self.input_user_password.setEchoMode(QLineEdit.Password)
        self.input_user_password.setMinimumHeight(42)
        password_left.addWidget(self.input_user_password, 1)
        self.btn_toggle_user_password = QPushButton("Göster")
        self.btn_toggle_user_password.setObjectName("secondaryButton")
        self.btn_toggle_user_password.setMinimumHeight(42)
        self.btn_toggle_user_password.setMinimumWidth(96)
        self.btn_toggle_user_password.clicked.connect(lambda: self.toggle_password_visibility("user_password"))
        password_left.addWidget(self.btn_toggle_user_password)
        form_layout.addLayout(password_left, 3, 0)

        password_right = QHBoxLayout()
        password_right.setSpacing(8)
        self.input_user_password_confirm = QLineEdit()
        self.input_user_password_confirm.setEchoMode(QLineEdit.Password)
        self.input_user_password_confirm.setMinimumHeight(42)
        password_right.addWidget(self.input_user_password_confirm, 1)
        self.btn_toggle_user_password_confirm = QPushButton("Göster")
        self.btn_toggle_user_password_confirm.setObjectName("secondaryButton")
        self.btn_toggle_user_password_confirm.setMinimumHeight(42)
        self.btn_toggle_user_password_confirm.setMinimumWidth(96)
        self.btn_toggle_user_password_confirm.clicked.connect(lambda: self.toggle_password_visibility("user_password_confirm"))
        password_right.addWidget(self.btn_toggle_user_password_confirm)
        form_layout.addLayout(password_right, 3, 1)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        form_layout.addLayout(btn_row, 4, 0, 1, 2)

        self.btn_user_add = QPushButton()
        self.btn_user_add.setObjectName("secondaryButton")
        self.btn_user_add.clicked.connect(self.add_or_update_user)
        self.btn_user_add.setMinimumHeight(42)
        btn_row.addWidget(self.btn_user_add)

        self.btn_user_clear = QPushButton()
        self.btn_user_clear.setObjectName("secondaryButton")
        self.btn_user_clear.clicked.connect(self.clear_user_form)
        self.btn_user_clear.setMinimumHeight(42)
        btn_row.addWidget(self.btn_user_clear)

        self.btn_user_delete = QPushButton()
        self.btn_user_delete.setObjectName("secondaryButton")
        self.btn_user_delete.clicked.connect(self.delete_selected_user)
        self.btn_user_delete.setMinimumHeight(42)
        btn_row.addWidget(self.btn_user_delete)
        btn_row.addStretch()

        self.users_perm_title = QLabel()
        self.users_perm_title.setObjectName("sectionTitle")
        ll.addWidget(self.users_perm_title)

        self.users_perm_desc = QLabel()
        self.users_perm_desc.setObjectName("sectionDesc")
        self.users_perm_desc.setWordWrap(True)
        ll.addWidget(self.users_perm_desc)

        self.user_permission_checks = {}

        self.users_perm_frame = QFrame()
        self.users_perm_frame.setObjectName("miniCard")
        perm_outer = QVBoxLayout(self.users_perm_frame)
        perm_outer.setContentsMargins(16, 16, 16, 16)
        perm_outer.setSpacing(14)
        ll.addWidget(self.users_perm_frame)

        def _build_perm_group(title_text: str, keys: list[str], columns: int = 2):
            frame = QFrame()
            frame.setObjectName("summaryCard")
            vbox = QVBoxLayout(frame)
            vbox.setContentsMargins(14, 14, 14, 14)
            vbox.setSpacing(10)

            title = QLabel(title_text)
            title.setObjectName("fieldLabel")
            vbox.addWidget(title)

            grid = QGridLayout()
            grid.setHorizontalSpacing(20)
            grid.setVerticalSpacing(12)
            vbox.addLayout(grid)

            for i, perm_key in enumerate(keys):
                chk = QCheckBox()
                chk.setMinimumHeight(26)
                self.user_permission_checks[perm_key] = chk
                grid.addWidget(chk, i // columns, i % columns)

            perm_outer.addWidget(frame)

        _build_perm_group(
            "Ana Sekmeler / Main Tabs",
            ["dashboard", "analysis", "notes", "responsibles", "muavin", "regular_ft", "user_management"],
            columns=2,
        )
        _build_perm_group(
            "Analiz Alt Yetkileri / Analysis Sub Permissions",
            ["analysis_tb", "analysis_plcc", "analysis_balance", "analysis_income"],
            columns=2,
        )
        _build_perm_group(
            "Muavin Sonuç Yetkileri / Subledger Result Permissions",
            [
                "muavin_user_based",
                "muavin_tax_based",
                "muavin_account_content",
            ],
            columns=2,
        )

        self.users_note = QLabel()
        self.users_note.setObjectName("softInfo")
        self.users_note.setWordWrap(True)
        ll.addWidget(self.users_note)
        ll.addStretch()

        rl = QVBoxLayout(right)
        rl.setContentsMargins(24, 24, 24, 24)
        rl.setSpacing(12)

        self.users_list_title = QLabel()
        self.users_list_title.setObjectName("sectionTitle")
        rl.addWidget(self.users_list_title)

        self.users_list_desc = QLabel()
        self.users_list_desc.setObjectName("sectionDesc")
        self.users_list_desc.setWordWrap(True)
        rl.addWidget(self.users_list_desc)

        self.users_table = QTableWidget()
        self.setup_table(self.users_table)
        self.users_table.cellClicked.connect(self.on_user_table_clicked)
        rl.addWidget(self.users_table, 1)

        return page

    def build_muavin_tab(self):
        page = QWidget()
        page_layout = QVBoxLayout(page)
        page_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        page_layout.addWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)

        outer = QVBoxLayout(container)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        card = QFrame()
        card.setObjectName("contentCard")
        outer.addWidget(card)

        main = QVBoxLayout(card)
        main.setContentsMargins(26, 24, 26, 26)
        main.setSpacing(18)

        title = QLabel("Muavin Analiz")
        title.setObjectName("sectionTitle")
        main.addWidget(title)
        self.muavin_title = title

        desc = QLabel("Muavin defteri yükleyerek dönem, belge türü, kullanıcı, karşıt hesap, vergi göstergesi ve masraf yeri bazlı çok boyutlu analiz alın.")
        desc.setWordWrap(True)
        desc.setObjectName("sectionDesc")
        main.addWidget(desc)
        self.muavin_desc = desc

        top_wrap = QFrame()
        top_wrap.setObjectName("miniCard")
        top_layout = QVBoxLayout(top_wrap)
        top_layout.setContentsMargins(16, 16, 16, 16)
        top_layout.setSpacing(12)
        main.addWidget(top_wrap)

        top = QHBoxLayout()
        top.setSpacing(12)
        top_layout.addLayout(top)

        self.btn_load_muavin = QPushButton("Muavin Yükle")
        self.btn_load_muavin.setObjectName("secondaryButton")
        self.btn_load_muavin.clicked.connect(self.load_muavin_file)
        self.btn_load_muavin.setMinimumHeight(42)
        top.addWidget(self.btn_load_muavin)

        self.btn_start_muavin = QPushButton("Başlat")
        self.btn_start_muavin.setObjectName("primaryButton")
        self.btn_start_muavin.clicked.connect(self.start_muavin_analysis)
        self.btn_start_muavin.setMinimumHeight(42)
        top.addWidget(self.btn_start_muavin)

        self.muavin_file_label = QLineEdit()
        self.muavin_file_label.setReadOnly(True)
        self.muavin_file_label.setPlaceholderText("Yüklenen muavin dosyası")
        self.muavin_file_label.setMinimumHeight(42)
        top.addWidget(self.muavin_file_label, 1)

        self.btn_export_muavin = QPushButton("Muavin Excel Çıktısı")
        self.btn_export_muavin.setObjectName("secondaryButton")
        self.btn_export_muavin.clicked.connect(self.export_muavin_analysis)
        self.btn_export_muavin.setMinimumHeight(42)
        top.addWidget(self.btn_export_muavin)

        action_row = QHBoxLayout()
        action_row.setSpacing(12)
        top_layout.addLayout(action_row)

        self.btn_muavin_risk_only = QPushButton("Sadece Riskli Kayıtlar")
        self.btn_muavin_risk_only.setCheckable(True)
        self.btn_muavin_risk_only.clicked.connect(self.toggle_muavin_risk_only)
        self.btn_muavin_risk_only.setMinimumHeight(40)
        action_row.addWidget(self.btn_muavin_risk_only)

        self.btn_muavin_cost_alarm_only = QPushButton("Sadece Masraf Alarmı")
        self.btn_muavin_cost_alarm_only.setCheckable(True)
        self.btn_muavin_cost_alarm_only.clicked.connect(self.toggle_muavin_cost_alarm_only)
        self.btn_muavin_cost_alarm_only.setMinimumHeight(40)
        action_row.addWidget(self.btn_muavin_cost_alarm_only)

        self.btn_muavin_autofit = QPushButton("Kolonları Otomatik Sığdır")
        self.btn_muavin_autofit.setObjectName("secondaryButton")
        self.btn_muavin_autofit.clicked.connect(self.auto_fit_muavin_tables)
        self.btn_muavin_autofit.setMinimumHeight(40)
        action_row.addWidget(self.btn_muavin_autofit)

        self.cmb_muavin_density = QComboBox()
        self.cmb_muavin_density.currentIndexChanged.connect(self.on_muavin_density_changed)
        self.cmb_muavin_density.setMinimumHeight(40)
        self.cmb_muavin_density.setMinimumWidth(180)
        action_row.addWidget(self.cmb_muavin_density)

        self.btn_muavin_expand_all = QPushButton("Tümünü Aç")
        self.btn_muavin_expand_all.setObjectName("secondaryButton")
        self.btn_muavin_expand_all.clicked.connect(lambda: self.set_all_muavin_sections(True))
        self.btn_muavin_expand_all.setMinimumHeight(40)
        action_row.addWidget(self.btn_muavin_expand_all)

        self.btn_muavin_collapse_all = QPushButton("Tümünü Kapat")
        self.btn_muavin_collapse_all.setObjectName("secondaryButton")
        self.btn_muavin_collapse_all.clicked.connect(lambda: self.set_all_muavin_sections(False))
        self.btn_muavin_collapse_all.setMinimumHeight(40)
        action_row.addWidget(self.btn_muavin_collapse_all)

        action_row.addStretch()

        self.muavin_info_pill = QLabel("-")
        self.muavin_info_pill.setObjectName("pillInfo")
        action_row.addWidget(self.muavin_info_pill)

        self.muavin_required_cols_label = QLabel("")
        self.muavin_required_cols_label.setObjectName("softInfo")
        self.muavin_required_cols_label.setWordWrap(True)
        top_layout.addWidget(self.muavin_required_cols_label)

        self.muavin_loaded_cols_label = QLabel("")
        self.muavin_loaded_cols_label.setObjectName("softInfo")
        self.muavin_loaded_cols_label.setWordWrap(True)
        top_layout.addWidget(self.muavin_loaded_cols_label)

        mapping_wrap = QFrame()
        mapping_wrap.setObjectName("miniCard")
        mapping_layout = QVBoxLayout(mapping_wrap)
        mapping_layout.setContentsMargins(16, 16, 16, 16)
        mapping_layout.setSpacing(10)
        main.addWidget(mapping_wrap)

        self.muavin_mapping_title = QLabel("Muavin Kolon Eşleme")
        self.muavin_mapping_title.setObjectName("fieldLabel")
        mapping_layout.addWidget(self.muavin_mapping_title)

        self.muavin_mapping_desc = QLabel("Farklı formatta gelen muavin raporlarında kaynak kolonları burada analiz alanlarına eşleyebilirsiniz.")
        self.muavin_mapping_desc.setObjectName("softInfo")
        self.muavin_mapping_desc.setWordWrap(True)
        mapping_layout.addWidget(self.muavin_mapping_desc)

        map_btn_row = QHBoxLayout()
        mapping_layout.addLayout(map_btn_row)
        self.btn_muavin_auto_map = QPushButton("Otomatik Eşleştir")
        self.btn_muavin_auto_map.setObjectName("secondaryButton")
        self.btn_muavin_auto_map.clicked.connect(self.handle_muavin_auto_map)
        map_btn_row.addWidget(self.btn_muavin_auto_map)
        self.btn_muavin_apply_map = QPushButton("Eşlemeyi Uygula")
        self.btn_muavin_apply_map.setObjectName("secondaryButton")
        self.btn_muavin_apply_map.clicked.connect(self.handle_muavin_apply_map)
        map_btn_row.addWidget(self.btn_muavin_apply_map)
        map_btn_row.addStretch()

        self.muavin_mapping_table = QTableWidget()
        self.setup_table(self.muavin_mapping_table)
        self.muavin_mapping_table.setMinimumHeight(420)
        mapping_layout.addWidget(self.muavin_mapping_table)

        filter_wrap = QFrame()
        filter_wrap.setObjectName("miniCard")
        filter_layout = QVBoxLayout(filter_wrap)
        filter_layout.setContentsMargins(16, 16, 16, 16)
        filter_layout.setSpacing(12)
        main.addWidget(filter_wrap)

        filter_title = QLabel("Filtreler")
        filter_title.setObjectName("fieldLabel")
        filter_layout.addWidget(filter_title)

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(12)
        filter_layout.addLayout(grid)

        self.lbl_muavin_account = QLabel("Hesap")
        self.lbl_muavin_statement = QLabel("Finansal Tablo")
        self.lbl_muavin_period = QLabel("Dönem")
        self.lbl_muavin_doc_type = QLabel("Belge Türü")
        self.lbl_muavin_user = QLabel("Kullanıcı")
        self.lbl_muavin_contra = QLabel("Satıcı / Cari")
        self.lbl_muavin_cost = QLabel("Masraf Yeri")
        self.lbl_muavin_search = QLabel("Serbest Arama")
        self.lbl_muavin_search_mode = QLabel("Arama Tipi")
        for lbl in [self.lbl_muavin_account, self.lbl_muavin_statement, self.lbl_muavin_period, self.lbl_muavin_doc_type, self.lbl_muavin_user, self.lbl_muavin_contra, self.lbl_muavin_cost, self.lbl_muavin_search, self.lbl_muavin_search_mode]:
            lbl.setObjectName("fieldLabel")

        grid.addWidget(self.lbl_muavin_account, 0, 0)
        grid.addWidget(self.lbl_muavin_statement, 0, 1)
        grid.addWidget(self.lbl_muavin_period, 0, 2)
        grid.addWidget(self.lbl_muavin_doc_type, 0, 3)

        self.cmb_muavin_account = QComboBox()
        self.cmb_muavin_account.currentIndexChanged.connect(self.refresh_muavin_tables)
        self.cmb_muavin_account.setMinimumHeight(40)
        grid.addWidget(self.cmb_muavin_account, 1, 0)

        self.cmb_muavin_statement = QComboBox()
        self.cmb_muavin_statement.currentIndexChanged.connect(self.refresh_muavin_tables)
        self.cmb_muavin_statement.setMinimumHeight(40)
        grid.addWidget(self.cmb_muavin_statement, 1, 1)

        self.cmb_muavin_period = QComboBox()
        self.cmb_muavin_period.currentIndexChanged.connect(self.refresh_muavin_tables)
        self.cmb_muavin_period.setMinimumHeight(40)
        grid.addWidget(self.cmb_muavin_period, 1, 2)

        self.cmb_muavin_doc_type = QComboBox()
        self.cmb_muavin_doc_type.currentIndexChanged.connect(self.refresh_muavin_tables)
        self.cmb_muavin_doc_type.setMinimumHeight(40)
        grid.addWidget(self.cmb_muavin_doc_type, 1, 3)

        grid.addWidget(self.lbl_muavin_user, 2, 0)
        grid.addWidget(self.lbl_muavin_contra, 2, 1)
        grid.addWidget(self.lbl_muavin_cost, 2, 2)
        grid.addWidget(self.lbl_muavin_search, 2, 3)
        grid.addWidget(self.lbl_muavin_search_mode, 2, 4)

        self.cmb_muavin_user = QComboBox()
        self.cmb_muavin_user.currentIndexChanged.connect(self.refresh_muavin_tables)
        self.cmb_muavin_user.setMinimumHeight(40)
        grid.addWidget(self.cmb_muavin_user, 3, 0)

        self.cmb_muavin_contra = QComboBox()
        self.cmb_muavin_contra.currentIndexChanged.connect(self.refresh_muavin_tables)
        self.cmb_muavin_contra.setMinimumHeight(40)
        grid.addWidget(self.cmb_muavin_contra, 3, 1)

        self.cmb_muavin_cost = QComboBox()
        self.cmb_muavin_cost.currentIndexChanged.connect(self.refresh_muavin_tables)
        self.cmb_muavin_cost.setMinimumHeight(40)
        grid.addWidget(self.cmb_muavin_cost, 3, 2)

        self.input_muavin_search = QLineEdit()
        self.input_muavin_search.setPlaceholderText("Metin / referans / satıcı-cari / belge no")
        self.input_muavin_search.textChanged.connect(self.refresh_muavin_tables)
        self.input_muavin_search.setMinimumHeight(40)
        grid.addWidget(self.input_muavin_search, 3, 3)

        self.cmb_muavin_search_mode = QComboBox()
        self.cmb_muavin_search_mode.currentIndexChanged.connect(self.on_muavin_search_mode_changed)
        self.cmb_muavin_search_mode.setMinimumHeight(40)
        grid.addWidget(self.cmb_muavin_search_mode, 3, 4)

        for col in range(5):
            grid.setColumnStretch(col, 1)

        cards_row = QHBoxLayout()
        cards_row.setSpacing(12)
        main.addLayout(cards_row)
        self.muavin_metric_cards = []
        for _ in range(4):
            frame = QFrame()
            frame.setObjectName("miniCard")
            frame.setMinimumHeight(98)
            fl = QVBoxLayout(frame)
            fl.setContentsMargins(14, 14, 14, 14)
            fl.setSpacing(6)
            t = QLabel("-")
            t.setObjectName("miniCardTitle")
            v = QLabel("0")
            v.setObjectName("summaryValue")
            s = QLabel("-")
            s.setObjectName("softInfo")
            s.setWordWrap(True)
            fl.addWidget(t)
            fl.addWidget(v)
            fl.addWidget(s)
            cards_row.addWidget(frame)
            self.muavin_metric_cards.append({"title": t, "value": v, "sub": s})

        self.muavin_result_buttons = {}
        self.muavin_tables = {}
        self.muavin_view_wrappers = {}
        self.muavin_section_buttons = {}
        self.muavin_section_wrappers = {}
        self.muavin_section_frames = {}

        result_wrap = QFrame()
        result_wrap.setObjectName("miniCard")
        result_layout = QVBoxLayout(result_wrap)
        result_layout.setContentsMargins(16, 16, 16, 16)
        result_layout.setSpacing(12)
        main.addWidget(result_wrap, 1)

        button_row = QHBoxLayout()
        button_row.setSpacing(10)
        result_layout.addLayout(button_row)

        self.btn_muavin_view_user = QPushButton("Kullanıcı Bazlı Analiz")
        self.btn_muavin_view_user.clicked.connect(lambda: self.set_muavin_result_view("user_based"))
        button_row.addWidget(self.btn_muavin_view_user)
        self.muavin_result_buttons["user_based"] = self.btn_muavin_view_user

        self.btn_muavin_view_tax = QPushButton("Vergisel Analiz")
        self.btn_muavin_view_tax.clicked.connect(lambda: self.set_muavin_result_view("tax_based"))
        button_row.addWidget(self.btn_muavin_view_tax)
        self.muavin_result_buttons["tax_based"] = self.btn_muavin_view_tax

        self.btn_muavin_view_account = QPushButton("Hesap İçerik Analiz")
        self.btn_muavin_view_account.clicked.connect(lambda: self.set_muavin_result_view("account_content"))
        button_row.addWidget(self.btn_muavin_view_account)
        self.muavin_result_buttons["account_content"] = self.btn_muavin_view_account
        button_row.addStretch()

        self.muavin_result_info = QLabel("-")
        self.muavin_result_info.setObjectName("softInfo")
        self.muavin_result_info.setWordWrap(True)
        result_layout.addWidget(self.muavin_result_info)

        def _build_result_section(parent_layout, title_text, table_key, min_height=220):
            section_frame = QFrame()
            section_frame.setObjectName("summaryCard")
            section_layout = QVBoxLayout(section_frame)
            section_layout.setContentsMargins(8, 8, 8, 8)
            section_layout.setSpacing(8)

            toggle_btn = QPushButton(f"▾ {title_text}")
            toggle_btn.setObjectName("secondaryButton")
            toggle_btn.setCheckable(False)
            toggle_btn.setMinimumHeight(42)
            toggle_btn.clicked.connect(lambda _=False, k=table_key: self.toggle_muavin_section(k))
            section_layout.addWidget(toggle_btn)

            content_wrap = QWidget()
            content_layout = QVBoxLayout(content_wrap)
            content_layout.setContentsMargins(6, 0, 6, 6)
            content_layout.setSpacing(6)

            table = QTableWidget()
            self.setup_table(table)
            table.setMinimumHeight(min_height)
            content_layout.addWidget(table, 1)

            section_layout.addWidget(content_wrap, 1)
            parent_layout.addWidget(section_frame)

            self.muavin_tables[table_key] = table
            self.muavin_section_frames[table_key] = section_frame
            self.muavin_section_buttons[table_key] = toggle_btn
            self.muavin_section_wrappers[table_key] = content_wrap
            if not hasattr(self, "muavin_section_states") or not isinstance(self.muavin_section_states, dict):
                self.muavin_section_states = {}
            if table_key not in self.muavin_section_states:
                self.muavin_section_states[table_key] = True
            content_wrap.setVisible(bool(self.muavin_section_states.get(table_key, True)))
            toggle_btn.setText(f"{'▾' if self.muavin_section_states.get(table_key, True) else '▸'} {title_text}")
            return table

        user_wrap = QFrame()
        user_wrap.setObjectName("miniCard")
        user_layout = QVBoxLayout(user_wrap)
        user_layout.setContentsMargins(0, 0, 0, 0)
        user_layout.setSpacing(10)
        _build_result_section(user_layout, "Kullanıcı Bazlı Analiz" if self.language == "tr" else "User-Based Analysis", "user_based", min_height=520)
        result_layout.addWidget(user_wrap, 1)
        self.muavin_view_wrappers["user_based"] = user_wrap

        tax_wrap = QFrame()
        tax_wrap.setObjectName("miniCard")
        tax_layout = QVBoxLayout(tax_wrap)
        tax_layout.setContentsMargins(0, 0, 0, 0)
        tax_layout.setSpacing(10)
        _build_result_section(tax_layout, "1. Mükerrer Belge Analizi" if self.language == "tr" else "1. Duplicate Document Analysis", "tax_dup", min_height=180)
        _build_result_section(tax_layout, "2. Vergisel Projeksiyon" if self.language == "tr" else "2. Tax Projection", "tax_projection", min_height=180)
        _build_result_section(tax_layout, "3. Vergi Göstergesi Bazlı Dönemsel Tutar / Belge Adedi" if self.language == "tr" else "3. Tax Indicator Period Amount / Document Count", "tax_indicator", min_height=180)
        result_layout.addWidget(tax_wrap, 1)
        self.muavin_view_wrappers["tax_based"] = tax_wrap

        account_wrap = QFrame()
        account_wrap.setObjectName("miniCard")
        account_layout = QVBoxLayout(account_wrap)
        account_layout.setContentsMargins(0, 0, 0, 0)
        account_layout.setSpacing(10)
        _build_result_section(account_layout, "1. Metin Açıklaması Yoğunluk Analizi" if self.language == "tr" else "1. Text Description Density Analysis", "account_text", min_height=180)
        _build_result_section(account_layout, "2. Belge Numarası ve Karşı Hesap Bazlı Hesap Çalışma Analizi" if self.language == "tr" else "2. Account Working Analysis by Document No and Contra Account", "account_relation", min_height=180)
        _build_result_section(account_layout, "3. Masraf Yeri Dönemsel Tutar Analizi" if self.language == "tr" else "3. Cost Center Period Amount Analysis", "account_cost", min_height=180)
        result_layout.addWidget(account_wrap, 1)
        self.muavin_view_wrappers["account_content"] = account_wrap

        self.set_muavin_result_view(getattr(self, "muavin_current_view", "user_based"))
        main.addStretch()
        return page

    def setup_table(self, table: QTableWidget):
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectRows)
        table.setSelectionMode(QTableWidget.SingleSelection)
        table.verticalHeader().setVisible(False)
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.horizontalHeader().setSectionsClickable(True)
        table.horizontalHeader().setMinimumSectionSize(60)
        table.setSortingEnabled(True)
        table.horizontalHeader().setStretchLastSection(False)
        table.setWordWrap(False)

    def enable_manual_column_resize(self, table: QTableWidget):
        if table is None:
            return
        header = table.horizontalHeader()
        for i in range(table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.Interactive)
        header.setStretchLastSection(False)

    def get_available_periods_from_loaded_files(self) -> List[str]:
        periods = []

        if self.tb_raw_df is not None and not self.tb_raw_df.empty:
            cols_norm = {normalize_col_name(c): c for c in self.tb_raw_df.columns}
            col_donem = cols_norm.get(normalize_col_name("Kayıt dönemi"))
            if col_donem:
                for v in self.tb_raw_df[col_donem].dropna().unique():
                    label = period_key_to_label(v)
                    if label == OPENING_PERIOD_TR or label in PERIOD_ORDER:
                        periods.append(label)

        if self.plcc_raw_df is not None and not self.plcc_raw_df.empty:
            cols_norm = {normalize_col_name(c): c for c in self.plcc_raw_df.columns}
            col_donem = cols_norm.get(normalize_col_name("Kayıt dönemi"))
            if col_donem:
                for v in self.plcc_raw_df[col_donem].dropna().unique():
                    label = period_key_to_label(v)
                    if label == OPENING_PERIOD_TR or label in PERIOD_ORDER:
                        periods.append(label)

        periods = sort_periods(periods)
        if not periods:
            periods = ["02-Mart", "03-Nisan"]
        return periods

    def sync_period_selection(self):
        self.available_periods = self.get_available_periods_from_loaded_files()
        selectable_periods = [p for p in self.available_periods if p != OPENING_PERIOD_TR]

        if not selectable_periods:
            selectable_periods = ["02-Mart", "03-Nisan"]

        if self.current_period not in selectable_periods:
            self.current_period = selectable_periods[-1]

        valid_previous = [p for p in selectable_periods if p != self.current_period]
        if self.previous_period not in valid_previous:
            self.previous_period = valid_previous[-1] if valid_previous else self.current_period


    def build_tb_rows(self) -> List[Dict]:
        if self.tb_raw_df is None or self.tb_raw_df.empty:
            return []

        df = self.tb_raw_df
        cols_norm = {normalize_col_name(c): c for c in df.columns}

        col_ana_hesap = cols_norm.get(normalize_col_name("Ana hesap"))
        col_hesap_adi = cols_norm.get(normalize_col_name("Ana hesap: Uzun metin"))
        col_tutar = cols_norm.get(normalize_col_name("Şirket kodu para birimi değeri"))
        col_mali_yil = cols_norm.get(normalize_col_name("Mali yıl"))
        col_donem = cols_norm.get(normalize_col_name("Kayıt dönemi"))

        missing = []
        if not col_ana_hesap:
            missing.append("Ana hesap")
        if not col_hesap_adi:
            missing.append("Ana hesap: Uzun metin")
        if not col_tutar:
            missing.append("Şirket kodu para birimi değeri")
        if not col_mali_yil:
            missing.append("Mali yıl")
        if not col_donem:
            missing.append("Kayıt dönemi")
        if missing:
            raise ValueError("TB dosyasında bulunamayan kolonlar: " + ", ".join(missing))

        periods = self.available_periods[:]
        aggregated: Dict[str, Dict] = {}

        for _, r in df.iterrows():
            ana_hesap = str(r[col_ana_hesap]).strip()
            if ana_hesap == "" or ana_hesap.lower() == "nan":
                continue

            hesap3 = normalize_hesap_prefix(ana_hesap)
            hesap_adi = str(r[col_hesap_adi]).strip()
            tutar = safe_float(r[col_tutar])
            mali_yil = str(r[col_mali_yil]).strip()
            donem_label = period_key_to_label(r[col_donem])

            if donem_label not in periods:
                continue

            yon_info = HESAP_YONU_MAP.get(hesap3, {})

            if ana_hesap not in aggregated:
                aggregated[ana_hesap] = {
                    "anaHesap": ana_hesap,
                    "anaHesapTam": ana_hesap,
                    "hesap": hesap3,
                    "sinif": hesap3[:1] if hesap3 else "-",
                    "hesapSinifi": yon_info.get("hesapSinifi", "-"),
                    "hesapAdi": hesap_adi or ana_hesap,
                    "hesapAdiEn": hesap_adi or ana_hesap,
                    "beklenenYon": yon_info.get("beklenenYon", "-"),
                    "maliYil": mali_yil,
                    "valuesNumeric": {p: 0.0 for p in periods},
                    "toplamNumeric": 0.0,
                }

            aggregated[ana_hesap]["valuesNumeric"][donem_label] += tutar
            aggregated[ana_hesap]["toplamNumeric"] += tutar

        class8_rules = {}
        class9_period_totals = {p: 0.0 for p in periods}
        class9_total = 0.0

        for full_hesap, item in aggregated.items():
            hesap_prefix = item.get("hesap", "")

            if str(hesap_prefix).startswith("8"):
                notes = []
                for p in periods:
                    if abs(item["valuesNumeric"].get(p, 0.0)) > 1e-9:
                        notes.append(f"{p} döneminde bakiye var")
                if abs(item["toplamNumeric"]) > 1e-9:
                    notes.append("Toplamda Bakiye var")
                if notes:
                    class8_rules[full_hesap] = join_notes(notes)

            if str(hesap_prefix).startswith("9"):
                for p in periods:
                    class9_period_totals[p] += item["valuesNumeric"].get(p, 0.0)
                class9_total += item["toplamNumeric"]

        class9_note = ""
        class9_note_parts = []
        for p in periods:
            if abs(class9_period_totals.get(p, 0.0)) > 1e-9:
                class9_note_parts.append(f"9'lu hesap grubu {p} döneminde sıfır vermiyor")
        if abs(class9_total) > 1e-9:
            class9_note_parts.append("9'lu hesap grubu toplamı sıfır vermiyor")
        if class9_note_parts:
            class9_note = join_notes(class9_note_parts)

        first_9_hesap = None
        sorted_keys = sorted(
            aggregated.keys(),
            key=lambda x: int("".join(ch for ch in str(x) if ch.isdigit())) if "".join(ch for ch in str(x) if ch.isdigit()) else str(x)
        )
        for key in sorted_keys:
            if str(aggregated[key].get("hesap", "")).startswith("9"):
                first_9_hesap = key
                break

        rows = []
        for full_hesap in sorted_keys:
            item = aggregated[full_hesap]

            current_numeric = item["valuesNumeric"].get(self.current_period, 0.0)
            previous_numeric = item["valuesNumeric"].get(self.previous_period, 0.0)

            note_tr, note_en = collect_matching_notes(
                self.notes,
                item["hesap"],
                item["anaHesapTam"],
                ""
            )

            extra_notes_tr = []
            if full_hesap in class8_rules:
                extra_notes_tr.append(class8_rules[full_hesap])
            if class9_note and full_hesap == first_9_hesap:
                extra_notes_tr.append(class9_note)

            note_tr = join_notes([note_tr] + extra_notes_tr)
            note_en = join_notes([note_en] + extra_notes_tr)

            toplam_yon = get_direction_from_amount(item["toplamNumeric"])
            kontrol = get_control_status(item["beklenenYon"], toplam_yon)

            sorumlu = collect_matching_responsibles(
                self.responsibles,
                item["hesap"],
                item["anaHesapTam"]
            )

            rows.append({
                **item,
                "values": {k: format_number(v) for k, v in item["valuesNumeric"].items()},
                "opening": format_number(item["valuesNumeric"].get(OPENING_PERIOD_TR, 0.0)),
                "toplam": format_number(item["toplamNumeric"]),
                "toplamYon": toplam_yon,
                "degisim": calc_change_percent(current_numeric, previous_numeric),
                "not": note_tr,
                "noteEn": note_en,
                "kontrol": kontrol,
                "sorumlu": sorumlu,
            })

        return rows

    def build_plcc_rows(self) -> Tuple[List[Dict], List[Dict]]:
        if self.plcc_raw_df is None or self.plcc_raw_df.empty:
            return [], []

        df = self.plcc_raw_df
        cols_norm = {normalize_col_name(c): c for c in df.columns}

        col_hesap = cols_norm.get(normalize_col_name("Ana hesap"))
        col_hesap_adi = cols_norm.get(normalize_col_name("Ana hesap: Uzun metin"))
        col_tutar = cols_norm.get(normalize_col_name("Şirket kodu para birimi değeri"))
        col_masraf = cols_norm.get(normalize_col_name("Masraf yeri"))
        col_donem = cols_norm.get(normalize_col_name("Kayıt dönemi"))
        col_masraf_adi = cols_norm.get(normalize_col_name("Masraf yeri: Kısa metin"))

        missing = []
        if not col_hesap:
            missing.append("Ana hesap")
        if not col_hesap_adi:
            missing.append("Ana hesap: Uzun metin")
        if not col_tutar:
            missing.append("Şirket kodu para birimi değeri")
        if not col_masraf:
            missing.append("Masraf yeri")
        if not col_donem:
            missing.append("Kayıt dönemi")
        if not col_masraf_adi:
            missing.append("Masraf yeri: Kısa metin")
        if missing:
            raise ValueError("PL-CC dosyasında bulunamayan kolonlar: " + ", ".join(missing))

        periods = self.available_periods[:]
        detail_map: Dict[Tuple[str, str], Dict] = {}
        subtotal_map: Dict[str, Dict] = {}

        for _, r in df.iterrows():
            hesap_tam = str(r[col_hesap]).strip()
            if hesap_tam == "" or hesap_tam.lower() == "nan":
                continue

            hesap = hesap_tam
            hesap_prefix = normalize_hesap_prefix(hesap_tam)
            hesap_adi = str(r[col_hesap_adi]).strip()
            tutar = safe_float(r[col_tutar])
            masraf = str(r[col_masraf]).strip()
            if masraf.lower() == "nan":
                masraf = ""
            masraf_adi = str(r[col_masraf_adi]).strip()
            if masraf_adi.lower() == "nan":
                masraf_adi = ""
            donem_label = period_key_to_label(r[col_donem])

            if donem_label not in periods:
                continue

            key = (masraf, hesap_tam)

            if key not in detail_map:
                detail_map[key] = {
                    "masrafYeri": masraf,
                    "ad": masraf_adi or masraf,
                    "adEn": masraf_adi or masraf,
                    "hesap": hesap,
                    "hesapPrefix": hesap_prefix,
                    "anaHesapTam": hesap_tam,
                    "hesapAdi": hesap_adi or hesap_tam,
                    "hesapAdiEn": hesap_adi or hesap_tam,
                    "valuesNumeric": {p: 0.0 for p in periods},
                    "toplamNumeric": 0.0,
                }

            detail_map[key]["valuesNumeric"][donem_label] += tutar
            detail_map[key]["toplamNumeric"] += tutar

            if masraf not in subtotal_map:
                subtotal_map[masraf] = {
                    "masrafYeri": masraf,
                    "ad": masraf_adi or masraf,
                    "adEn": masraf_adi or masraf,
                    "valuesNumeric": {p: 0.0 for p in periods},
                    "toplamNumeric": 0.0,
                }

            subtotal_map[masraf]["valuesNumeric"][donem_label] += tutar
            subtotal_map[masraf]["toplamNumeric"] += tutar

        detail_rows = []
        for item in detail_map.values():
            current_numeric = item["valuesNumeric"].get(self.current_period, 0.0)
            previous_numeric = item["valuesNumeric"].get(self.previous_period, 0.0)

            note_tr, note_en = collect_matching_notes(
                self.notes,
                item["hesap"],
                item["anaHesapTam"],
                item["masrafYeri"]
            )

            if str(item["masrafYeri"]).strip() == "":
                extra_tr = "Masraf Yeri Yoktur"
                extra_en = "Cost Center Missing"
                note_tr = join_notes([note_tr, extra_tr])
                note_en = join_notes([note_en, extra_en])

            expected = HESAP_YONU_MAP.get(item.get("hesapPrefix", normalize_hesap_prefix(item["hesap"])), {}).get("beklenenYon", "-")
            actual = get_direction_from_amount(item["toplamNumeric"])
            kontrol = get_control_status(expected, actual)

            sorumlu = collect_matching_responsibles(
                self.responsibles,
                item["hesap"],
                item["anaHesapTam"]
            )
            detail_rows.append({
                **item,
                "values": {k: format_number(v) for k, v in item["valuesNumeric"].items()},
                "toplam": format_number(item["toplamNumeric"]),
                "degisim": calc_change_percent(current_numeric, previous_numeric),
                "not": note_tr,
                "noteEn": note_en,
                "beklenenYon": expected,
                "toplamYon": actual,
                "kontrol": kontrol,
                "sorumlu": sorumlu,
            })

        subtotals = []
        for item in subtotal_map.values():
            current_numeric = item["valuesNumeric"].get(self.current_period, 0.0)
            previous_numeric = item["valuesNumeric"].get(self.previous_period, 0.0)
            subtotals.append({
                **item,
                "values": {k: format_number(v) for k, v in item["valuesNumeric"].items()},
                "toplam": format_number(item["toplamNumeric"]),
                "degisim": calc_change_percent(current_numeric, previous_numeric),
            })

        detail_rows.sort(key=lambda x: (masraf_sort_key(x["masrafYeri"]), str(x["anaHesapTam"])))
        subtotals.sort(key=lambda x: masraf_sort_key(x["masrafYeri"]))
        return detail_rows, subtotals
    def format_elapsed_time(self, elapsed_seconds: float) -> str:
        total_seconds = max(0, int(round(float(elapsed_seconds))))
        minutes, seconds = divmod(total_seconds, 60)
        hours, minutes = divmod(minutes, 60)
        if hours > 0:
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        return f"{minutes:02d}:{seconds:02d}"

    def _build_processing_text(self, text: str) -> str:
        if self.busy_started_at is None:
            return text
        elapsed = self.format_elapsed_time(time.monotonic() - self.busy_started_at)
        prefix = "Geçen Süre" if self.language == "tr" else "Elapsed"
        return f"{text} • {prefix}: {elapsed}"

    def _update_processing_value(self, progress: Optional[int] = None):
        if progress is not None:
            self.progress.setValue(max(0, min(100, int(progress))))
        current_progress = self.progress.value()
        if self.busy_started_at is not None:
            elapsed = self.format_elapsed_time(time.monotonic() - self.busy_started_at)
            self.proc_value.setText(f"{int(current_progress)}% • {elapsed}")
        else:
            self.proc_value.setText(f"{int(current_progress)}%")

    def refresh_busy_elapsed(self):
        if self.busy_started_at is None:
            return
        self.proc_text.setText(self._build_processing_text(self.busy_last_message or self.proc_text.text()))
        self._update_processing_value(None)

    def _deferred_info(self, title: str, message: str, delay_ms: int = 75):
        QTimer.singleShot(delay_ms, lambda t=title, m=message: QMessageBox.information(self, t, m))

    def _deferred_error(self, title: str, message: str, delay_ms: int = 75):
        QTimer.singleShot(delay_ms, lambda t=title, m=message: QMessageBox.critical(self, t, m))

    def on_tb_file_loaded_async(self, payload):
        self.clear_active_worker()
        try:
            self.set_processing_message("TB dosyası işlendi, dönemler hazırlanıyor..." if self.language == "tr" else "TB file processed, preparing periods...", 65)
            self.tb_raw_df = payload.get("df", pd.DataFrame()) if isinstance(payload, dict) else pd.DataFrame()
            self.tb_file_path = payload.get("path", "") if isinstance(payload, dict) else ""
            self.analysis_has_run = False
            self.available_periods = self.get_available_periods_from_loaded_files()
            self.sync_period_selection()
            self.refresh_all()
            msg = "TB dosyası yüklendi." if self.language == "tr" else "TB file loaded."
            self.end_busy_state(msg, 100)
            self._deferred_info("Bilgi" if self.language == "tr" else "Info", msg)
        except Exception as e:
            self.end_busy_state("TB dosyası yüklenemedi." if self.language == "tr" else "TB file could not be loaded.", 0)
            self._deferred_error("Hata" if self.language == "tr" else "Error", ("TB dosyası yüklenemedi." if self.language == "tr" else "TB file could not be loaded.") + "\n\n" + str(e))

    def on_tb_file_failed_async(self, error_text: str):
        self.clear_active_worker()
        self.end_busy_state("TB dosyası yüklenemedi." if self.language == "tr" else "TB file could not be loaded.", 0)
        self._deferred_error("Hata" if self.language == "tr" else "Error", str(error_text))

    def on_plcc_file_loaded_async(self, payload):
        self.clear_active_worker()
        try:
            self.set_processing_message("PL-CC dosyası işlendi, dönemler hazırlanıyor..." if self.language == "tr" else "PL-CC file processed, preparing periods...", 65)
            self.plcc_raw_df = payload.get("df", pd.DataFrame()) if isinstance(payload, dict) else pd.DataFrame()
            self.plcc_file_path = payload.get("path", "") if isinstance(payload, dict) else ""
            self.analysis_has_run = False
            self.available_periods = self.get_available_periods_from_loaded_files()
            self.sync_period_selection()
            self.refresh_all()
            msg = "PL-CC dosyası yüklendi." if self.language == "tr" else "PL-CC file loaded."
            self.end_busy_state(msg, 100)
            self._deferred_info("Bilgi" if self.language == "tr" else "Info", msg)
        except Exception as e:
            self.end_busy_state("PL-CC dosyası yüklenemedi." if self.language == "tr" else "PL-CC file could not be loaded.", 0)
            self._deferred_error("Hata" if self.language == "tr" else "Error", ("PL-CC dosyası yüklenemedi." if self.language == "tr" else "PL-CC file could not be loaded.") + "\n\n" + str(e))

    def on_plcc_file_failed_async(self, error_text: str):
        self.clear_active_worker()
        self.end_busy_state("PL-CC dosyası yüklenemedi." if self.language == "tr" else "PL-CC file could not be loaded.", 0)
        self._deferred_error("Hata" if self.language == "tr" else "Error", str(error_text))

    def _load_regular_ft_file_async(self, file_type: str, dialog_title: str):
        path, _ = QFileDialog.getOpenFileName(self, dialog_title, BASE_DIR, "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        label = {"faggl": "FAGGL", "eba": "EBA", "zfi052": "ZFI052"}.get(file_type, file_type.upper())
        begin_text = (f"{label} dosyası okunuyor..." if self.language == "tr" else f"Reading {label} file...")
        self.begin_busy_state(begin_text, 5)
        started = self.start_background_worker(
            load_single_excel_payload,
            lambda payload, ft=file_type: self.on_regular_ft_file_loaded(payload, ft),
            self.on_regular_ft_file_failed,
            path,
            label,
            self.language,
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def _finish_regular_ft_ui_load(self):
        try:
            self.refresh_regular_ft_table()
            self.tabs.setCurrentIndex(5)
            msg = "Düzenli gelen fatura analizi hazırlandı." if self.language == "tr" else "Recurring invoice analysis is ready."
            self.end_busy_state(msg, 100)
            self._deferred_info(self.t("info"), msg)
        except Exception as e:
            self.clear_active_worker()
            self.end_busy_state("Düzenli gelen fatura analizi oluşturulamadı." if self.language == "tr" else "Recurring invoice analysis could not be created.", 0)
            self._deferred_error(self.t("error"), ("Düzenli gelen fatura analizi oluşturulamadı." if self.language == "tr" else "Recurring invoice analysis could not be created.") + "\n\n" + str(e))

    def set_processing_message(self, text: str, progress: Optional[int] = None):
        self.busy_last_message = text
        self.proc_text.setText(self._build_processing_text(text))
        self._update_processing_value(progress)

    def begin_busy_state(self, text: str, progress: int = 0):
        self.is_busy = True
        self.busy_started_at = time.monotonic()
        self.busy_last_message = text
        self.busy_timer.start()
        QApplication.setOverrideCursor(Qt.WaitCursor)
        for btn in [
            getattr(self, "start_button", None),
            getattr(self, "btn_upload_tb", None),
            getattr(self, "btn_upload_plcc", None),
            getattr(self, "btn_load_muavin", None),
            getattr(self, "btn_start_muavin", None),
            getattr(self, "btn_export", None),
            getattr(self, "btn_export_muavin", None),
            getattr(self, "btn_regular_ft_load_faggl", None),
            getattr(self, "btn_regular_ft_load_eba", None),
            getattr(self, "btn_regular_ft_load_zfi052", None),
            getattr(self, "btn_regular_ft_start", None),
            getattr(self, "btn_regular_ft_export", None),
        ]:
            if btn is not None:
                btn.setEnabled(False)
        self.set_processing_message(text, progress)

    def end_busy_state(self, text: Optional[str] = None, progress: Optional[int] = None):
        elapsed_text = self.format_elapsed_time(time.monotonic() - self.busy_started_at) if self.busy_started_at is not None else "00:00"
        self.is_busy = False
        self.busy_timer.stop()
        try:
            QApplication.restoreOverrideCursor()
        except Exception:
            pass
        for btn in [
            getattr(self, "start_button", None),
            getattr(self, "btn_upload_tb", None),
            getattr(self, "btn_upload_plcc", None),
            getattr(self, "btn_load_muavin", None),
            getattr(self, "btn_start_muavin", None),
            getattr(self, "btn_export", None),
            getattr(self, "btn_export_muavin", None),
            getattr(self, "btn_regular_ft_load_faggl", None),
            getattr(self, "btn_regular_ft_load_eba", None),
            getattr(self, "btn_regular_ft_load_zfi052", None),
            getattr(self, "btn_regular_ft_start", None),
            getattr(self, "btn_regular_ft_export", None),
        ]:
            if btn is not None:
                btn.setEnabled(True)
        if text is not None or progress is not None:
            final_text = text or self.busy_last_message or self.proc_text.text()
            final_prefix = "Toplam Süre" if self.language == "tr" else "Total Time"
            final_text = f"{final_text} • {final_prefix}: {elapsed_text}"
            self.busy_started_at = None
            self.busy_last_message = final_text
            self.proc_text.setText(final_text)
            self._update_processing_value(self.progress.value() if progress is None else progress)
        else:
            self.busy_started_at = None
            self.busy_last_message = self.proc_text.text()
            self._update_processing_value(self.progress.value())

    def start_background_worker(self, fn, on_success, on_error, *args, **kwargs):
        if self.active_worker is not None and self.active_worker.isRunning():
            return False
        worker = BackgroundWorker(fn, *args, **kwargs)
        worker.progress.connect(self.on_background_worker_progress)
        worker.succeeded.connect(on_success)
        worker.failed.connect(on_error)
        self.active_worker = worker
        worker.start()
        return True

    def on_background_worker_progress(self, progress: int, message: str):
        self.busy_last_message = message
        self.proc_text.setText(self._build_processing_text(message))
        self._update_processing_value(progress)

    def clear_active_worker(self):
        worker = self.active_worker
        self.active_worker = None
        if worker is not None:
            try:
                worker.deleteLater()
            except Exception:
                pass


    def on_background_export_success(self, payload):
        self.clear_active_worker()
        export_path = ""
        export_paths = []
        part_count = 1
        row_count = 0
        if isinstance(payload, dict):
            export_path = str(payload.get("path", "")).strip()
            export_paths = [str(x).strip() for x in payload.get("paths", []) if str(x).strip()]
            part_count = int(payload.get("part_count", 1) or 1)
            row_count = int(payload.get("row_count", 0) or 0)
        if not export_paths and export_path:
            export_paths = [export_path]
        self.end_busy_state("İşlem tamamlandı. Dosya açılıyor..." if self.language == "tr" else "Operation completed. Opening file...", 100)
        if export_paths:
            if part_count > 1:
                if self.language == "tr":
                    msg = f"Çıktı verisi {row_count:,} satır olduğu için dosya otomatik olarak {part_count} parçaya bölündü:\n\n".replace(",", ".")
                    msg += "\n".join(os.path.basename(p) for p in export_paths)
                else:
                    msg = f"The export contained {row_count:,} rows, so it was automatically split into {part_count} parts:\n\n"
                    msg += "\n".join(os.path.basename(p) for p in export_paths)
            else:
                msg = (f"Dosya '{os.path.basename(export_paths[0])}' olarak oluşturuldu." if self.language == "tr"
                       else f"File '{os.path.basename(export_paths[0])}' has been created.")
            QMessageBox.information(
                self,
                "Bilgi" if self.language == "tr" else "Info",
                msg
            )
            for idx, p in enumerate(export_paths[:3]):
                self.schedule_open_saved_file(p, delay_ms=250 + (idx * 250))

    def on_background_export_failed(self, error_text: str):
        self.clear_active_worker()
        self.end_busy_state("İşlem tamamlanamadı." if self.language == "tr" else "Operation failed.", 0)
        QMessageBox.critical(
            self,
            "Hata" if self.language == "tr" else "Error",
            (f"İşlem tamamlanamadı.\n\n{error_text}" if self.language == "tr"
             else f"The operation could not be completed.\n\n{error_text}")
        )

    def background_build_main_analysis(self, progress_cb=None):
        _safe_progress(progress_cb, 10, "Dönemler hazırlanıyor..." if self.language == "tr" else "Preparing periods...")
        available_periods = self.get_available_periods_from_loaded_files()
        selectable_periods = [p for p in available_periods if p != OPENING_PERIOD_TR] or ["02-Mart", "03-Nisan"]
        current_period = self.current_period if self.current_period in selectable_periods else selectable_periods[-1]
        valid_previous = [p for p in selectable_periods if p != current_period]
        previous_period = self.previous_period if self.previous_period in valid_previous else (valid_previous[-1] if valid_previous else current_period)

        old_current = self.current_period
        old_previous = self.previous_period
        old_periods = list(self.available_periods)

        self.available_periods = available_periods
        self.current_period = current_period
        self.previous_period = previous_period

        tb_rows = []
        plcc_detail = []
        plcc_subtotal = []
        errors = []

        if self.tb_file_path:
            try:
                _safe_progress(progress_cb, 35, "TB analizi oluşturuluyor..." if self.language == "tr" else "Building TB analysis...")
                tb_rows = self.build_tb_rows()
            except Exception as e:
                errors.append(f"TB hata: {e}" if self.language == "tr" else f"TB error: {e}")

        if self.plcc_file_path:
            try:
                _safe_progress(progress_cb, 70, "PL-CC analizi oluşturuluyor..." if self.language == "tr" else "Building PL-CC analysis...")
                plcc_detail, plcc_subtotal = self.build_plcc_rows()
            except Exception as e:
                errors.append(f"PL-CC hata: {e}" if self.language == "tr" else f"PL-CC error: {e}")

        self.available_periods = old_periods
        self.current_period = old_current
        self.previous_period = old_previous

        return {
            "available_periods": available_periods,
            "current_period": current_period,
            "previous_period": previous_period,
            "tb_rows_cache": tb_rows,
            "plcc_detail_cache": plcc_detail,
            "plcc_subtotal_cache": plcc_subtotal,
            "errors": errors,
        }

    def on_main_analysis_loaded(self, payload):
        self.clear_active_worker()
        errors = payload.get("errors", []) if isinstance(payload, dict) else []
        if errors:
            self.analysis_has_run = False
            self.end_busy_state("Analiz tamamlanamadı." if self.language == "tr" else "Analysis failed.", 0)
            QMessageBox.critical(self, "Analiz Hatası" if self.language == "tr" else "Analysis Error", "\n\n".join(errors))
            return

        self.available_periods = payload.get("available_periods", self.available_periods)
        self.current_period = payload.get("current_period", self.current_period)
        self.previous_period = payload.get("previous_period", self.previous_period)
        self.tb_rows_cache = payload.get("tb_rows_cache", [])
        self.plcc_detail_cache = payload.get("plcc_detail_cache", [])
        self.plcc_subtotal_cache = payload.get("plcc_subtotal_cache", [])
        self.analysis_has_run = True
        self.refresh_all()
        self.tabs.setCurrentIndex(1)
        if self.tb_file_path:
            self.set_active_view("tb")
        elif self.plcc_file_path:
            self.set_active_view("plcc")
        self.end_busy_state("Analiz başarıyla hazırlandı." if self.language == "tr" else "Analysis is ready.", 100)
        QMessageBox.information(self, "Bilgi" if self.language == "tr" else "Info", "Analiz başarıyla hazırlandı." if self.language == "tr" else "Analysis is ready.")

    def on_main_analysis_failed(self, error_text: str):
        self.clear_active_worker()
        self.analysis_has_run = False
        self.end_busy_state("Analiz tamamlanamadı." if self.language == "tr" else "Analysis failed.", 0)
        QMessageBox.critical(self, "Analiz Hatası" if self.language == "tr" else "Analysis Error", str(error_text))

    def on_tb_loaded(self, payload):
        self.clear_active_worker()
        self.tb_raw_df = payload.get("df", pd.DataFrame()) if isinstance(payload, dict) else pd.DataFrame()
        self.tb_file_path = payload.get("path", "") if isinstance(payload, dict) else ""
        self.analysis_has_run = False
        self.available_periods = self.get_available_periods_from_loaded_files()
        self.sync_period_selection()
        self.refresh_all()
        self.end_busy_state("TB dosyası yüklendi." if self.language == "tr" else "TB file loaded.", 100)
        QMessageBox.information(self, "Bilgi" if self.language == "tr" else "Info", "TB dosyası yüklendi." if self.language == "tr" else "TB file loaded.")

    def on_tb_load_failed(self, error_text: str):
        self.clear_active_worker()
        self.end_busy_state("TB dosyası yüklenemedi." if self.language == "tr" else "TB file could not be loaded.", 0)
        QMessageBox.critical(self, "Hata" if self.language == "tr" else "Error", f"TB dosyası okunamadı.\n\n{error_text}" if self.language == "tr" else f"TB file could not be read.\n\n{error_text}")

    def on_plcc_loaded(self, payload):
        self.clear_active_worker()
        self.plcc_raw_df = payload.get("df", pd.DataFrame()) if isinstance(payload, dict) else pd.DataFrame()
        self.plcc_file_path = payload.get("path", "") if isinstance(payload, dict) else ""
        self.analysis_has_run = False
        self.available_periods = self.get_available_periods_from_loaded_files()
        self.sync_period_selection()
        self.refresh_all()
        self.end_busy_state("PL-CC dosyası yüklendi." if self.language == "tr" else "PL-CC file loaded.", 100)
        QMessageBox.information(self, "Bilgi" if self.language == "tr" else "Info", "PL-CC dosyası yüklendi." if self.language == "tr" else "PL-CC file loaded.")

    def on_plcc_load_failed(self, error_text: str):
        self.clear_active_worker()
        self.end_busy_state("PL-CC dosyası yüklenemedi." if self.language == "tr" else "PL-CC file could not be loaded.", 0)
        QMessageBox.critical(self, "Hata" if self.language == "tr" else "Error", f"PL-CC dosyası okunamadı.\n\n{error_text}" if self.language == "tr" else f"PL-CC file could not be read.\n\n{error_text}")

    def on_regular_ft_file_loaded(self, payload, file_type: str):
        self.clear_active_worker()
        loaded_df = payload.get("df", pd.DataFrame()) if isinstance(payload, dict) else pd.DataFrame()
        loaded_path = payload.get("path", "") if isinstance(payload, dict) else ""
        if file_type == "faggl":
            self.regular_ft_faggl_df = loaded_df
            self.regular_ft_faggl_path = loaded_path
            label_tr, label_en = "FAGGL dosyası yüklendi.", "FAGGL file loaded."
        elif file_type == "eba":
            self.regular_ft_eba_df = loaded_df
            self.regular_ft_eba_path = loaded_path
            label_tr, label_en = "EBA dosyası yüklendi.", "EBA file loaded."
        else:
            self.regular_ft_zfi052_df = loaded_df
            self.regular_ft_zfi052_path = loaded_path
            label_tr, label_en = "ZFI052 dosyası yüklendi.", "ZFI052 file loaded."
        self.regular_ft_analysis_ready = False
        self.refresh_regular_ft_ui()
        msg = label_tr if self.language == "tr" else label_en
        self.end_busy_state(msg, 100)
        self._deferred_info("Bilgi" if self.language == "tr" else "Info", msg)

    def on_regular_ft_file_failed(self, error_text: str):
        self.clear_active_worker()
        self.end_busy_state("Dosya yüklenemedi." if self.language == "tr" else "File could not be loaded.", 0)
        self._deferred_error("Hata" if self.language == "tr" else "Error", str(error_text))

    def background_build_regular_ft_analysis(self, progress_cb=None):
        self.prepare_regular_ft_analysis(progress_cb=progress_cb)
        return {
            "base_output": self.regular_ft_base_output_df.copy(),
            "output": self.regular_ft_output_df.copy(),
            "periods": list(self.regular_ft_periods),
            "current_period": self.regular_ft_current_period,
            "previous_period": self.regular_ft_previous_period,
            "user_period_map": dict(self.regular_ft_user_period_map),
            "eba_pending_period_dict": dict(self.regular_ft_eba_pending_period_dict),
            "eba_invoice_period_dict": dict(self.regular_ft_eba_invoice_period_dict),
            "faggl_invoice_keys_map": dict(self.regular_ft_faggl_invoice_keys_map),
            "zfi_invoice_map": dict(self.regular_ft_zfi_invoice_map),
            "zfi_period_map": dict(self.regular_ft_zfi_period_map),
            "zfi_vendor_map": dict(self.regular_ft_zfi_vendor_map),
        }

    def on_regular_ft_analysis_loaded(self, payload):
        self.clear_active_worker()
        self.regular_ft_base_output_df = payload.get("base_output", pd.DataFrame())
        self.regular_ft_output_df = payload.get("output", pd.DataFrame())
        self.regular_ft_periods = payload.get("periods", [])
        self.regular_ft_current_period = payload.get("current_period", self.regular_ft_current_period)
        self.regular_ft_previous_period = payload.get("previous_period", self.regular_ft_previous_period)
        self.regular_ft_user_period_map = payload.get("user_period_map", {})
        self.regular_ft_eba_pending_period_dict = payload.get("eba_pending_period_dict", {})
        self.regular_ft_eba_invoice_period_dict = payload.get("eba_invoice_period_dict", {})
        self.regular_ft_faggl_invoice_keys_map = payload.get("faggl_invoice_keys_map", {})
        self.regular_ft_zfi_invoice_map = payload.get("zfi_invoice_map", {})
        self.regular_ft_zfi_period_map = payload.get("zfi_period_map", {})
        self.regular_ft_zfi_vendor_map = payload.get("zfi_vendor_map", {})
        self.regular_ft_analysis_ready = True
        self.refresh_regular_ft_ui()
        QTimer.singleShot(0, self._finish_regular_ft_ui_load)

    def on_regular_ft_analysis_failed(self, error_text: str):
        self.clear_active_worker()
        self.end_busy_state("Düzenli gelen fatura analizi oluşturulamadı." if self.language == "tr" else "Recurring invoice analysis could not be created.", 0)
        self._deferred_error(self.t("error"), ("Düzenli gelen fatura analizi oluşturulamadı." if self.language == "tr" else "Recurring invoice analysis could not be created.") + "\n\n" + str(error_text))

    def schedule_open_saved_file(self, path: str, delay_ms: int = 250):
        normalized_path = os.path.normpath(path)
        QTimer.singleShot(delay_ms, lambda p=normalized_path: open_file_after_save(p))

    def refresh_all(self):
        self.refresh_header()
        self.refresh_sidebar()
        self.refresh_tabs_text()
        self.refresh_dashboard()
        self.refresh_analysis_ui()
        self.refresh_analysis_tables()
        self.refresh_notes_ui()
        self.refresh_notes_table()
        self.refresh_responsibles_ui()
        self.refresh_responsibles_table()
        self.refresh_users_ui()
        self.refresh_users_table()
        self.refresh_muavin_ui()
        self.refresh_muavin_tables()
        self.refresh_regular_ft_ui()
        self.refresh_regular_ft_table()
        self.update_progress()

    def refresh_header(self):
        self.header_badge.setText(self.t("premiumView") + f" • {self.current_user}")
        self.header_title.setText(self.t("appTitle"))
        self.header_desc.setText(self.t("appDesc"))
        self.start_button.setText(self.t("start"))

        self.btn_tr.setObjectName("langButtonActive" if self.language == "tr" else "langButton")
        self.btn_en.setObjectName("langButtonActive" if self.language == "en" else "langButton")
        self.btn_tr.style().unpolish(self.btn_tr)
        self.btn_tr.style().polish(self.btn_tr)
        self.btn_en.style().unpolish(self.btn_en)
        self.btn_en.style().polish(self.btn_en)

        tb_findings = len([r for r in self.tb_rows_cache if r.get("not") or r.get("kontrol") == "Ters duruyor"])
        plcc_findings = len([r for r in self.plcc_detail_cache if r.get("not") or r.get("kontrol") == "Ters duruyor"])

        raw_cards = [
            {"titleKey": "loadedTb", "value": os.path.basename(self.tb_file_path) if self.tb_file_path else "-"},
            {"titleKey": "loadedPlcc", "value": os.path.basename(self.plcc_file_path) if self.plcc_file_path else "-"},
            {"titleKey": "definedNotes", "value": f"{len(self.notes)} kayıt", "valueEn": f"{len(self.notes)} records"},
            {"titleKey": "findings", "value": "Detaylı görüntüle", "valueEn": "View details"},
        ]

        for idx, item in enumerate(raw_cards):
            card = self.summary_cards[idx]
            card["title"].setText(self.t(item["titleKey"]))
            if item["titleKey"] == "findings":
                card["value"].setText(f'{self.t("tb")}: {tb_findings} • {self.t("plcc")}: {plcc_findings}')
                card["btn1"].setText(f'{self.t("findingsTb")} ({tb_findings})')
                card["btn2"].setText(f'{self.t("findingsPlcc")} ({plcc_findings})')
                card["btn1"].show()
                card["btn2"].show()
            else:
                card["value"].setText(item.get("valueEn") if self.language == "en" and item.get("valueEn") else item["value"])
                card["btn1"].hide()
                card["btn2"].hide()

    def refresh_sidebar(self):
        self.sidebar_title.setText(self.t("panel"))
        self.sidebar_desc.setText(self.t("panelDesc"))
        self.btn_upload_tb.setText(self.t("uploadTb"))
        self.btn_upload_plcc.setText(self.t("uploadPlcc"))
        self.btn_note_defs.setText(self.t("noteDefinitions"))
        self.btn_responsibles.setText(self.t("responsibleAssignments"))
        self.btn_muavin.setText(self.t("muavinAnalysis"))
        self.btn_regular_ft.setText(self.t("regularFtAnalysis"))
        self.proc_label.setText(self.t("processingStatus"))
        self.proc_text.setText(self.t("processingText"))

    def refresh_tabs_text(self):
        self.tabs.setTabText(0, self.t("dashboard"))
        self.tabs.setTabText(1, self.t("analysis"))
        self.tabs.setTabText(2, self.t("noteDefinitions"))
        self.tabs.setTabText(3, self.t("responsibleAssignments"))
        self.tabs.setTabText(4, self.t("muavinAnalysis"))
        self.tabs.setTabText(5, self.t("regularFtAnalysis"))
        self.tabs.setTabText(6, self.t("userManagement"))
        self.apply_user_permissions()

    def refresh_dashboard(self):
        self.dashboard_title.setText(self.t("dashboardTitle"))
        self.dashboard_desc.setText(self.t("dashboardDesc"))

        self.lbl_tb_file.setText(self.t("tbFile"))
        self.lbl_plcc_file.setText(self.t("plccFile"))
        self.lbl_curr_period.setText(self.t("currentPeriod"))
        self.lbl_prev_period.setText(self.t("previousPeriod"))

        self.tb_file_input.setText(self.tb_file_path or "")
        self.plcc_file_input.setText(self.plcc_file_path or "")

        self.fill_period_combos()

        self.dashboard_pills[0].setText(self.t("tbReady") if self.tb_file_path else self.t("notLoadedTb"))
        self.dashboard_pills[1].setText(self.t("plccReady") if self.plcc_file_path else self.t("notLoadedPlcc"))
        self.dashboard_pills[2].setText(self.t("analysisReady") if self.analysis_has_run else self.t("start"))

        self.dashboard_pills[0].setObjectName("dashboardPillOk")
        self.dashboard_pills[1].setObjectName("dashboardPillOk")
        self.dashboard_pills[2].setObjectName("dashboardPillInfo")

        for pill in self.dashboard_pills:
            pill.setFixedHeight(32)
            pill.setMinimumWidth(110)
            pill.setMaximumWidth(170)
            pill.style().unpolish(pill)
            pill.style().polish(pill)

        metrics = compute_dashboard_financial_metrics(self.tb_rows_cache, self.current_period) if self.tb_rows_cache else {
            "aktif_toplam": 0.0,
            "pasif_toplam": 0.0,
            "net_kar": 0.0,
        }
        metric_defs = [
            ("Aktif Toplam", metrics["aktif_toplam"]),
            ("Pasif Toplam", metrics["pasif_toplam"]),
            ("Net Kar", metrics["net_kar"]),
        ]
        if hasattr(self, "dashboard_metric_cards"):
            for card, (title, value) in zip(self.dashboard_metric_cards, metric_defs):
                card["title"].setText(title)
                card["value"].setText(format_number(value))
                card["period"].setText(self.period_display_label(self.current_period))

    def refresh_analysis_ui(self):
        self.analysis_title.setText(self.t("analysisTitle"))
        self.analysis_desc.setText(self.t("analysisDesc"))
        self.search_input.setPlaceholderText(self.t("searchTb") if self.active_view == "tb" else self.t("searchPlcc"))
        if hasattr(self, "cmb_analysis_search_mode"):
            self.cmb_analysis_search_mode.blockSignals(True)
            self.cmb_analysis_search_mode.clear()
            self.cmb_analysis_search_mode.addItems(["İçinde Bul" if self.language == "tr" else "Contains", "Tam Eşleşme Bul" if self.language == "tr" else "Exact Match"])
            self.cmb_analysis_search_mode.setCurrentIndex(1 if getattr(self, "analysis_search_mode", "contains") == "exact" else 0)
            self.cmb_analysis_search_mode.blockSignals(False)
        self.period_info.setText(f"{self.t('periodCompare')}: {self.period_display_label(self.current_period)} → {self.period_display_label(self.previous_period)}")
        self.btn_export.setText(self.t("exportExcel"))
        self.btn_auto_fit.setText(self.t("autoFitColumns"))
        self.btn_view_tb.setText(self.t("tb"))
        self.btn_view_plcc.setText(self.t("plcc"))
        self.btn_show_all.setText(self.t("showAll"))
        self.btn_findings.setText(self.t("findingsOnly"))
        self.btn_tb_all.setText(self.t("allAccounts"))
        self.btn_tb_balance.setText(self.t("balanceSheet"))
        self.btn_tb_income.setText(self.t("incomeStatement"))
        self.toggle_info.setText(self.t("toggleInfo"))

        self.btn_view_tb.setObjectName("toggleButtonActive" if self.active_view == "tb" else "toggleButton")
        self.btn_view_plcc.setObjectName("toggleButtonActive" if self.active_view == "plcc" else "toggleButton")
        self.btn_show_all.setObjectName("filterSoftActive" if self.analysis_filter == "all" else "filterSoft")
        self.btn_findings.setObjectName("filterSoftActive" if self.analysis_filter == "findings" else "filterSoft")
        self.btn_tb_all.setObjectName("filterSoftActive" if self.tb_financial_filter == "all" else "filterSoft")
        self.btn_tb_balance.setObjectName("filterSoftActive" if self.tb_financial_filter == "balance" else "filterSoft")
        self.btn_tb_income.setObjectName("filterSoftActive" if self.tb_financial_filter == "income" else "filterSoft")

        perms = getattr(self, "current_user_permissions", normalize_user_permissions(None))
        tb_filter_visible = self.active_view == "tb" and perms.get("analysis_tb", True)
        self.btn_tb_all.setVisible(tb_filter_visible)
        self.btn_tb_balance.setVisible(tb_filter_visible and perms.get("analysis_balance", True))
        self.btn_tb_income.setVisible(tb_filter_visible and perms.get("analysis_income", True))

        for btn in [self.btn_view_tb, self.btn_view_plcc, self.btn_show_all, self.btn_findings, self.btn_tb_all, self.btn_tb_balance, self.btn_tb_income]:
            btn.style().unpolish(btn)
            btn.style().polish(btn)

        self.cmb_density.blockSignals(True)
        self.cmb_density.clear()
        self.cmb_density.addItems([self.t("compact"), self.t("normal"), self.t("wide")])
        self.cmb_density.setCurrentIndex({"compact": 0, "normal": 1, "wide": 2}[self.row_density])
        self.cmb_density.blockSignals(False)

    def refresh_analysis_tables(self):
        self.populate_tb_table()
        self.populate_plcc_table()
        self.tb_table.setVisible(self.active_view == "tb")
        self.plcc_table.setVisible(self.active_view == "plcc")
        self.apply_row_heights()



    def refresh_notes_ui(self):
        self.notes_title.setText(self.t("noteDefTitle"))
        self.notes_desc.setText(self.t("noteDefDesc"))
        self.note_paste_hint.setText("Format: Hesap(3 hane) | Ana Hesap(tam) | Masraf Yeri | Not. Ana Hesap doluysa önce ona göre eşleştirilir.")
        self.lbl_note_paste.setText(self.t("pasteArea"))
        self.input_note_paste.setHorizontalHeaderLabels([self.t("account"), "Ana Hesap", self.t("costCenter"), self.t("note")])
        self.btn_note_bulk_save.setText(self.t("bulkSave"))
        self.btn_note_paste_clear.setText(self.t("clearAll"))

        self.note_match_info.setText("Eşleşme önceliği: Ana Hesap > Hesap(3 hane). Masraf Yeri varsa ek filtre olarak uygulanır.")
        self.saved_notes_title.setText(self.t("savedNotes"))
        self.saved_notes_desc.setText(self.t("savedNotesDesc"))
        self.notes_control_info.setText(self.t("notesControlInfo"))
        self.btn_notes_save_changes.setText("Değişiklikleri Kaydet")
        self.btn_notes_delete_selected.setText("Seçili Notu Sil")

        self.cmb_notes_density.blockSignals(True)
        self.cmb_notes_density.clear()
        self.cmb_notes_density.addItems([self.t("compact"), self.t("normal"), self.t("wide")])
        self.cmb_notes_density.setCurrentIndex({"compact": 0, "normal": 1, "wide": 2}[self.notes_row_density])
        self.cmb_notes_density.blockSignals(False)

    def refresh_notes_table(self):
        headers = [self.t("account"), "Ana Hesap", self.t("costCenter"), self.t("note")]
        self.notes_table.clear()
        self.notes_table.setColumnCount(4)
        self.notes_table.setHorizontalHeaderLabels(headers)
        self.notes_table.setRowCount(len(self.notes))

        for r, row in enumerate(self.notes):
            set_table_item(self.notes_table, r, 0, str(normalize_hesap_prefix(row.get("hesap", ""))))
            set_table_item(self.notes_table, r, 1, str(row.get("anaHesap", "")))
            set_table_item(self.notes_table, r, 2, str(row.get("masrafYeri", "")) or self.t("general"))
            set_table_item(self.notes_table, r, 3, row.get("noteEn", "") if self.language == "en" else row.get("not", ""))

        for i, w in enumerate([110, 150, 120, 420]):
            self.notes_table.setColumnWidth(i, w)
        self.enable_manual_column_resize(self.notes_table)
        self.apply_notes_row_heights()

    def refresh_responsibles_ui(self):
        self.resp_title.setText(self.t("responsibleAssignTitle"))
        self.resp_desc.setText(self.t("responsibleAssignDesc"))
        self.lbl_resp_paste.setText(self.t("pasteArea"))
        self.input_resp_paste.setHorizontalHeaderLabels([self.t("account"), "Ana Hesap", self.t("responsible")])
        self.btn_resp_save.setText(self.t("bulkSave"))
        self.btn_resp_clear.setText(self.t("clearAll"))
        self.resp_format_info.setText("Eşleşme önceliği: Ana Hesap > Hesap(3 hane).")
        self.saved_resp_title.setText(self.t("savedResponsibles"))
        self.saved_resp_desc.setText(self.t("savedResponsiblesDesc"))
        self.btn_resp_save_changes.setText("Değişiklikleri Kaydet")
        self.btn_resp_delete_selected.setText("Seçili Sorumluyu Sil")

    def refresh_responsibles_table(self):
        headers = [self.t("account"), "Ana Hesap", self.t("responsible")]
        self.responsibles_table.clear()
        self.responsibles_table.setColumnCount(3)
        self.responsibles_table.setHorizontalHeaderLabels(headers)

        sorted_rows = sorted(
            self.responsibles,
            key=lambda x: (
                str(x.get("anaHesap", "")),
                normalize_hesap_prefix(x.get("hesap", "")),
                str(x.get("sorumlu", "")).lower()
            )
        )
        self.responsibles_table.setRowCount(len(sorted_rows))

        for r, row in enumerate(sorted_rows):
            set_table_item(self.responsibles_table, r, 0, normalize_hesap_prefix(row.get("hesap", "")))
            set_table_item(self.responsibles_table, r, 1, str(row.get("anaHesap", "")))
            set_table_item(self.responsibles_table, r, 2, str(row.get("sorumlu", "")))

        for i, w in enumerate([110, 150, 280]):
            self.responsibles_table.setColumnWidth(i, w)
        self.enable_manual_column_resize(self.responsibles_table)


    def load_regular_ft_faggl(self):
        self._load_regular_ft_file_async("faggl", "FAGGL Excel Seç" if self.language == "tr" else "Select FAGGL Excel")

    def load_regular_ft_eba(self):
        self._load_regular_ft_file_async("eba", "EBA Excel Seç" if self.language == "tr" else "Select EBA Excel")

    def load_regular_ft_zfi052(self):
        self._load_regular_ft_file_async("zfi052", "ZFI052 Excel Seç" if self.language == "tr" else "Select ZFI052 Excel")

    def toggle_regular_ft_risk_only(self):
        self.regular_ft_risk_only = self.btn_regular_ft_risk_only.isChecked()
        self.refresh_regular_ft_ui()
        self.refresh_regular_ft_table()

    def on_regular_ft_density_changed(self):
        idx = self.cmb_regular_ft_density.currentIndex()
        self.regular_ft_row_density = {0: "compact", 1: "normal", 2: "wide"}.get(idx, "normal")
        self.apply_regular_ft_row_heights()

    def apply_regular_ft_row_heights(self):
        if not hasattr(self, "regular_ft_table"):
            return
        height = 30 if self.regular_ft_row_density == "compact" else 48 if self.regular_ft_row_density == "wide" else 38
        for r in range(self.regular_ft_table.rowCount()):
            self.regular_ft_table.setRowHeight(r, height)

    def auto_fit_regular_ft_table(self):
        if not hasattr(self, "regular_ft_table") or self.regular_ft_table.columnCount() == 0:
            return
        header = self.regular_ft_table.horizontalHeader()
        for i in range(self.regular_ft_table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
            self.regular_ft_table.resizeColumnToContents(i)
            width = self.regular_ft_table.columnWidth(i)
            self.regular_ft_table.setColumnWidth(i, min(max(width + 18, 95), 280))
        self.enable_manual_column_resize(self.regular_ft_table)

    def refresh_regular_ft_ui(self):
        if not hasattr(self, "regular_ft_title"):
            return
        self.regular_ft_title.setText(self.t("regularFtAnalysis"))
        self.regular_ft_desc.setText(self.t("regularFtDesc"))
        self.btn_regular_ft_load_faggl.setText(f"FAGGL {self.t('loadData')}")
        self.btn_regular_ft_load_eba.setText(f"EBA {self.t('loadData')}")
        self.btn_regular_ft_load_zfi052.setText(f"ZFI052 {self.t('loadData')}")
        self.btn_regular_ft_start.setText(self.t("runAnalysis"))
        self.btn_regular_ft_risk_only.setText(self.t("showRiskyOnly"))
        self.btn_regular_ft_autofit.setText(self.t("autoFitColumns"))
        self.btn_regular_ft_export.setText(self.t("exportExcel"))
        self.lbl_regular_ft_currency.setText(self.t("currency"))
        self.lbl_regular_ft_user.setText(self.t("userName"))
        self.lbl_regular_ft_vendor.setText(self.t("vendorName"))
        self.lbl_regular_ft_eba_status.setText(self.t("ebaStatus"))
        self.lbl_regular_ft_search.setText(self.t("searchOutput"))
        self.lbl_regular_ft_search_mode.setText("Arama Tipi" if self.language == "tr" else "Search Type")
        self.lbl_regular_ft_current_period.setText(self.t("currentPeriod"))
        self.lbl_regular_ft_previous_period.setText(self.t("previousPeriod"))
        self.regular_ft_faggl_path_input.setText(self.regular_ft_faggl_path or "")
        self.regular_ft_eba_path_input.setText(self.regular_ft_eba_path or "")
        self.regular_ft_zfi052_path_input.setText(self.regular_ft_zfi052_path or "")
        self.btn_regular_ft_risk_only.setObjectName("filterSoftActive" if self.regular_ft_risk_only else "filterSoft")
        self.btn_regular_ft_risk_only.style().unpolish(self.btn_regular_ft_risk_only)
        self.btn_regular_ft_risk_only.style().polish(self.btn_regular_ft_risk_only)
        self.cmb_regular_ft_density.blockSignals(True)
        self.cmb_regular_ft_density.clear()
        self.cmb_regular_ft_density.addItems([self.t("compact"), self.t("normal"), self.t("wide")])
        self.cmb_regular_ft_density.setCurrentIndex({"compact": 0, "normal": 1, "wide": 2}.get(self.regular_ft_row_density, 1))
        self.cmb_regular_ft_density.blockSignals(False)
        if hasattr(self, "cmb_regular_ft_search_mode"):
            self.cmb_regular_ft_search_mode.blockSignals(True)
            self.cmb_regular_ft_search_mode.clear()
            self.cmb_regular_ft_search_mode.addItems(["İçinde Bul" if self.language == "tr" else "Contains", "Tam Eşleşme Bul" if self.language == "tr" else "Exact Match"])
            self.cmb_regular_ft_search_mode.setCurrentIndex(1 if getattr(self, "regular_ft_search_mode", "contains") == "exact" else 0)
            self.cmb_regular_ft_search_mode.blockSignals(False)
        self.fill_regular_ft_combos()
        self.fill_regular_ft_period_combos()

    def fill_regular_ft_combos(self):
        if not hasattr(self, "cmb_regular_ft_currency"):
            return
        df = self.regular_ft_output_df.copy() if isinstance(self.regular_ft_output_df, pd.DataFrame) else pd.DataFrame()
        all_label = self.regular_ft_all_label()
        currency_opts = [all_label] + sorted([x for x in df.get("Döviz Türü", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if x])
        user_opts = [all_label] + sorted([x for x in df.get("Kullanıcı Adı", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if x])
        vendor_opts = [all_label] + sorted([x for x in df.get("Satıcı Adı", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if x])
        eba_opts = [all_label] + sorted([x for x in df.get("EBA Son Durum", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if x])
        defs = [
            (self.cmb_regular_ft_currency, currency_opts, "regular_ft_currency_filter"),
            (self.cmb_regular_ft_user, user_opts, "regular_ft_user_filter"),
            (self.cmb_regular_ft_vendor, vendor_opts, "regular_ft_vendor_filter"),
            (self.cmb_regular_ft_eba_status, eba_opts, "regular_ft_eba_status_filter"),
        ]
        for combo, opts, attr in defs:
            combo.blockSignals(True)
            combo.clear()
            combo.addItems(opts)
            current_val = getattr(self, attr, all_label)
            idx = combo.findText(current_val)
            if idx < 0:
                setattr(self, attr, all_label)
                idx = 0
            combo.setCurrentIndex(idx)
            combo.blockSignals(False)


    def fill_regular_ft_period_combos(self):
        if not hasattr(self, "cmb_regular_ft_current_period"):
            return
        periods = list(self.regular_ft_periods or [])
        self.cmb_regular_ft_current_period.blockSignals(True)
        self.cmb_regular_ft_previous_period.blockSignals(True)
        self.cmb_regular_ft_current_period.clear()
        self.cmb_regular_ft_previous_period.clear()
        for period in periods:
            display = self.regular_ft_period_display(period)
            self.cmb_regular_ft_current_period.addItem(display, period)
            self.cmb_regular_ft_previous_period.addItem(display, period)
        if periods:
            if self.regular_ft_current_period not in periods:
                self.regular_ft_current_period = periods[-1]
            if self.regular_ft_previous_period not in periods:
                self.regular_ft_previous_period = periods[-2] if len(periods) >= 2 else periods[-1]
            cur_idx = max(0, self.cmb_regular_ft_current_period.findData(self.regular_ft_current_period))
            prev_idx = max(0, self.cmb_regular_ft_previous_period.findData(self.regular_ft_previous_period))
            self.cmb_regular_ft_current_period.setCurrentIndex(cur_idx)
            self.cmb_regular_ft_previous_period.setCurrentIndex(prev_idx)
        self.cmb_regular_ft_current_period.blockSignals(False)
        self.cmb_regular_ft_previous_period.blockSignals(False)

    def on_regular_ft_period_compare_changed(self):
        if not hasattr(self, "cmb_regular_ft_current_period") or self.cmb_regular_ft_current_period.count() == 0:
            return
        current_val = self.cmb_regular_ft_current_period.currentData() or self.cmb_regular_ft_current_period.currentText()
        previous_val = self.cmb_regular_ft_previous_period.currentData() or self.cmb_regular_ft_previous_period.currentText()
        if current_val:
            self.regular_ft_current_period = str(current_val)
        if previous_val:
            self.regular_ft_previous_period = str(previous_val)
        if self.regular_ft_current_period == self.regular_ft_previous_period and len(self.regular_ft_periods) >= 2:
            idx = self.regular_ft_periods.index(self.regular_ft_current_period)
            fallback = self.regular_ft_periods[idx - 1] if idx > 0 else self.regular_ft_periods[1]
            self.regular_ft_previous_period = fallback
            self.fill_regular_ft_period_combos()
        if self.regular_ft_analysis_ready:
            self.recalculate_regular_ft_current_fields()
            self.fill_regular_ft_combos()
            self.refresh_regular_ft_table()

    def recalculate_regular_ft_current_fields(self):
        base = self.regular_ft_base_output_df.copy() if isinstance(self.regular_ft_base_output_df, pd.DataFrame) else pd.DataFrame()
        if base.empty:
            self.regular_ft_output_df = pd.DataFrame()
            return
        periods = list(self.regular_ft_periods or [])
        for idx, row in base.iterrows():
            vendor = normalize_text_value(row.get("Satıcı", ""))
            currency = normalize_text_value(row.get("Döviz Türü", ""))
            current_user = self.regular_ft_user_period_map.get((vendor, currency, self.regular_ft_current_period), "")
            if not current_user:
                last_users = [self.regular_ft_user_period_map.get((vendor, currency, p), "") for p in periods]
                current_user = next((u for u in reversed(last_users) if normalize_text_value(u)), "")
            eba_info = self.regular_ft_eba_pending_period_dict.get((vendor, self.regular_ft_current_period), {})
            eba_adet = safe_int(eba_info.get("count", 0))
            eba_tutar = safe_float(eba_info.get("amount", 0.0))
            eba_son_durum = normalize_text_value(eba_info.get("status", ""))
            eba_kural = "Sadece tanımlı statüler" if eba_adet or abs(eba_tutar) > 1e-12 or eba_son_durum else "Statü eşleşmesi yok"
            current_eba_invoices = self.regular_ft_eba_invoice_period_dict.get((vendor, self.regular_ft_current_period), [])
            faggl_invoice_keys = self.regular_ft_faggl_invoice_keys_map.get((vendor, self.regular_ft_current_period), [])
            zfi_exact_from_eba = safe_unique_join([self.regular_ft_zfi_invoice_map.get((vendor, inv_no), "") for inv_no in current_eba_invoices], sep=" | ", limit=10)
            zfi_exact_from_faggl = safe_unique_join([self.regular_ft_zfi_invoice_map.get((vendor, inv_no), "") for inv_no in faggl_invoice_keys], sep=" | ", limit=10)
            zfi_period = self.regular_ft_zfi_period_map.get((vendor, self.regular_ft_current_period), "")
            zfi_vendor = self.regular_ft_zfi_vendor_map.get(vendor, "")
            zfi_muhatap = zfi_exact_from_eba or zfi_exact_from_faggl or zfi_period or zfi_vendor
            if zfi_exact_from_eba:
                zfi_kural = "EBA Fatura No exact"
            elif zfi_exact_from_faggl:
                zfi_kural = "FAGGL Belge Anahtarı exact"
            elif zfi_period:
                zfi_kural = "Dönem fallback"
            elif zfi_vendor:
                zfi_kural = "Satıcı fallback"
            else:
                zfi_kural = ""

            period_counts = [safe_int(row.get(f"{p} Adet", 0)) for p in periods]
            period_amounts = [safe_float(row.get(f"{p} Tutar", 0.0)) for p in periods]
            current_count = safe_int(row.get(f"{self.regular_ft_current_period} Adet", 0)) if self.regular_ft_current_period else 0
            prev_count = safe_int(row.get(f"{self.regular_ft_previous_period} Adet", 0)) if self.regular_ft_previous_period else 0
            current_amount = safe_float(row.get(f"{self.regular_ft_current_period} Tutar", 0.0)) if self.regular_ft_current_period else 0.0
            prev_amount = safe_float(row.get(f"{self.regular_ft_previous_period} Tutar", 0.0)) if self.regular_ft_previous_period else 0.0
            adet_degisim = calc_change_percent_text(current_count, prev_count) if self.regular_ft_previous_period else "-"
            tutar_degisim = calc_change_percent_text(current_amount, prev_amount) if self.regular_ft_previous_period else "-"
            analiz_notu = build_regular_note(period_counts, period_amounts, eba_adet, self.regular_ft_current_period, self.regular_ft_previous_period)
            risk_skoru, risk_durumu, risk_kurali = evaluate_regular_risk(analiz_notu, eba_adet, zfi_muhatap, adet_degisim, tutar_degisim, eba_kural, zfi_kural)

            base.at[idx, "Kullanıcı Adı"] = current_user
            base.at[idx, "EBA Adet"] = eba_adet
            base.at[idx, "EBA Tutar"] = eba_tutar
            base.at[idx, "EBA Son Durum"] = eba_son_durum
            base.at[idx, "EBA Kural"] = eba_kural
            base.at[idx, "ZFI052 Muhatap"] = zfi_muhatap
            base.at[idx, "ZFI052 Kural"] = zfi_kural
            base.at[idx, "Adet % Değişim"] = adet_degisim
            base.at[idx, "Tutar % Değişim"] = tutar_degisim
            base.at[idx, "Analiz Notu"] = analiz_notu
            base.at[idx, "Risk Skoru"] = risk_skoru
            base.at[idx, "Risk Durumu"] = risk_durumu
            base.at[idx, "Risk Kuralı"] = risk_kurali

        base["Riskli"] = base["Risk Durumu"].astype(str).isin(["Kritik", "Yüksek"])
        base = base.sort_values(by=["Riskli", "Risk Skoru", "Satıcı Adı", "Satıcı", "Döviz Türü"], ascending=[False, False, True, True, True]).reset_index(drop=True)
        self.regular_ft_output_df = base

    def start_regular_ft_analysis(self):
        if self.regular_ft_faggl_df is None or self.regular_ft_eba_df is None or self.regular_ft_zfi052_df is None:
            QMessageBox.warning(self, self.t("warning"), "FAGGL, EBA ve ZFI052 dosyalarının üçünü de yükleyin." if self.language == "tr" else "Load all three files: FAGGL, EBA, and ZFI052.")
            return
        self.begin_busy_state("Düzenli gelen fatura analizi hazırlanıyor..." if self.language == "tr" else "Preparing recurring invoice analysis...", 5)
        started = self.start_background_worker(
            self.background_build_regular_ft_analysis,
            self.on_regular_ft_analysis_loaded,
            self.on_regular_ft_analysis_failed,
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def prepare_regular_ft_analysis(self, progress_cb=None):
        faggl = self.regular_ft_faggl_df.copy()
        eba = self.regular_ft_eba_df.copy()
        zfi = self.regular_ft_zfi052_df.copy()

        _safe_progress(progress_cb, 15, "FAGGL verisi hazırlanıyor..." if self.language == "tr" else "Preparing FAGGL data...")
        faggl = faggl.rename(columns={c: str(c).strip() for c in faggl.columns})
        required_faggl = [
            "Mali yıl/dönem", "Belge para birimi değeri", "Belge para birimi anahtarı",
            "Satıcı", "Satıcı hesabı: Ad 1", "Kullanıcı adı"
        ]
        missing = [c for c in required_faggl if c not in faggl.columns]
        if missing:
            raise ValueError("FAGGL dosyasında eksik kolonlar: " + ", ".join(missing))

        for optional_col in ["Referans", "Belge numarası", "Belge metni", "Belge tarihi"]:
            if optional_col not in faggl.columns:
                faggl[optional_col] = ""

        faggl["Satıcı"] = faggl["Satıcı"].map(normalize_vendor_code)
        faggl["Satıcı Adı"] = faggl["Satıcı hesabı: Ad 1"].map(normalize_text_value)
        faggl["Kullanıcı adı"] = faggl["Kullanıcı adı"].map(normalize_text_value)
        faggl["Döviz Türü"] = faggl["Belge para birimi anahtarı"].map(normalize_text_value)
        faggl["Dönem Label"] = faggl["Mali yıl/dönem"].map(parse_fiscal_period_label)
        faggl["Tutar"] = faggl["Belge para birimi değeri"].map(safe_float)
        faggl["Belge Ref"] = faggl["Referans"].map(normalize_invoice_no)
        faggl["Belge No"] = faggl["Belge numarası"].map(normalize_invoice_no)
        faggl["Belge Anahtarı"] = faggl.apply(
            lambda x: x["Belge Ref"] or x["Belge No"] or f"ROW_{x.name}",
            axis=1
        )
        faggl = faggl[(faggl["Satıcı"] != "") & (faggl["Dönem Label"] != "")].copy()
        faggl["sort_key"] = faggl["Dönem Label"].map(regular_period_sort_key)
        faggl = faggl.sort_values(["Satıcı", "Döviz Türü", "sort_key", "Belge Anahtarı"]).reset_index(drop=True)

        periods = sorted(faggl["Dönem Label"].dropna().unique().tolist(), key=regular_period_sort_key)
        self.regular_ft_periods = periods
        self.regular_ft_current_period = periods[-1] if periods else ""
        self.regular_ft_previous_period = periods[-2] if len(periods) >= 2 else ""

        _safe_progress(progress_cb, 35, "EBA verisi hazırlanıyor..." if self.language == "tr" else "Preparing EBA data...")
        eba = eba.rename(columns={c: str(c).strip() for c in eba.columns})
        required_eba = ["Süreç Durumu", "Satıcı Kodu", "Fatura No", "Fatura Tarihi", "KDV'siz Fatura Tutarı"]
        missing = [c for c in required_eba if c not in eba.columns]
        if missing:
            raise ValueError("EBA dosyasında eksik kolonlar: " + ", ".join(missing))

        allowed_statuses = {
            "Fatura Sorumlusu Grubunda",
            "Asistan Onayında",
            "Fatura Onay Aşamasında",
            "Vergi Grubunda",
            "Muhasebe Son Kontrol Onayında",
        }

        for optional_col in ["Satıcı Adı"]:
            if optional_col not in eba.columns:
                eba[optional_col] = ""

        eba["Satıcı Kodu"] = eba["Satıcı Kodu"].map(normalize_vendor_code)
        eba["Fatura No"] = eba["Fatura No"].map(normalize_invoice_no)
        eba["EBA Dönem"] = eba["Fatura Tarihi"].map(month_year_tr_from_date)
        eba["KDV'siz Fatura Tutarı"] = eba["KDV'siz Fatura Tutarı"].map(safe_float)
        eba["Süreç Durumu"] = eba["Süreç Durumu"].map(normalize_text_value)
        eba["Satıcı Adı"] = eba["Satıcı Adı"].map(normalize_text_value)
        eba = eba[(eba["Satıcı Kodu"] != "") & (eba["Fatura No"] != "") & (eba["EBA Dönem"] != "")].copy()

        eba_pending = eba[eba["Süreç Durumu"].isin(allowed_statuses)].copy()
        eba_pending_unique = eba_pending.sort_values(["Satıcı Kodu", "EBA Dönem", "Fatura No"]).drop_duplicates(["Satıcı Kodu", "EBA Dönem", "Fatura No"], keep="last") if not eba_pending.empty else pd.DataFrame(columns=eba.columns)
        eba_all_unique = eba.sort_values(["Satıcı Kodu", "EBA Dönem", "Fatura No"]).drop_duplicates(["Satıcı Kodu", "EBA Dönem", "Fatura No"], keep="last") if not eba.empty else pd.DataFrame(columns=eba.columns)

        eba_pending_period_map = eba_pending_unique.groupby(["Satıcı Kodu", "EBA Dönem"]).agg(
            EBA_Adet_Pending=("Fatura No", "nunique"),
            EBA_Tutar_Pending=("KDV'siz Fatura Tutarı", "sum"),
            EBA_Son_Durum_Pending=("Süreç Durumu", lambda s: safe_unique_join(s.unique(), sep=" | ", limit=10)),
        ).reset_index() if not eba_pending_unique.empty else pd.DataFrame(columns=["Satıcı Kodu", "EBA Dönem", "EBA_Adet_Pending", "EBA_Tutar_Pending", "EBA_Son_Durum_Pending"])

        eba_all_period_map = eba_all_unique.groupby(["Satıcı Kodu", "EBA Dönem"]).agg(
            EBA_Adet_All=("Fatura No", "nunique"),
            EBA_Tutar_All=("KDV'siz Fatura Tutarı", "sum"),
            EBA_Son_Durum_All=("Süreç Durumu", lambda s: safe_unique_join(s.unique(), sep=" | ", limit=10)),
        ).reset_index() if not eba_all_unique.empty else pd.DataFrame(columns=["Satıcı Kodu", "EBA Dönem", "EBA_Adet_All", "EBA_Tutar_All", "EBA_Son_Durum_All"])

        eba_vendor_map = eba_all_unique.groupby(["Satıcı Kodu"]).agg(
            EBA_Adet_Vendor=("Fatura No", "nunique"),
            EBA_Tutar_Vendor=("KDV'siz Fatura Tutarı", "sum"),
            EBA_Son_Durum_Vendor=("Süreç Durumu", lambda s: safe_unique_join(s.unique(), sep=" | ", limit=10)),
        ).reset_index() if not eba_all_unique.empty else pd.DataFrame(columns=["Satıcı Kodu", "EBA_Adet_Vendor", "EBA_Tutar_Vendor", "EBA_Son_Durum_Vendor"])

        self.set_processing_message("ZFI052 verisi hazırlanıyor...", 50)
        zfi = zfi.rename(columns={c: str(c).strip() for c in zfi.columns})
        required_zfi = ["GIB ID", "Satıcı", "Sorumlu"]
        missing = [c for c in required_zfi if c not in zfi.columns]
        if missing:
            raise ValueError("ZFI052 dosyasında eksik kolonlar: " + ", ".join(missing))

        zfi["Satıcı"] = zfi["Satıcı"].map(normalize_vendor_code)
        zfi["GIB ID"] = zfi["GIB ID"].map(normalize_invoice_no)
        zfi["Sorumlu"] = zfi["Sorumlu"].map(normalize_text_value)
        zfi["ZFI Dönem"] = zfi["Belge Tarihi"].map(month_year_tr_from_date) if "Belge Tarihi" in zfi.columns else ""
        zfi = zfi[(zfi["Satıcı"] != "")].copy()

        zfi_invoice_map = zfi[(zfi["GIB ID"] != "")].groupby(["Satıcı", "GIB ID"])["Sorumlu"].agg(lambda s: safe_unique_join(s.unique(), sep=" | ", limit=10)).to_dict()
        zfi_period_map = zfi[zfi["ZFI Dönem"].astype(str).str.strip() != ""].groupby(["Satıcı", "ZFI Dönem"])["Sorumlu"].agg(lambda s: safe_unique_join(s.unique(), sep=" | ", limit=10)).to_dict() if "ZFI Dönem" in zfi.columns else {}
        zfi_vendor_map = zfi.groupby(["Satıcı"])["Sorumlu"].agg(lambda s: safe_unique_join(s.unique(), sep=" | ", limit=10)).to_dict()

        self.set_processing_message("Dönemsel çıktı oluşturuluyor...", 70)
        agg = faggl.groupby(["Satıcı", "Satıcı Adı", "Döviz Türü", "Dönem Label"], as_index=False).agg(
            Adet=("Belge Anahtarı", "nunique"),
            Tutar=("Tutar", "sum"),
            Kullanıcı_Adı=("Kullanıcı adı", lambda s: list([x for x in s if str(x).strip()])[-1] if any(str(x).strip() for x in s) else ""),
        )

        self.regular_ft_user_period_map = {
            (normalize_text_value(r["Satıcı"]), normalize_text_value(r["Döviz Türü"]), normalize_text_value(r["Dönem Label"])): normalize_text_value(r["Kullanıcı_Adı"])
            for _, r in agg.iterrows()
        }
        self.regular_ft_eba_pending_period_dict = {
            (normalize_text_value(r["Satıcı Kodu"]), normalize_text_value(r["EBA Dönem"])): {
                "count": safe_int(r["EBA_Adet_Pending"]),
                "amount": safe_float(r["EBA_Tutar_Pending"]),
                "status": normalize_text_value(r["EBA_Son_Durum_Pending"]),
            }
            for _, r in eba_pending_period_map.iterrows()
        } if not eba_pending_period_map.empty else {}
        self.regular_ft_eba_invoice_period_dict = {}
        if not eba_all_unique.empty:
            for (vendor, period), sub in eba_all_unique.groupby(["Satıcı Kodu", "EBA Dönem"]):
                self.regular_ft_eba_invoice_period_dict[(normalize_text_value(vendor), normalize_text_value(period))] = sub["Fatura No"].dropna().astype(str).unique().tolist()
        self.regular_ft_faggl_invoice_keys_map = {}
        for (vendor, period), sub in faggl.groupby(["Satıcı", "Dönem Label"]):
            self.regular_ft_faggl_invoice_keys_map[(normalize_text_value(vendor), normalize_text_value(period))] = sub["Belge Anahtarı"].dropna().astype(str).unique().tolist()
        self.regular_ft_zfi_invoice_map = dict(zfi_invoice_map)
        self.regular_ft_zfi_period_map = dict(zfi_period_map)
        self.regular_ft_zfi_vendor_map = dict(zfi_vendor_map)

        rows = []
        base_groups = agg.groupby(["Satıcı", "Satıcı Adı", "Döviz Türü"], sort=False)
        for (vendor, vendor_name, currency), grp in base_groups:
            period_count_map = {r["Dönem Label"]: safe_int(r["Adet"]) for _, r in grp.iterrows()}
            period_amount_map = {r["Dönem Label"]: safe_float(r["Tutar"]) for _, r in grp.iterrows()}
            row = {
                "Satıcı": vendor,
                "Satıcı Adı": vendor_name,
                "Döviz Türü": currency,
                "Kullanıcı Adı": "",
                "EBA Adet": 0,
                "EBA Tutar": 0.0,
                "EBA Son Durum": "",
                "EBA Kural": "",
                "ZFI052 Muhatap": "",
                "ZFI052 Kural": "",
                "Adet % Değişim": "-",
                "Tutar % Değişim": "-",
                "Analiz Notu": "",
                "Risk Skoru": 0,
                "Risk Durumu": "Normal",
                "Risk Kuralı": "",
            }
            for period in periods:
                row[f"{period} Adet"] = safe_int(period_count_map.get(period, 0))
                row[f"{period} Tutar"] = safe_float(period_amount_map.get(period, 0.0))
            rows.append(row)

        output = pd.DataFrame(rows)
        if output.empty:
            self.regular_ft_base_output_df = output
            self.regular_ft_output_df = output
            return

        self.regular_ft_base_output_df = output
        self.recalculate_regular_ft_current_fields()
        _safe_progress(progress_cb, 92, "Düzenli gelen fatura analizi tamamlandı..." if self.language == "tr" else "Recurring invoice analysis completed...")

    def get_filtered_regular_ft_df(self):
        df = self.regular_ft_output_df.copy() if isinstance(self.regular_ft_output_df, pd.DataFrame) else pd.DataFrame()
        if df.empty:
            return df
        self.regular_ft_currency_filter = self.cmb_regular_ft_currency.currentText() or self.regular_ft_all_label()
        self.regular_ft_user_filter = self.cmb_regular_ft_user.currentText() or self.regular_ft_all_label()
        self.regular_ft_vendor_filter = self.cmb_regular_ft_vendor.currentText() or self.regular_ft_all_label()
        self.regular_ft_eba_status_filter = self.cmb_regular_ft_eba_status.currentText() or self.regular_ft_all_label()
        if hasattr(self, "cmb_regular_ft_current_period") and self.cmb_regular_ft_current_period.count():
            self.regular_ft_current_period = str(self.cmb_regular_ft_current_period.currentData() or self.regular_ft_current_period)
        if hasattr(self, "cmb_regular_ft_previous_period") and self.cmb_regular_ft_previous_period.count():
            self.regular_ft_previous_period = str(self.cmb_regular_ft_previous_period.currentData() or self.regular_ft_previous_period)
        self.regular_ft_search_text = self.input_regular_ft_search.text().strip().lower()

        if self.regular_ft_currency_filter != self.regular_ft_all_label():
            df = df[df["Döviz Türü"] == self.regular_ft_currency_filter]
        if self.regular_ft_user_filter != self.regular_ft_all_label():
            df = df[df["Kullanıcı Adı"] == self.regular_ft_user_filter]
        if self.regular_ft_vendor_filter != self.regular_ft_all_label():
            df = df[df["Satıcı Adı"] == self.regular_ft_vendor_filter]
        if self.regular_ft_eba_status_filter != self.regular_ft_all_label():
            df = df[df["EBA Son Durum"].astype(str).str.contains(re.escape(self.regular_ft_eba_status_filter), na=False)]
        if self.regular_ft_risk_only:
            df = df[df["Riskli"] == True]
        if self.regular_ft_search_text:
            mask = df_text_search_mask(
                df,
                ["Satıcı", "Satıcı Adı", "Döviz Türü", "Kullanıcı Adı", "EBA Son Durum", "ZFI052 Muhatap", "Analiz Notu", "Risk Durumu", "Risk Kuralı"],
                self.regular_ft_search_text,
                getattr(self, "regular_ft_search_mode", "contains"),
            )
            df = df[mask]
        return df.copy()

    def refresh_regular_ft_table(self):
        if not hasattr(self, "regular_ft_table"):
            return
        df = self.get_filtered_regular_ft_df()
        if df.empty:
            self.populate_simple_table(self.regular_ft_table, [self.regular_ft_column_label("Bilgi")], [[self.t("noData")]])
            for card, title in zip(self.regular_ft_metric_cards, [self.t("vendors"), self.t("period"), self.t("riskyRecords"), self.t("ebaRecords")]):
                card["title"].setText(title)
                card["value"].setText("0")
                card["sub"].setText("-")
            self.regular_ft_info_pill.setText(self.t("readyNoData"))
            return

        internal_headers = ["Satıcı", "Satıcı Adı", "Döviz Türü", "Kullanıcı Adı"]
        for period in self.regular_ft_periods:
            internal_headers.extend([f"{period} Adet", f"{period} Tutar"])
        internal_headers.extend([
            "EBA Adet", "EBA Tutar", "EBA Son Durum", "EBA Kural",
            "ZFI052 Muhatap", "ZFI052 Kural", "Adet % Değişim", "Tutar % Değişim",
            "Analiz Notu", "Risk Skoru", "Risk Durumu", "Risk Kuralı"
        ])
        display_headers = [self.regular_ft_column_label(h) for h in internal_headers]

        rows = []
        badges = []
        for _, r in df.iterrows():
            row = []
            badge_row = []
            for h in internal_headers:
                val = r.get(h, "")
                sval = format_number(val) if h.endswith("Tutar") else str(val)
                if self.language == "en" and h == "Risk Durumu":
                    sval = {"Kritik": "Critical", "Yüksek": "High", "Orta": "Medium", "Normal": "Normal"}.get(sval, sval)
                row.append(sval)
                badge = None
                if h in ["EBA Son Durum", "EBA Kural", "ZFI052 Muhatap", "ZFI052 Kural", "Analiz Notu", "Adet % Değişim", "Tutar % Değişim", "Risk Skoru", "Risk Durumu", "Risk Kuralı"]:
                    badge = risk_badge_from_text(sval)
                if self.language == "en" and h == "Analiz Notu" and sval:
                    sval = sval.replace("Son Dönem Ft.Yok", "No Invoice in Current Period").replace("Yeni", "New").replace("Aralıklı Gelen", "Intermittent Flow").replace("EBA Kayıt Var", "EBA Record Exists").replace("Adet Beklenti Farkı", "Count Expectation Gap").replace("Tutar Sapması", "Amount Variance")
                if h == "Analiz Notu" and sval:
                    badge = "danger"
                if h == "Risk Skoru":
                    badge = numeric_risk_badge(parse_sort_value(sval))
                if h == "Risk Durumu":
                    if sval in ["Kritik", "Critical"]:
                        badge = "danger"
                    elif sval in ["Yüksek", "Orta", "High", "Medium"]:
                        badge = "warn"
                    else:
                        badge = "success"
                badge_row.append(badge)
            rows.append(row)
            badges.append(badge_row)

        self.populate_simple_table(self.regular_ft_table, display_headers, rows, badge_matrix=badges)
        self.apply_regular_ft_row_heights()
        rows_txt = "satır" if self.language == "tr" else "rows"
        self.regular_ft_info_pill.setText(f"{len(df)} {rows_txt} • {self.t('currentPeriod')}: {self.regular_ft_period_display(self.regular_ft_current_period) or '-'}")

        total_vendors = df["Satıcı"].nunique()
        risk_count = int(df["Riskli"].sum()) if "Riskli" in df.columns else 0
        total_eba = int(df["EBA Adet"].fillna(0).astype(float).sum()) if "EBA Adet" in df.columns else 0
        self.regular_ft_metric_cards[0]["title"].setText(self.t("vendors"))
        self.regular_ft_metric_cards[0]["value"].setText(str(total_vendors))
        self.regular_ft_metric_cards[0]["sub"].setText(self.t("filteredUniqueVendors"))
        self.regular_ft_metric_cards[1]["title"].setText(self.t("period"))
        self.regular_ft_metric_cards[1]["value"].setText(self.regular_ft_period_display(self.regular_ft_current_period) or "-")
        self.regular_ft_metric_cards[1]["sub"].setText(f"{self.t('previous')}: {self.regular_ft_period_display(self.regular_ft_previous_period) or '-'}")
        avg_risk = float(df["Risk Skoru"].fillna(0).astype(float).mean()) if "Risk Skoru" in df.columns else 0.0
        self.regular_ft_metric_cards[2]["title"].setText(self.t("riskyRecords"))
        self.regular_ft_metric_cards[2]["value"].setText(str(risk_count))
        self.regular_ft_metric_cards[2]["sub"].setText(f"{self.t('averageScore')}: {avg_risk:,.1f}".replace(",", "X").replace(".", ",").replace("X", "."))
        self.regular_ft_metric_cards[3]["title"].setText(self.t("ebaRecords"))
        self.regular_ft_metric_cards[3]["value"].setText(str(total_eba))
        self.regular_ft_metric_cards[3]["sub"].setText(self.t("currentPeriodEbaTotal"))

    def export_regular_ft_analysis(self):
        if self.regular_ft_table.rowCount() == 0:
            QMessageBox.warning(self, self.t("warning"), self.t("noExportData"))
            return
        default_name = build_versioned_filename("Recurring_Invoice_Analysis" if self.language == "en" else "Duzenli_Gelen_Ft_Analiz")
        path, _ = QFileDialog.getSaveFileName(self, "Excel Çıktısı Kaydet", default_name, "Excel Files (*.xlsx)")
        if not path:
            return
        if not str(path).lower().endswith(".xlsx"):
            path = f"{path}.xlsx"

        headers = [self.regular_ft_table.horizontalHeaderItem(i).text() for i in range(self.regular_ft_table.columnCount())]
        rows = []
        for r in range(self.regular_ft_table.rowCount()):
            rows.append([
                self.regular_ft_table.item(r, c).text() if self.regular_ft_table.item(r, c) else ""
                for c in range(self.regular_ft_table.columnCount())
            ])

        self.begin_busy_state("Düzenli gelen fatura Excel çıktısı hazırlanıyor..." if self.language == "tr" else "Preparing recurring invoice Excel export...", 15)
        started = self.start_background_worker(
            write_single_sheet_excel,
            self.on_background_export_success,
            self.on_background_export_failed,
            path,
            headers,
            rows,
            "Recurring_Invoice_Analysis" if self.language == "en" else "Duzenli_Gelen_Ft_Analiz",
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def get_permission_labels(self) -> Dict[str, str]:
        tr = self.language == "tr"
        return {
            "dashboard": "Dashboard",
            "analysis": "Analiz" if tr else "Analysis",
            "analysis_tb": "Analiz > TB" if tr else "Analysis > TB",
            "analysis_plcc": "Analiz > PL-CC" if tr else "Analysis > PL-CC",
            "analysis_balance": "Analiz > Bilanço" if tr else "Analysis > Balance Sheet",
            "analysis_income": "Analiz > Gelir Tablosu" if tr else "Analysis > Income Statement",
            "notes": "Not Tanımları" if tr else "Note Definitions",
            "responsibles": "Sorumlu Tayin" if tr else "Responsible Assignment",
            "muavin": "Muavin Analiz" if tr else "Subledger Analysis",
            "muavin_user_based": "Muavin > Kullanıcı Bazlı Analiz" if tr else "Subledger > User-Based Analysis",
            "muavin_tax_based": "Muavin > Vergisel Analiz" if tr else "Subledger > Tax Analysis",
            "muavin_account_content": "Muavin > Hesap İçerik Analiz" if tr else "Subledger > Account Content Analysis",
            "muavin_finding": "Muavin > Bulgu Metinleri" if tr else "Subledger > Findings",
            "muavin_text": "Muavin > Metin / Açıklama Analizi" if tr else "Subledger > Text / Description Analysis",
            "muavin_period": "Muavin > Dönem Özeti" if tr else "Subledger > Period Summary",
            "muavin_doctype": "Muavin > Belge Türü Özeti" if tr else "Subledger > Document Type Summary",
            "muavin_risk_user": "Muavin > Riskli Kullanıcılar" if tr else "Subledger > High-Risk Users",
            "muavin_user": "Muavin > Kullanıcı x Belge Türü" if tr else "Subledger > User x Document Type",
            "muavin_contra": "Muavin > Satıcı / Cari" if tr else "Subledger > Vendor / Offset Account",
            "muavin_drilldown": "Muavin > Satıcı / Cari Detayı" if tr else "Subledger > Vendor / Offset Detail",
            "muavin_document_lines": "Muavin > Belge Alt Satırları" if tr else "Subledger > Document Lines",
            "muavin_account_doc_relation": "Muavin > Hesap + Belge İlişkisi" if tr else "Subledger > Account + Document Relation",
            "muavin_late7": "Muavin > 7 Gün Sonrası İşlemler" if tr else "Subledger > Entries After Day 7",
            "muavin_doccheck": "Muavin > Denk./Ters Kayıt Kontrolü" if tr else "Subledger > Clearing / Reversal Control",
            "muavin_dupref": "Muavin > Mükerrer Referanslar" if tr else "Subledger > Duplicate References",
            "muavin_duprefdetail": "Muavin > Referans Detay Liste" if tr else "Subledger > Reference Detail List",
            "muavin_taxref": "Muavin > Vergisel Kontrol - Referans" if tr else "Subledger > Tax Control - Reference",
            "muavin_taxvendor": "Muavin > Vergisel Kontrol - Satıcı" if tr else "Subledger > Tax Control - Vendor",
            "muavin_tax": "Muavin > Vergi Göstergesi Analizi" if tr else "Subledger > Tax Indicator Analysis",
            "muavin_cost": "Muavin > Masraf Yeri ve Trend" if tr else "Subledger > Cost Center and Trend",
            "regular_ft": "Düzenli Gelen Ft Analiz" if tr else "Recurring Invoice Analysis",
            "user_management": "Kullanıcı Yönetimi" if tr else "User Management",
        }

    def get_user_permissions(self, username: str) -> Dict[str, bool]:
        username = str(username).strip().lower()
        for item in load_users():
            if str(item.get("username", "")).strip().lower() == username:
                return normalize_user_permissions(item.get("permissions"))
        return normalize_user_permissions(None)

    def collect_permission_form_values(self) -> Dict[str, bool]:
        values = normalize_user_permissions(None)
        for key, chk in getattr(self, "user_permission_checks", {}).items():
            values[key] = bool(chk.isChecked())
        return values

    def set_permission_form_values(self, permissions: Optional[Dict] = None):
        perms = normalize_user_permissions(permissions)
        if not hasattr(self, "user_permission_checks"):
            return
        for key, chk in self.user_permission_checks.items():
            chk.blockSignals(True)
            chk.setChecked(bool(perms.get(key, True)))
            chk.blockSignals(False)

    def apply_user_permissions(self):
        self.current_user_permissions = self.get_user_permissions(self.current_user)
        perms = self.current_user_permissions
        is_admin = self.is_current_user_admin()
        if hasattr(self, "tabs"):
            try:
                self.tabs.setTabVisible(0, perms.get("dashboard", True))
                self.tabs.setTabVisible(1, perms.get("analysis", True))
                self.tabs.setTabVisible(2, perms.get("notes", True))
                self.tabs.setTabVisible(3, perms.get("responsibles", True))
                self.tabs.setTabVisible(4, perms.get("muavin", True))
                self.tabs.setTabVisible(5, perms.get("regular_ft", True))
                self.tabs.setTabVisible(6, is_admin and perms.get("user_management", True))
            except Exception:
                pass
            current_idx = self.tabs.currentIndex()
            if hasattr(self.tabs, "isTabVisible"):
                if current_idx < 0 or not self.tabs.isTabVisible(current_idx):
                    for idx in range(self.tabs.count()):
                        try:
                            visible = self.tabs.isTabVisible(idx)
                        except Exception:
                            visible = True
                        if visible:
                            self.tabs.setCurrentIndex(idx)
                            break
        if hasattr(self, "btn_note_defs"):
            self.btn_note_defs.setVisible(perms.get("notes", True))
            self.btn_responsibles.setVisible(perms.get("responsibles", True))
            self.btn_muavin.setVisible(perms.get("muavin", True))
            self.btn_regular_ft.setVisible(perms.get("regular_ft", True))
        if hasattr(self, "btn_view_tb"):
            self.btn_view_tb.setVisible(perms.get("analysis_tb", True))
            self.btn_view_plcc.setVisible(perms.get("analysis_plcc", True))
            self.btn_tb_balance.setVisible(self.active_view == "tb" and perms.get("analysis_balance", True))
            self.btn_tb_income.setVisible(self.active_view == "tb" and perms.get("analysis_income", True))
            if self.active_view == "tb" and not perms.get("analysis_tb", True) and perms.get("analysis_plcc", True):
                self.active_view = "plcc"
            elif self.active_view == "plcc" and not perms.get("analysis_plcc", True) and perms.get("analysis_tb", True):
                self.active_view = "tb"
            if self.tb_financial_filter == "balance" and not perms.get("analysis_balance", True):
                self.tb_financial_filter = "all"
            if self.tb_financial_filter == "income" and not perms.get("analysis_income", True):
                self.tb_financial_filter = "all"

        if hasattr(self, "muavin_tables"):
            muavin_main = perms.get("muavin", True)
            group_perms = self.get_muavin_group_permissions(perms)
            if hasattr(self, "muavin_info_pill"):
                self.muavin_info_pill.setVisible(muavin_main)
            if hasattr(self, "btn_load_muavin"):
                self.btn_load_muavin.setVisible(muavin_main)
                self.btn_start_muavin.setVisible(muavin_main)
                self.btn_export_muavin.setVisible(muavin_main)
                self.btn_muavin_risk_only.setVisible(muavin_main)
                self.btn_muavin_cost_alarm_only.setVisible(muavin_main)
                self.btn_muavin_autofit.setVisible(muavin_main)
                self.cmb_muavin_density.setVisible(muavin_main)
                self.muavin_file_label.setVisible(muavin_main)
                if hasattr(self, "btn_muavin_expand_all"):
                    self.btn_muavin_expand_all.setVisible(False)
                if hasattr(self, "btn_muavin_collapse_all"):
                    self.btn_muavin_collapse_all.setVisible(False)
            if hasattr(self, "muavin_mapping_frame"):
                self.muavin_mapping_frame.setVisible(muavin_main)
            if hasattr(self, "muavin_required_columns_info"):
                self.muavin_required_columns_info.setVisible(muavin_main)
            if hasattr(self, "muavin_loaded_columns_info"):
                self.muavin_loaded_columns_info.setVisible(muavin_main)
            if hasattr(self, "cmb_muavin_account"):
                for widget in [
                    self.lbl_muavin_account, self.lbl_muavin_statement, self.lbl_muavin_period, self.lbl_muavin_doc_type,
                    self.lbl_muavin_user, self.lbl_muavin_contra, self.lbl_muavin_cost, self.lbl_muavin_search,
                    self.cmb_muavin_account, self.cmb_muavin_statement, self.cmb_muavin_period, self.cmb_muavin_doc_type,
                    self.cmb_muavin_user, self.cmb_muavin_contra, self.cmb_muavin_cost, self.input_muavin_search,
                ]:
                    widget.setVisible(muavin_main)
            if hasattr(self, "muavin_metric_cards"):
                for card in self.muavin_metric_cards:
                    parent = card["title"].parentWidget()
                    if parent is not None:
                        parent.setVisible(muavin_main)
            if hasattr(self, "muavin_result_buttons"):
                for key, btn in self.muavin_result_buttons.items():
                    btn.setVisible(group_perms.get(key, True))
            if hasattr(self, "muavin_tables"):
                current_view = getattr(self, "muavin_current_view", "user_based")
                if not group_perms.get(current_view, False):
                    for candidate in ["user_based", "tax_based", "account_content"]:
                        if group_perms.get(candidate, False):
                            current_view = candidate
                            break
                    self.muavin_current_view = current_view
                for key, table in self.muavin_tables.items():
                    table.setVisible(muavin_main and group_perms.get(key, True) and key == self.muavin_current_view)
            if hasattr(self, "muavin_result_info"):
                self.muavin_result_info.setVisible(muavin_main)
            if hasattr(self, "muavin_desc"):
                self.muavin_desc.setText(
                    ("Kullanıcı, vergisel ve hesap içerik analizlerini üç ayrı sonuç görünümünde izleyebilirsiniz." if self.language == "tr"
                     else "You can review user-based, tax-based, and account-content analyses in three result views.")
                )


    def _set_password_echo_mode(self, line_edit, button, visible: bool):
        if line_edit is None or button is None:
            return
        line_edit.setEchoMode(QLineEdit.Normal if visible else QLineEdit.Password)
        tr = self.language == "tr"
        button.setText(("Gizle" if tr else "Hide") if visible else ("Göster" if tr else "Show"))

    def toggle_password_visibility(self, target: str):
        mapping = {
            "user_password": (getattr(self, "input_user_password", None), getattr(self, "btn_toggle_user_password", None)),
            "user_password_confirm": (getattr(self, "input_user_password_confirm", None), getattr(self, "btn_toggle_user_password_confirm", None)),
            "login_password": (getattr(self, "input_login_password", None), getattr(self, "btn_toggle_login_password", None)),
        }
        line_edit, button = mapping.get(target, (None, None))
        if line_edit is None or button is None:
            return
        visible = line_edit.echoMode() != QLineEdit.Normal
        self._set_password_echo_mode(line_edit, button, visible)


    def refresh_users_ui(self):
        is_admin = self.is_current_user_admin()
        tr = self.language == "tr"
        self.users_title.setText("Kullanıcı Yönetimi" if tr else "User Management")
        self.users_desc.setText("Programa giriş için kullanıcı adı ve şifre yönetin. Varsayılan ilk giriş admin / admin olarak tanımlıdır." if tr else "Manage usernames and passwords for program access. The default first login is admin / admin.")
        self.users_current_info.setText((f"Aktif kullanıcı: {self.current_user}" if tr else f"Active user: {self.current_user}"))
        self.lbl_user_username.setText("Kullanıcı Adı" if tr else "Username")
        self.lbl_user_password.setText("Şifre" if tr else "Password")
        self.lbl_user_password_confirm.setText("Şifre Tekrar" if tr else "Confirm Password")
        self.users_perm_title.setText("Sekme, Alt Sekme ve Muavin Blok Yetkileri" if tr else "Tab, Sub-Tab, and Subledger Block Permissions")
        self.users_perm_desc.setText("İşaretli alanlar kullanıcı tarafından görüntülenir ve kullanılabilir. Muavin içindeki alt analiz blokları da ayrı ayrı yetkilendirilebilir." if tr else "Checked areas are visible and usable by the user. Subledger analysis blocks can also be authorized separately.")
        labels = self.get_permission_labels()
        for key, chk in self.user_permission_checks.items():
            chk.setText(labels.get(key, key))
        self.btn_user_add.setText("Kullanıcıyı Kaydet / Güncelle" if tr else "Save / Update User")
        self.btn_user_clear.setText("Temizle" if tr else "Clear")
        self.btn_user_delete.setText("Seçili Kullanıcıyı Sil" if tr else "Delete Selected User")
        self.users_note.setText("Not: Yalnızca admin kullanıcısı yeni kullanıcı tanımlayabilir, şifre güncelleyebilir ve kullanıcı silebilir." if tr else "Note: Only the admin user can add new users, update passwords, and delete users.")
        self.users_list_title.setText("Kayıtlı Kullanıcılar" if tr else "Registered Users")
        self.users_list_desc.setText("Listeden kullanıcı seçtiğinizde form otomatik dolar." if tr else "When you select a user from the list, the form is filled automatically.")
        self._set_password_echo_mode(self.input_user_password, self.btn_toggle_user_password, self.input_user_password.echoMode() == QLineEdit.Normal)
        self._set_password_echo_mode(self.input_user_password_confirm, self.btn_toggle_user_password_confirm, self.input_user_password_confirm.echoMode() == QLineEdit.Normal)
        for w in [self.input_user_username, self.input_user_password, self.input_user_password_confirm, self.btn_user_add, self.btn_user_delete, self.btn_toggle_user_password, self.btn_toggle_user_password_confirm]:
            w.setEnabled(is_admin)
        for chk in self.user_permission_checks.values():
            chk.setEnabled(is_admin)
        self.btn_user_clear.setEnabled(True)
        if hasattr(self, "users_perm_frame"):
            self.users_perm_frame.setVisible(is_admin)
            self.users_perm_title.setVisible(is_admin)
            self.users_perm_desc.setVisible(is_admin)
        if not is_admin:
            self.users_note.setText("Bu alan yalnızca admin kullanıcısı tarafından yönetilebilir." if tr else "This area can only be managed by the admin user.")

    def refresh_users_table(self):
        self.user_records = load_users()
        headers = ["Kullanıcı Adı", "Rol", "Yetkili Alanlar"] if self.language == "tr" else ["Username", "Role", "Authorized Areas"]
        self.users_table.clear()
        self.users_table.setColumnCount(3)
        self.users_table.setHorizontalHeaderLabels(headers)
        self.users_table.setRowCount(len(self.user_records))
        labels = self.get_permission_labels()
        for r, row in enumerate(sorted(self.user_records, key=lambda x: x.get("username", "").lower())):
            set_table_item(self.users_table, r, 0, str(row.get("username", "")))
            role = "Admin" if row.get("is_admin") else ("Kullanıcı" if self.language == "tr" else "User")
            set_table_item(self.users_table, r, 1, role, badge="info" if row.get("is_admin") else None)
            perms = normalize_user_permissions(row.get("permissions"))
            active_labels = [labels.get(k, k) for k, v in perms.items() if v]
            set_table_item(self.users_table, r, 2, safe_unique_join(active_labels, limit=14))
        self.users_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.users_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.users_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        for r in range(self.users_table.rowCount()):
            self.users_table.setRowHeight(r, 38)

    def is_current_user_admin(self) -> bool:
        for item in load_users():
            if str(item.get("username", "")).strip().lower() == str(self.current_user).strip().lower():
                return bool(item.get("is_admin", False))
        return False

    def clear_user_form(self):
        self.input_user_username.clear()
        self.input_user_password.clear()
        self.input_user_password_confirm.clear()
        self._set_password_echo_mode(self.input_user_password, self.btn_toggle_user_password, False)
        self._set_password_echo_mode(self.input_user_password_confirm, self.btn_toggle_user_password_confirm, False)
        self.set_permission_form_values(normalize_user_permissions(None))
        self.users_table.clearSelection()

    def on_user_table_clicked(self, row: int, col: int):
        item = self.users_table.item(row, 0)
        if item is None:
            return
        username = item.text()
        self.input_user_username.setText(username)
        self.input_user_password.clear()
        self.input_user_password_confirm.clear()
        self._set_password_echo_mode(self.input_user_password, self.btn_toggle_user_password, False)
        self._set_password_echo_mode(self.input_user_password_confirm, self.btn_toggle_user_password_confirm, False)
        self.set_permission_form_values(self.get_user_permissions(username))

    def add_or_update_user(self):
        if not self.is_current_user_admin():
            QMessageBox.warning(self, self.t("warning"), "Sadece admin kullanıcı yönetimi yapabilir." if self.language == "tr" else "Only the admin user can manage users.")
            return
        username = self.input_user_username.text().strip()
        password = self.input_user_password.text()
        password_confirm = self.input_user_password_confirm.text()
        if not username:
            QMessageBox.warning(self, self.t("warning"), "Kullanıcı adı boş olamaz." if self.language == "tr" else "Username cannot be empty.")
            return
        if not password:
            QMessageBox.warning(self, self.t("warning"), "Şifre boş olamaz." if self.language == "tr" else "Password cannot be empty.")
            return
        if password != password_confirm:
            QMessageBox.warning(self, self.t("warning"), "Şifreler eşleşmiyor." if self.language == "tr" else "Passwords do not match.")
            return
        users = load_users()
        updated = False
        permissions = self.collect_permission_form_values()
        if username.lower() == "admin":
            permissions["user_management"] = True
        else:
            permissions["user_management"] = False
        for user in users:
            if str(user.get("username", "")).strip().lower() == username.lower():
                user["password_hash"] = hash_password(password)
                user["permissions"] = permissions
                updated = True
                break
        if not updated:
            users.append({"username": username, "password_hash": hash_password(password), "is_admin": username.lower() == "admin", "permissions": permissions})
        save_users(users)
        if username.lower() == str(self.current_user).strip().lower():
            self.current_user_permissions = permissions
        self.user_records = load_users()
        self.refresh_users_table()
        self.clear_user_form()
        QMessageBox.information(self, "Bilgi" if self.language == "tr" else "Info", ("Kullanıcı kaydedildi / güncellendi." if self.language == "tr" else "User saved / updated."))

    def delete_selected_user(self):
        if not self.is_current_user_admin():
            QMessageBox.warning(self, self.t("warning"), "Sadece admin kullanıcı yönetimi yapabilir." if self.language == "tr" else "Only the admin user can manage users.")
            return
        username = self.input_user_username.text().strip()
        if not username:
            QMessageBox.warning(self, self.t("warning"), "Silmek için kullanıcı seçin." if self.language == "tr" else "Select a user to delete.")
            return
        if username.lower() == "admin":
            QMessageBox.warning(self, self.t("warning"), "Admin kullanıcısı silinemez." if self.language == "tr" else "The admin user cannot be deleted.")
            return
        if username.lower() == str(self.current_user).lower():
            QMessageBox.warning(self, self.t("warning"), "Aktif kullanıcıyı silemezsiniz." if self.language == "tr" else "You cannot delete the active user.")
            return
        users = [u for u in load_users() if str(u.get("username", "")).strip().lower() != username.lower()]
        save_users(users)
        self.user_records = load_users()
        self.refresh_users_table()
        self.clear_user_form()
        QMessageBox.information(self, "Bilgi" if self.language == "tr" else "Info", ("Kullanıcı silindi." if self.language == "tr" else "User deleted."))

    def get_muavin_field_definitions(self):
        return [
            {"key": "yilay", "tr": "Yıl/ay", "en": "Year/Month", "required": True, "aliases": ["yıl/ay", "yil/ay", "year/month", "period", "posting period"]},
            {"key": "ana_hesap", "tr": "Ana hesap", "en": "Main Account", "required": True, "aliases": ["ana hesap", "main account", "gl account", "account"]},
            {"key": "ana_hesap_adi", "tr": "DK hesabı uzun metni", "en": "G/L Account Long Text", "required": True, "aliases": ["dk hesabı uzun metni", "g/l account long text", "gl account long text", "account description", "hesap açıklaması"]},
            {"key": "referans", "tr": "Referans", "en": "Reference", "required": False, "aliases": ["referans", "reference", "assignment"]},
            {"key": "belge_numarasi", "tr": "Belge numarası", "en": "Document Number", "required": False, "aliases": ["belge numarası", "document number", "belge no", "document no"]},
            {"key": "belge_turu", "tr": "Belge türü", "en": "Document Type", "required": False, "aliases": ["belge türü", "document type"]},
            {"key": "karsi_hesap_tanimi", "tr": "Karşıt kayıt hesabı tanımı", "en": "Offset Account Description", "required": False, "aliases": ["karşıt kayıt hesabı tanımı", "karsit kayit hesabi tanimi", "offset account description", "contra account description"]},
            {"key": "islem_kodu", "tr": "İşlem kodu", "en": "Transaction Code", "required": False, "aliases": ["işlem kodu", "islem kodu", "transaction code", "tcode", "t-code"]},
            {"key": "up_tutar", "tr": "UP cinsinden tutar", "en": "Amount in Local Currency", "required": False, "aliases": ["up cinsinden tutar", "amount in local currency", "company code currency amount", "yerel para tutarı"]},
            {"key": "belge_pb_tutar", "tr": "Belge PB cinsinden tutar", "en": "Amount in Document Currency", "required": False, "aliases": ["belge pb cinsinden tutar", "amount in document currency", "document currency amount"]},
            {"key": "belge_pb", "tr": "Belge para birimi", "en": "Document Currency", "required": False, "aliases": ["belge para birimi", "document currency", "currency"]},
            {"key": "karsi_hesap", "tr": "Karşıt kayıt hesabı", "en": "Offset Account", "required": False, "aliases": ["karşıt kayıt hesabı", "offset account", "contra account", "karsit kayit hesabi"]},
            {"key": "denklestirme", "tr": "Denkleştirme belgesi", "en": "Clearing Document", "required": False, "aliases": ["denkleştirme belgesi", "clearing document", "clearing doc"]},
            {"key": "metin", "tr": "Metin", "en": "Text", "required": False, "aliases": ["metin", "text", "item text"]},
            {"key": "ters_kayit", "tr": "Ters kayıt blg.no.", "en": "Reversal Document No.", "required": False, "aliases": ["ters kayıt blg.no.", "reversal document no.", "reversal doc", "reverse document"]},
            {"key": "kullanici", "tr": "Kullanıcı adı", "en": "User Name", "required": False, "aliases": ["kullanıcı adı", "kullanici adi", "user name", "username", "created by"]},
            {"key": "vergi_gostergesi", "tr": "Vergi göstergesi", "en": "Tax Indicator", "required": False, "aliases": ["vergi göstergesi", "tax indicator", "tax code"]},
            {"key": "belge_tarihi", "tr": "Belge Tarihi", "en": "Document Date", "required": False, "aliases": ["belge tarihi", "document date", "doc date"]},
            {"key": "kayit_tarihi", "tr": "Kayıt Tarihi", "en": "Posting Date", "required": False, "aliases": ["kayıt tarihi", "kayit tarihi", "posting date", "entry date"]},
            {"key": "giris_tarihi", "tr": "Giriş Tarihi", "en": "Created Date", "required": False, "aliases": ["giriş tarihi", "giris tarihi", "created date", "creation date", "created on"]},
            {"key": "masraf_yeri", "tr": "Masraf yeri", "en": "Cost Center", "required": False, "aliases": ["masraf yeri", "cost center"]},
            {"key": "masraf_yeri_tanimi", "tr": "Masraf yeri tanımı", "en": "Cost Center Description", "required": False, "aliases": ["masraf yeri tanımı", "cost center description", "cost center text"]},
        ]

    def get_muavin_available_columns(self) -> List[str]:
        if self.muavin_available_columns:
            return list(self.muavin_available_columns)
        if self.muavin_raw_df is not None and not self.muavin_raw_df.empty:
            return [str(c).strip() for c in self.muavin_raw_df.columns]
        return []

    def load_muavin_raw_data_from_sources(self):
        if not self.muavin_file_paths:
            return
        mapping = self.get_muavin_selected_mapping()
        selected_cols = [str(v).strip() for v in mapping.values() if str(v).strip()]
        if not selected_cols:
            selected_cols = self.get_muavin_available_columns()
        selected_cols = list(dict.fromkeys(selected_cols))
        loaded_frames = []
        for idx, file_path in enumerate(self.muavin_file_paths, start=1):
            self.set_processing_message(
                (f"Muavin kaynak dosyaları yükleniyor... ({idx}/{len(self.muavin_file_paths)})" if self.language == "tr" else f"Loading subledger source files... ({idx}/{len(self.muavin_file_paths)})"),
                min(35, 8 + int((idx / max(len(self.muavin_file_paths), 1)) * 20))
            )
            df_part = read_excel_selected_columns(file_path, usecols=selected_cols if selected_cols else None)
            df_part["_source_file"] = os.path.basename(file_path)
            loaded_frames.append(df_part)
        self.muavin_raw_df = pd.concat(loaded_frames, ignore_index=True) if loaded_frames else pd.DataFrame()

    def guess_muavin_column_mapping(self):
        defs = self.get_muavin_field_definitions()
        raw_cols = self.get_muavin_available_columns()
        if not raw_cols:
            self.muavin_column_mapping = {}
            return {}
        norm_to_actual = {normalize_col_name(c): str(c) for c in raw_cols}
        mapping = {}
        used_actual = set()
        for fd in defs:
            selected = ""
            targets = [fd["tr"], fd["en"]] + fd.get("aliases", [])
            for cand in targets:
                norm = normalize_col_name(cand)
                if norm in norm_to_actual and norm_to_actual[norm] not in used_actual:
                    selected = norm_to_actual[norm]
                    break
            if not selected:
                for cand in targets:
                    norm = normalize_col_name(cand)
                    for raw_norm, actual in norm_to_actual.items():
                        if actual in used_actual:
                            continue
                        if norm and (norm in raw_norm or raw_norm in norm):
                            selected = actual
                            break
                    if selected:
                        break
            if selected:
                used_actual.add(selected)
            mapping[fd["key"]] = selected
        self.muavin_column_mapping = mapping
        return mapping

    def refresh_muavin_mapping_table(self):
        if not hasattr(self, "muavin_mapping_table"):
            return
        defs = self.get_muavin_field_definitions()
        raw_cols = self.get_muavin_available_columns()
        if not self.muavin_column_mapping:
            self.guess_muavin_column_mapping()
        self.muavin_mapping_table.clear()
        self.muavin_mapping_table.setRowCount(len(defs))
        self.muavin_mapping_combos = {}
        self.muavin_mapping_table.setColumnCount(4)
        self.muavin_mapping_table.setHorizontalHeaderLabels([
            "Field" if self.language == "en" else "Alan",
            "Required" if self.language == "en" else "Zorunlu",
            "Suggested Source" if self.language == "en" else "Önerilen Kaynak",
            "Selected Source Column" if self.language == "en" else "Seçilen Kaynak Kolon",
        ])
        for r, fd in enumerate(defs):
            label = fd["en"] if self.language == "en" else fd["tr"]
            set_table_item(self.muavin_mapping_table, r, 0, label)
            set_table_item(self.muavin_mapping_table, r, 1, ("Yes" if self.language == "en" else "Evet") if fd["required"] else ("No" if self.language == "en" else "Hayır"))
            suggested = self.muavin_column_mapping.get(fd["key"], "")
            set_table_item(self.muavin_mapping_table, r, 2, suggested or "-")
            combo = QComboBox()
            combo.addItem("")
            combo.addItems([str(c) for c in raw_cols])
            if suggested:
                idx = combo.findText(suggested)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
            self.muavin_mapping_table.setCellWidget(r, 3, combo)
            self.muavin_mapping_combos[fd["key"]] = combo
        header = self.muavin_mapping_table.horizontalHeader()
        for i in range(4):
            header.setSectionResizeMode(i, QHeaderView.Interactive)
        self.muavin_mapping_table.resizeColumnsToContents()

    def apply_muavin_mapping_from_table(self):
        defs = self.get_muavin_field_definitions()
        mapping = {}
        for fd in defs:
            combo = self.muavin_mapping_combos.get(fd["key"]) if hasattr(self, "muavin_mapping_combos") else None
            mapping[fd["key"]] = combo.currentText().strip() if combo is not None else self.muavin_column_mapping.get(fd["key"], "")
        self.muavin_column_mapping = mapping

    def get_muavin_selected_mapping(self):
        if hasattr(self, "muavin_mapping_combos") and self.muavin_mapping_combos:
            self.apply_muavin_mapping_from_table()
        if not self.muavin_column_mapping:
            self.guess_muavin_column_mapping()
        return dict(self.muavin_column_mapping)

    def handle_muavin_auto_map(self):
        self.guess_muavin_column_mapping()
        self.refresh_muavin_mapping_table()
        self.muavin_loaded_cols_label.setText(self.build_muavin_loaded_columns_status_text())

    def handle_muavin_apply_map(self):
        self.apply_muavin_mapping_from_table()
        self.muavin_loaded_cols_label.setText(self.build_muavin_loaded_columns_status_text())
        QMessageBox.information(self, "Bilgi" if self.language == "tr" else "Info", "Kolon eşleme uygulandı." if self.language == "tr" else "Column mapping applied.")

    def get_muavin_required_columns_map(self):
        defs = self.get_muavin_field_definitions()
        return {
            "required": [d["tr"] for d in defs if d.get("required")],
            "used": [d["tr"] for d in defs if not d.get("required")],
        }

    def build_muavin_required_columns_text(self):
        col_map = self.get_muavin_required_columns_map()
        req = ", ".join(col_map["required"])
        used = ", ".join(col_map["used"])
        if self.language == "en":
            return f"Required columns: {req} | Additional columns used in analysis: {used} | Custom mapping is supported below. Multiple files are supported and only mapped columns are read during analysis for speed."
        return f"Zorunlu kolonlar: {req} | Analizde kullanılan ek kolonlar: {used} | Aşağıda özel kolon eşleme yapılabilir. Birden fazla dosya desteklenir ve hız için analiz sırasında yalnızca eşlenen kolonlar okunur."

    def build_muavin_loaded_columns_status_text(self):
        raw_cols = self.get_muavin_available_columns()
        if not raw_cols:
            return ""
        mapping = self.get_muavin_selected_mapping()
        defs = self.get_muavin_field_definitions()
        required_defs = [d for d in defs if d.get("required")]
        found_required = [d["tr"] for d in required_defs if mapping.get(d["key"], "")]
        missing_required = [d["tr"] for d in required_defs if not mapping.get(d["key"], "")]
        matched_count = len([k for k, v in mapping.items() if v])
        preview = ", ".join([str(c) for c in raw_cols[:12]])
        if len(raw_cols) > 12:
            preview += f" (+{len(raw_cols) - 12})"
        file_count = len(self.muavin_file_paths) if self.muavin_file_paths else (1 if self.muavin_file_path else 0)
        if self.language == "en":
            return (
                f"Loaded file count: {file_count} | Loaded raw columns: {preview} | Found required: {len(found_required)}/{len(required_defs)} | "
                f"Mapped analysis fields: {matched_count}/{len(defs)} | Missing required: "
                f"{', '.join(missing_required) if missing_required else 'None'}"
            )
        return (
            f"Yüklenen dosya sayısı: {file_count} | Yüklenen ham kolonlar: {preview} | Bulunan zorunlu: {len(found_required)}/{len(required_defs)} | "
            f"Eşlenen analiz alanları: {matched_count}/{len(defs)} | Eksik zorunlu: "
            f"{', '.join(missing_required) if missing_required else 'Yok'}"
        )

    def refresh_muavin_ui(self):
        if hasattr(self, "muavin_title"):
            is_en = self.language == "en"
            self.muavin_title.setText("Subledger Analysis" if is_en else self.t("muavinAnalysis"))
            self.muavin_desc.setText("Load the raw subledger file and analyze periods, document types, users, offset accounts, tax indicators, and cost centers in one screen." if is_en else "Muavin defteri yükleyerek dönem, belge türü, kullanıcı, karşıt hesap, vergi göstergesi ve masraf yeri bazlı çok boyutlu analiz alın.")
            self.btn_load_muavin.setText("Load Subledger File(s)" if is_en else "Muavin Dosya(ları) Yükle")
            self.btn_start_muavin.setText("Run" if is_en else "Başlat")
            self.btn_export_muavin.setText("Export Subledger Excel" if is_en else "Muavin Excel Çıktısı")
            self.btn_muavin_autofit.setText("Auto Fit Columns" if is_en else "Kolonları Otomatik Sığdır")
            if hasattr(self, "btn_muavin_expand_all"):
                self.btn_muavin_expand_all.setVisible(False)
            if hasattr(self, "btn_muavin_collapse_all"):
                self.btn_muavin_collapse_all.setVisible(False)
            self.btn_muavin_risk_only.setText("Risky Records Only" if is_en else "Sadece Riskli Kayıtlar")
            self.btn_muavin_cost_alarm_only.setText("Cost Alarm Only" if is_en else "Sadece Masraf Alarmı")
            self.btn_muavin_risk_only.setObjectName("filterSoftActive" if getattr(self, "muavin_risk_only", False) else "filterSoft")
            self.btn_muavin_risk_only.style().unpolish(self.btn_muavin_risk_only)
            self.btn_muavin_risk_only.style().polish(self.btn_muavin_risk_only)
            self.btn_muavin_cost_alarm_only.setObjectName("filterSoftActive" if getattr(self, "muavin_cost_alarm_only", False) else "filterSoft")
            self.btn_muavin_cost_alarm_only.style().unpolish(self.btn_muavin_cost_alarm_only)
            self.btn_muavin_cost_alarm_only.style().polish(self.btn_muavin_cost_alarm_only)
            self.muavin_file_label.setText(self.muavin_file_path or "")
            self.muavin_file_label.setPlaceholderText("Loaded raw subledger file(s)" if is_en else "Yüklenen ham muavin dosya(ları)")
            self.muavin_required_cols_label.setText(self.build_muavin_required_columns_text())
            self.muavin_loaded_cols_label.setText(self.build_muavin_loaded_columns_status_text())
            if hasattr(self, "muavin_mapping_title"):
                self.muavin_mapping_title.setText("Subledger Column Mapping" if is_en else "Muavin Kolon Eşleme")
                self.muavin_mapping_desc.setText("If you load a different subledger format, map the source columns to analysis fields here." if is_en else "Farklı formatta gelen muavin raporlarında kaynak kolonları burada analiz alanlarına eşleyebilirsiniz.")
                self.btn_muavin_auto_map.setText("Auto Match" if is_en else "Otomatik Eşleştir")
                self.btn_muavin_apply_map.setText("Apply Mapping" if is_en else "Eşlemeyi Uygula")
                self.refresh_muavin_mapping_table()
            self.fill_muavin_combos()
            self.cmb_muavin_density.blockSignals(True)
            self.cmb_muavin_density.clear()
            self.cmb_muavin_density.addItems([self.t("compact"), self.t("normal"), self.t("wide")])
            self.cmb_muavin_density.setCurrentIndex({"compact": 0, "normal": 1, "wide": 2}.get(getattr(self, "muavin_row_density", "normal"), 1))
            self.cmb_muavin_density.blockSignals(False)
            if hasattr(self, "cmb_muavin_account"):
                self.cmb_muavin_account.setToolTip("Hesap filtresi")
                self.cmb_muavin_statement.setToolTip("Bilanço / Gelir Tablosu filtresi")
                self.cmb_muavin_period.setToolTip("Dönem filtresi")
                self.cmb_muavin_doc_type.setToolTip("Belge türü filtresi")
                self.cmb_muavin_user.setToolTip("Kullanıcı filtresi")
                self.cmb_muavin_contra.setToolTip("Satıcı / cari filtresi")
                self.cmb_muavin_cost.setToolTip("Masraf yeri filtresi")
            if hasattr(self, "cmb_muavin_search_mode"):
                self.lbl_muavin_search_mode.setText("Arama Tipi" if not is_en else "Search Type")
                self.cmb_muavin_search_mode.blockSignals(True)
                self.cmb_muavin_search_mode.clear()
                self.cmb_muavin_search_mode.addItems(["İçinde Bul" if not is_en else "Contains", "Tam Eşleşme Bul" if not is_en else "Exact Match"])
                self.cmb_muavin_search_mode.setCurrentIndex(1 if getattr(self, "muavin_search_mode", "contains") == "exact" else 0)
                self.cmb_muavin_search_mode.blockSignals(False)
            if hasattr(self, "btn_muavin_view_user"):
                self.btn_muavin_view_user.setText("Kullanıcı Bazlı Analiz" if not is_en else "User-Based Analysis")
                self.btn_muavin_view_tax.setText("Vergisel Analiz" if not is_en else "Tax Analysis")
                self.btn_muavin_view_account.setText("Hesap İçerik Analiz" if not is_en else "Account Content Analysis")
                self.set_muavin_result_view(getattr(self, "muavin_current_view", "user_based"))

    def set_muavin_result_view(self, view_key: str):
        if view_key not in ["user_based", "tax_based", "account_content"]:
            view_key = "user_based"
        self.muavin_current_view = view_key
        for key, btn in getattr(self, "muavin_result_buttons", {}).items():
            btn.setObjectName("toggleButtonActive" if key == view_key else "toggleButton")
            btn.style().unpolish(btn)
            btn.style().polish(btn)
        for key, wrapper in getattr(self, "muavin_view_wrappers", {}).items():
            wrapper.setVisible(key == view_key)
        if getattr(self, "muavin_analysis_ready", False):
            try:
                self.refresh_muavin_tables()
            except Exception:
                pass

    def get_muavin_group_permissions(self, perms: Dict[str, bool]) -> Dict[str, bool]:
        muavin_main = perms.get("muavin", True)
        return {
            "user_based": muavin_main and perms.get("muavin_user_based", True),
            "tax_based": muavin_main and perms.get("muavin_tax_based", True),
            "account_content": muavin_main and perms.get("muavin_account_content", True),
        }

    def toggle_muavin_section(self, key: str):
        state = not getattr(self, "muavin_section_states", {}).get(key, True)
        if not hasattr(self, "muavin_section_states") or not isinstance(self.muavin_section_states, dict):
            self.muavin_section_states = {}
        self.muavin_section_states[key] = state
        table = self.muavin_section_wrappers.get(key)
        if table is not None:
            table.setVisible(state)
        btn = self.muavin_section_buttons.get(key)
        if btn is not None:
            label = btn.text().split(" ", 1)[1] if " " in btn.text() else btn.text()
            btn.setText(f"{'▾' if state else '▸'} {label}")

    def set_all_muavin_sections(self, state: bool):
        if not hasattr(self, "muavin_section_states") or not isinstance(self.muavin_section_states, dict):
            self.muavin_section_states = {}
        for key in list(getattr(self, "muavin_tables", {}).keys()):
            self.muavin_section_states[key] = bool(state)
            table = self.muavin_section_wrappers.get(key)
            if table is not None:
                table.setVisible(bool(state))
            btn = self.muavin_section_buttons.get(key)
            if btn is not None:
                label = btn.text().split(" ", 1)[1] if " " in btn.text() else btn.text()
                btn.setText(f"{'▾' if state else '▸'} {label}")


    def on_analysis_search_mode_changed(self):
        self.analysis_search_mode = "exact" if self.cmb_analysis_search_mode.currentIndex() == 1 else "contains"
        self.refresh_analysis_tables()

    def on_regular_ft_search_mode_changed(self):
        self.regular_ft_search_mode = "exact" if self.cmb_regular_ft_search_mode.currentIndex() == 1 else "contains"
        self.refresh_regular_ft_table()

    def on_muavin_search_mode_changed(self):
        self.muavin_search_mode = "exact" if self.cmb_muavin_search_mode.currentIndex() == 1 else "contains"
        self.refresh_muavin_tables()

    def toggle_muavin_risk_only(self):
        self.muavin_risk_only = self.btn_muavin_risk_only.isChecked()
        self.refresh_muavin_ui()
        self.refresh_muavin_tables()

    def toggle_muavin_cost_alarm_only(self):
        self.muavin_cost_alarm_only = self.btn_muavin_cost_alarm_only.isChecked()
        self.refresh_muavin_ui()
        self.refresh_muavin_tables()

    def on_muavin_density_changed(self):
        idx = self.cmb_muavin_density.currentIndex()
        self.muavin_row_density = {0: "compact", 1: "normal", 2: "wide"}.get(idx, "normal")
        self.apply_muavin_row_heights()

    def apply_muavin_row_heights(self):
        height = 30 if self.muavin_row_density == "compact" else 48 if self.muavin_row_density == "wide" else 38
        for table in getattr(self, "muavin_tables", {}).values():
            for r in range(table.rowCount()):
                table.setRowHeight(r, height)

    def auto_fit_muavin_tables(self):
        for table in getattr(self, "muavin_tables", {}).values():
            if not table.isVisible() or table.columnCount() == 0:
                continue
            table.setUpdatesEnabled(False)
            header = table.horizontalHeader()
            for i in range(table.columnCount()):
                header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
                table.resizeColumnToContents(i)
                width = table.columnWidth(i)
                table.setColumnWidth(i, min(max(width + 18, 90), 360))
            table.setUpdatesEnabled(True)
            self.enable_manual_column_resize(table)

    def on_muavin_contra_double_click(self, row: int, col: int):
        table = self.muavin_tables.get("contra")
        if table is None or row < 0:
            return
        item = table.item(row, 1)
        if item is None:
            return
        self.muavin_selected_contra_name = item.text().strip()
        self.muavin_selected_document = ""
        self.refresh_muavin_tables()

    def on_muavin_drilldown_double_click(self, row: int, col: int):
        table = self.muavin_tables.get("drilldown")
        if table is None or row < 0:
            return
        item = table.item(row, 3)
        if item is None:
            return
        self.muavin_selected_document = item.text().strip()
        self.refresh_muavin_tables()

    def on_muavin_duplicate_ref_double_click(self, row: int, col: int):
        table = self.muavin_tables.get("dupref")
        if table is None or row < 0:
            return
        ref_item = table.item(row, 0)
        contra_item = table.item(row, 1)
        contra_name_item = table.item(row, 2)
        if ref_item is None:
            return
        ref_val = normalize_text_value(ref_item.text())
        contra_val = normalize_text_value(contra_item.text()) if contra_item else ""
        contra_name_val = normalize_text_value(contra_name_item.text()) if contra_name_item else ""
        self.muavin_selected_reference = f"{ref_val}||{contra_val or contra_name_val}" if ref_val else ""
        self.refresh_muavin_tables()

    def fill_muavin_combos(self):
        if not hasattr(self, "cmb_muavin_account"):
            return

        combo_defs = [
            (self.cmb_muavin_account, ["Tümü"] + self.muavin_accounts, "muavin_selected_account"),
            (self.cmb_muavin_statement, ["Tümü", "Bilanço", "Gelir Tablosu", "Maliyet / Nazım", "Nazım", "Diğer"], "muavin_statement_filter"),
            (self.cmb_muavin_period, ["Tümü"] + self.muavin_periods, "muavin_period_filter"),
            (self.cmb_muavin_doc_type, ["Tümü"] + self.muavin_doc_types, "muavin_doc_type_filter"),
            (self.cmb_muavin_user, ["Tümü"] + self.muavin_users, "muavin_user_filter"),
            (self.cmb_muavin_contra, ["Tümü"] + self.muavin_contras, "muavin_contra_filter"),
            (self.cmb_muavin_cost, ["Tümü"] + self.muavin_cost_centers, "muavin_cost_filter"),
        ]
        for combo, options, attr_name in combo_defs:
            combo.blockSignals(True)
            combo.clear()
            combo.addItems(options)
            current_val = getattr(self, attr_name, "Tümü")
            idx = combo.findText(current_val)
            if idx < 0:
                current_val = "Tümü"
                setattr(self, attr_name, current_val)
                idx = 0
            combo.setCurrentIndex(idx)
            combo.blockSignals(False)

        if hasattr(self, "input_muavin_search"):
            self.input_muavin_search.blockSignals(True)
            self.input_muavin_search.setText(self.muavin_search_text)
            self.input_muavin_search.blockSignals(False)
    def load_muavin_file(self):
        dialog_title = "Muavin Excel Dosyaları Seç" if self.language == "tr" else "Select Subledger Excel Files"
        paths, _ = QFileDialog.getOpenFileNames(self, dialog_title, BASE_DIR, "Excel Files (*.xlsx *.xls)")
        if not paths:
            return
        self.begin_busy_state("Muavin dosya başlıkları okunuyor..." if self.language == "tr" else "Reading subledger file headers...", 5)
        started = self.start_background_worker(
            load_muavin_headers_payload,
            self.on_muavin_headers_loaded,
            self.on_muavin_headers_failed,
            list(paths),
            self.language,
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def on_muavin_headers_loaded(self, payload):
        self.clear_active_worker()
        self.muavin_file_paths = list(payload.get("file_paths", []))
        self.muavin_available_columns = list(payload.get("available_columns", []))
        self.muavin_raw_df = None
        self.muavin_clean_df = None
        self.muavin_file_path = payload.get("display_path", "")
        self.muavin_analysis_ready = False
        self.muavin_current_view = "user_based"
        self.guess_muavin_column_mapping()
        self.refresh_muavin_ui()
        self.tabs.setCurrentIndex(4)
        summary = self.build_muavin_loaded_columns_status_text()
        self.end_busy_state("Muavin dosyaları hazır." if self.language == "tr" else "Subledger files are ready.", 100)
        if self.language == "en":
            QMessageBox.information(self, "Info", f"{len(self.muavin_file_paths)} subledger file(s) loaded. Run Analysis will read only the mapped columns for speed.\n\n{summary}")
        else:
            QMessageBox.information(self, "Bilgi", f"{len(self.muavin_file_paths)} adet muavin dosyası yüklendi. Hız için Başlat aşamasında sadece eşlenen kolonlar okunacaktır.\n\n{summary}")

    def on_muavin_headers_failed(self, error_text: str):
        self.clear_active_worker()
        self.end_busy_state("Muavin dosyaları okunamadı." if self.language == "tr" else "Subledger files could not be read.", 0)
        if self.language == "en":
            QMessageBox.critical(self, "Error", f"Subledger file(s) could not be read.\n\n{error_text}")
        else:
            QMessageBox.critical(self, "Hata", f"Muavin dosyaları okunamadı.\n\n{error_text}")

    def start_muavin_analysis(self):
        if (self.muavin_raw_df is None or self.muavin_raw_df.empty) and not self.muavin_file_paths:
            QMessageBox.warning(self, "Uyarı" if self.language == "tr" else "Warning", "Önce muavin dosyası yükleyin." if self.language == "tr" else "Load subledger file(s) first.")
            return
        self.begin_busy_state("Muavin analizi hazırlanıyor..." if self.language == "tr" else "Preparing subledger analysis...", 5)
        mapping = self.get_muavin_selected_mapping()
        if self.muavin_file_paths:
            started = self.start_background_worker(
                build_muavin_analysis_payload_from_files,
                self.on_muavin_analysis_loaded,
                self.on_muavin_analysis_failed,
                list(self.muavin_file_paths),
                mapping,
                self.language,
            )
        else:
            started = self.start_background_worker(
                build_muavin_analysis_payload,
                self.on_muavin_analysis_loaded,
                self.on_muavin_analysis_failed,
                self.muavin_raw_df,
                mapping,
                self.language,
            )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def on_muavin_analysis_loaded(self, payload):
        self.clear_active_worker()
        self.muavin_raw_df = payload.get("raw_df", pd.DataFrame())
        self.muavin_clean_df = payload.get("clean_df", pd.DataFrame())
        self.muavin_accounts = payload.get("accounts", [])
        self.muavin_periods = payload.get("periods", [])
        self.muavin_doc_types = payload.get("doc_types", [])
        self.muavin_users = payload.get("users", [])
        self.muavin_contras = payload.get("contras", [])
        self.muavin_cost_centers = payload.get("cost_centers", [])
        if self.muavin_selected_account not in (["Tümü"] + self.muavin_accounts):
            self.muavin_selected_account = "Tümü"
        self.set_processing_message("Muavin tabloları oluşturuluyor..." if self.language == "tr" else "Building subledger tables...", 85)
        self.muavin_analysis_ready = True
        self.refresh_muavin_ui()
        self.refresh_muavin_tables()
        self.tabs.setCurrentIndex(4)
        self.end_busy_state("Muavin analizi hazırlandı." if self.language == "tr" else "Subledger analysis is ready.", 100)
        QMessageBox.information(self, "Bilgi" if self.language == "tr" else "Info", "Muavin analizi hazırlandı." if self.language == "tr" else "Subledger analysis is ready.")

    def on_muavin_analysis_failed(self, error_text: str):
        self.clear_active_worker()
        self.end_busy_state("Muavin analizi oluşturulamadı." if self.language == "tr" else "Subledger analysis could not be created.", 0)
        QMessageBox.critical(self, "Hata" if self.language == "tr" else "Error", f"{'Muavin analizi oluşturulamadı.' if self.language == 'tr' else 'Subledger analysis could not be created.'}\n\n{error_text}")
    def prepare_muavin_dataframe(self):
        if (self.muavin_raw_df is None or self.muavin_raw_df.empty) and self.muavin_file_paths:
            self.load_muavin_raw_data_from_sources()

        if self.muavin_raw_df is None or self.muavin_raw_df.empty:
            self.muavin_clean_df = pd.DataFrame()
            self.muavin_accounts = []
            self.muavin_periods = []
            self.muavin_doc_types = []
            self.muavin_users = []
            self.muavin_contras = []
            self.muavin_cost_centers = []
            return

        self.set_processing_message("Muavin verisi kopyalanıyor ve kolonlar hazırlanıyor...", 12)
        df = self.muavin_raw_df.copy()
        mapping = self.get_muavin_selected_mapping()
        rename_map = {}
        wanted = {k: v for k, v in mapping.items() if v}
        defs_by_key = {d["key"]: d for d in self.get_muavin_field_definitions()}
        missing = []
        for req_key in ["yilay", "ana_hesap", "ana_hesap_adi"]:
            if not wanted.get(req_key):
                missing.append(defs_by_key[req_key]["tr"])
        if missing:
            raise ValueError("Muavin dosyasında bulunamayan kolonlar: " + ", ".join(missing))
        for new, old in wanted.items():
            if old and str(old).strip() in df.columns:
                rename_map[str(old).strip()] = new
        df = df.rename(columns=rename_map)

        self.set_processing_message("Muavin kolonları normalize ediliyor...", 22)
        self.set_processing_message("Muavin kolonları normalize ediliyor...", 22)
        for col in ["ana_hesap","ana_hesap_adi","referans","belge_numarasi","belge_turu","karsi_hesap_tanimi","islem_kodu","belge_pb","karsi_hesap","denklestirme","metin","ters_kayit","kullanici","vergi_gostergesi","masraf_yeri","masraf_yeri_tanimi","belge_tarihi","kayit_tarihi","giris_tarihi"]:
            if col not in df.columns:
                df[col] = ""
            df[col] = df[col].map(normalize_text_value)
        for col in ["up_tutar","belge_pb_tutar"]:
            if col not in df.columns:
                df[col] = 0.0
            df[col] = df[col].map(safe_float)

        self.set_processing_message("Temel muavin göstergeleri hazırlanıyor...", 32)
        df = ensure_muavin_derived_columns(df)
        df["donem"] = df["yilay"].map(parse_muavin_period)
        df["hesap_prefix"] = df["ana_hesap"].map(normalize_hesap_prefix)
        df["finansal_tablo"] = df["hesap_prefix"].map(classify_financial_statement)
        df["has_denklestirme"] = df["denklestirme"].astype(str).str.strip() != ""
        df["has_ters_kayit"] = df["ters_kayit"].astype(str).str.strip() != ""
        df["belge_anahtar"] = df.apply(lambda x: normalize_text_value(x["belge_numarasi"]) or normalize_text_value(x["referans"]), axis=1)
        df = df[(df["ana_hesap"].astype(str).str.strip() != "") | (df["ana_hesap_adi"].astype(str).str.strip() != "")]
        df = df[df["donem"].astype(str).str.strip() != ""].copy()
        df = df[~((df["belge_anahtar"] == "") & (df["belge_turu"] == "") & (df["karsi_hesap"] == "") & (df["karsi_hesap_tanimi"] == "") & (df["metin"] == "") & (df["up_tutar"].abs() < 1e-12))].copy()

        df["duplicate_ref_contra_key"] = df.apply(
            lambda x: (
                f"{normalize_text_value(x['referans'])}||{normalize_text_value(x['karsi_hesap']) or normalize_text_value(x['karsi_hesap_tanimi'])}"
                if normalize_text_value(x["referans"])
                else ""
            ),
            axis=1
        )

        open_ref_df = df[
            (df["referans"].astype(str).str.strip() != "")
            & (~df["has_ters_kayit"])
            & (
                (df["karsi_hesap"].astype(str).str.strip() != "")
                | (df["karsi_hesap_tanimi"].astype(str).str.strip() != "")
            )
        ].copy()
        if not open_ref_df.empty:
            ref_contra_doc_counts = open_ref_df.groupby("duplicate_ref_contra_key")["belge_anahtar"].nunique().to_dict()
            df["duplicate_open_reference"] = df["duplicate_ref_contra_key"].map(
                lambda x: ref_contra_doc_counts.get(str(x).strip(), 0) > 1 if str(x).strip() else False
            )
            df["duplicate_open_reference_count"] = df["duplicate_ref_contra_key"].map(
                lambda x: int(ref_contra_doc_counts.get(str(x).strip(), 0)) if str(x).strip() else 0
            )
        else:
            df["duplicate_open_reference"] = False
            df["duplicate_open_reference_count"] = 0

        self.set_processing_message("Kullanıcı risk skorları hesaplanıyor...", 45)
        grp_ud = df.groupby(["belge_turu", "kullanici"]).agg(
            satir=("ana_hesap", "size"),
            belge_sayisi=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
            tutar=("up_tutar", lambda s: float(s.abs().sum())),
            ters_orani=("has_ters_kayit", "mean"),
            denkl_orani=("has_denklestirme", "mean"),
            cari_sayisi=("karsi_hesap", lambda s: s.replace("", pd.NA).dropna().nunique()),
        ).reset_index()
        if not grp_ud.empty:
            grp_ud["risk_skoru"] = grp_ud["satir"] * 0.8 + grp_ud["belge_sayisi"] * 1.2 + grp_ud["tutar"].rank(pct=True) * 35 + grp_ud["ters_orani"] * 30 + (1 - grp_ud["denkl_orani"]) * 10 + grp_ud["cari_sayisi"].rank(pct=True) * 15
            risk_map = {(r["belge_turu"], r["kullanici"]): float(r["risk_skoru"]) for _, r in grp_ud.iterrows()}
            df["risk_skoru"] = df.apply(lambda x: risk_map.get((x["belge_turu"], x["kullanici"]), 0.0), axis=1)
        else:
            df["risk_skoru"] = 0.0

        self.set_processing_message("Satıcı / cari yoğunluk ve anormallik skorları hesaplanıyor...", 58)
        grp_contra = df.groupby(["karsi_hesap", "karsi_hesap_tanimi"]).agg(
            satir=("ana_hesap", "size"),
            belge_sayisi=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
            tutar=("up_tutar", lambda s: float(s.abs().sum())),
            ters_orani=("has_ters_kayit", "mean"),
            vergi_sayisi=("vergi_gostergesi", lambda s: s.replace("", pd.NA).dropna().nunique()),
            belge_turu_sayisi=("belge_turu", lambda s: s.replace("", pd.NA).dropna().nunique()),
        ).reset_index()
        if not grp_contra.empty:
            mt, st = grp_contra["tutar"].mean(), grp_contra["tutar"].std(ddof=0)
            ms, ss = grp_contra["satir"].mean(), grp_contra["satir"].std(ddof=0)
            grp_contra["yogunluk_skoru"] = grp_contra["tutar"].rank(pct=True) * 50 + grp_contra["satir"].rank(pct=True) * 30 + grp_contra["belge_sayisi"].rank(pct=True) * 20
            grp_contra["anomali_skoru"] = grp_contra.apply(lambda r: max(0.0, safe_zscore(r["tutar"], mt, st)) * 30 + max(0.0, safe_zscore(r["satir"], ms, ss)) * 20 + r["ters_orani"] * 30 + (1 if r["vergi_sayisi"] >= 3 else 0) * 10 + (1 if r["belge_turu_sayisi"] >= 4 else 0) * 10, axis=1)
            contra_map = {(r["karsi_hesap"], r["karsi_hesap_tanimi"]): (float(r["yogunluk_skoru"]), float(r["anomali_skoru"])) for _, r in grp_contra.iterrows()}
            df["yogunluk_skoru"] = df.apply(lambda x: contra_map.get((x["karsi_hesap"], x["karsi_hesap_tanimi"]), (0.0, 0.0))[0], axis=1)
            df["anomali_skoru"] = df.apply(lambda x: contra_map.get((x["karsi_hesap"], x["karsi_hesap_tanimi"]), (0.0, 0.0))[1], axis=1)
        else:
            df["yogunluk_skoru"] = 0.0
            df["anomali_skoru"] = 0.0

        self.set_processing_message("Masraf yeri trendleri hazırlanıyor...", 68)
        cc = df.groupby(["masraf_yeri", "masraf_yeri_tanimi", "donem"])["up_tutar"].sum().reset_index()
        if not cc.empty:
            cc = cc.sort_values(["masraf_yeri", "donem"])
            cc["prev_tutar"] = cc.groupby(["masraf_yeri", "masraf_yeri_tanimi"])["up_tutar"].shift(1).fillna(0.0)
            cc["degisim_yuzde"] = cc.apply(lambda r: 0.0 if abs(float(r["prev_tutar"])) < 1e-12 and abs(float(r["up_tutar"])) < 1e-12 else (999.0 if abs(float(r["prev_tutar"])) < 1e-12 else ((float(r["up_tutar"]) - float(r["prev_tutar"])) / abs(float(r["prev_tutar"])) * 100.0)), axis=1)
            cc["cost_change_alarm"] = cc.apply(lambda r: trend_alarm_text(float(r["up_tutar"]), float(r["prev_tutar"])), axis=1)
            cc_map = {(r["masraf_yeri"], r["masraf_yeri_tanimi"], r["donem"]): (float(r["prev_tutar"]), float(r["degisim_yuzde"]), r["cost_change_alarm"]) for _, r in cc.iterrows()}
            df["prev_cost_tutar"] = df.apply(lambda x: cc_map.get((x["masraf_yeri"], x["masraf_yeri_tanimi"], x["donem"]), (0.0, 0.0, "Normal"))[0], axis=1)
            df["cost_degisim_yuzde"] = df.apply(lambda x: cc_map.get((x["masraf_yeri"], x["masraf_yeri_tanimi"], x["donem"]), (0.0, 0.0, "Normal"))[1], axis=1)
            df["cost_change_alarm"] = df.apply(lambda x: cc_map.get((x["masraf_yeri"], x["masraf_yeri_tanimi"], x["donem"]), (0.0, 0.0, "Normal"))[2], axis=1)
        else:
            df["prev_cost_tutar"] = 0.0
            df["cost_degisim_yuzde"] = 0.0
            df["cost_change_alarm"] = "Normal"

        self.set_processing_message("Belge ilişkisel kontrolleri hazırlanıyor...", 78)
        doc_grp = df.groupby("belge_anahtar").agg(
            doc_satir=("ana_hesap", "size"),
            doc_net_tutar=("up_tutar", "sum"),
            doc_abs_tutar=("up_tutar", lambda s: float(s.abs().sum())),
            doc_unique_accounts=("ana_hesap", lambda s: s.replace("", pd.NA).dropna().nunique()),
            doc_has_denkl=("has_denklestirme", "max"),
            doc_has_ters=("has_ters_kayit", "max"),
            doc_duplicate_ref=("duplicate_open_reference", "max"),
            doc_duplicate_ref_count=("duplicate_open_reference_count", "max"),
        ).reset_index()
        def _doc_status(r):
            anahtar = normalize_text_value(r["belge_anahtar"])
            if not anahtar:
                return "Belge anahtarı yok"
            if r.get("doc_duplicate_ref"):
                return f"Ters kayıt boşken referans + karşıt hesap mükerrer ({int(r.get('doc_duplicate_ref_count', 0))} belge)"
            if r["doc_has_denkl"] and r["doc_has_ters"]:
                return "Denkleştirme + ters kayıt birlikte"
            if abs(float(r["doc_net_tutar"])) <= 1.0 and int(r["doc_unique_accounts"]) >= 2:
                return "Belge netleşmiş"
            if r["doc_has_denkl"] and abs(float(r["doc_net_tutar"])) > 1.0:
                return "Denkleşmiş fakat net bakiye var"
            if r["doc_has_ters"] and abs(float(r["doc_net_tutar"])) > 1.0:
                return "Ters kayıt var, bakiye kontrolü gerekli"
            if int(r["doc_satir"]) >= 2 and not r["doc_has_denkl"]:
                return "Açık belge / ilişki kontrol"
            return "Normal"
        if not doc_grp.empty:
            doc_grp["doc_relation_status"] = doc_grp.apply(_doc_status, axis=1)
            doc_map = {r["belge_anahtar"]: (int(r["doc_satir"]), float(r["doc_net_tutar"]), float(r["doc_abs_tutar"]), int(r["doc_unique_accounts"]), bool(r["doc_has_denkl"]), bool(r["doc_has_ters"]), bool(r["doc_duplicate_ref"]), int(r["doc_duplicate_ref_count"]), r["doc_relation_status"]) for _, r in doc_grp.iterrows()}
            df["doc_satir"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[0])
            df["doc_net_tutar"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[1])
            df["doc_abs_tutar"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[2])
            df["doc_unique_accounts"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[3])
            df["doc_duplicate_ref"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[6])
            df["doc_duplicate_ref_count"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[7])
            df["doc_relation_status"] = df["belge_anahtar"].map(lambda x: doc_map.get(x, (0,0.0,0.0,0,False,False,False,0,"Normal"))[8])
        else:
            df["doc_satir"] = 0
            df["doc_net_tutar"] = 0.0
            df["doc_abs_tutar"] = 0.0
            df["doc_unique_accounts"] = 0
            df["doc_duplicate_ref"] = False
            df["doc_duplicate_ref_count"] = 0
            df["doc_relation_status"] = "Normal"

        df = build_muavin_audit_columns(df)
        df["risk_flag"] = (
            (df["risk_skoru"] >= 60)
            | (df["anomali_skoru"] >= 60)
            | (df["cost_change_alarm"].astype(str).str.contains("Alarm|Yeni Hareket", na=False))
            | (df["doc_relation_status"].astype(str) != "Normal")
            | (df["late_7day_flag"] == True)
            | (df["audit_risk_score"] >= 40)
        )

        self.set_processing_message("Muavin filtre listeleri hazırlanıyor...", 92)
        self.muavin_clean_df = df
        self.muavin_accounts = sorted(df["ana_hesap"].dropna().astype(str).unique().tolist())
        self.muavin_periods = sorted(df["donem"].dropna().astype(str).unique().tolist())
        self.muavin_doc_types = sorted([x for x in df["belge_turu"].dropna().astype(str).unique().tolist() if x])
        self.muavin_users = sorted([x for x in df["kullanici"].dropna().astype(str).unique().tolist() if x])
        self.muavin_contras = sorted([x for x in df["karsi_hesap_tanimi"].dropna().astype(str).unique().tolist() if x])
        self.muavin_cost_centers = sorted([f"{r['masraf_yeri']} - {r['masraf_yeri_tanimi']}".strip(" -") for _, r in df[["masraf_yeri","masraf_yeri_tanimi"]].drop_duplicates().iterrows() if normalize_text_value(r["masraf_yeri"]) or normalize_text_value(r["masraf_yeri_tanimi"])])
        if self.muavin_selected_account not in (["Tümü"] + self.muavin_accounts):
            self.muavin_selected_account = "Tümü"
    def get_filtered_muavin_df(self):
        if self.muavin_clean_df is None or self.muavin_clean_df.empty:
            return pd.DataFrame()
        df = self.muavin_clean_df.copy()
        if hasattr(self, "cmb_muavin_account") and self.cmb_muavin_account.count():
            self.muavin_selected_account = self.cmb_muavin_account.currentText() or "Tümü"
        if hasattr(self, "cmb_muavin_statement") and self.cmb_muavin_statement.count():
            self.muavin_statement_filter = self.cmb_muavin_statement.currentText() or "Tümü"
        if hasattr(self, "cmb_muavin_period") and self.cmb_muavin_period.count():
            self.muavin_period_filter = self.cmb_muavin_period.currentText() or "Tümü"
        if hasattr(self, "cmb_muavin_doc_type") and self.cmb_muavin_doc_type.count():
            self.muavin_doc_type_filter = self.cmb_muavin_doc_type.currentText() or "Tümü"
        if hasattr(self, "cmb_muavin_user") and self.cmb_muavin_user.count():
            self.muavin_user_filter = self.cmb_muavin_user.currentText() or "Tümü"
        if hasattr(self, "cmb_muavin_contra") and self.cmb_muavin_contra.count():
            self.muavin_contra_filter = self.cmb_muavin_contra.currentText() or "Tümü"
        if hasattr(self, "cmb_muavin_cost") and self.cmb_muavin_cost.count():
            self.muavin_cost_filter = self.cmb_muavin_cost.currentText() or "Tümü"
        if hasattr(self, "input_muavin_search"):
            self.muavin_search_text = self.input_muavin_search.text().strip().lower()

        if self.muavin_selected_account != "Tümü":
            df = df[df["ana_hesap"] == self.muavin_selected_account]
        if self.muavin_statement_filter != "Tümü":
            df = df[df["finansal_tablo"] == self.muavin_statement_filter]
        if self.muavin_period_filter != "Tümü":
            df = df[df["donem"] == self.muavin_period_filter]
        if self.muavin_doc_type_filter != "Tümü":
            df = df[df["belge_turu"] == self.muavin_doc_type_filter]
        if self.muavin_user_filter != "Tümü":
            df = df[df["kullanici"] == self.muavin_user_filter]
        if self.muavin_contra_filter != "Tümü":
            df = df[df["karsi_hesap_tanimi"] == self.muavin_contra_filter]
        if self.muavin_cost_filter != "Tümü":
            if " - " in self.muavin_cost_filter:
                cc, cc_name = self.muavin_cost_filter.split(" - ", 1)
                df = df[(df["masraf_yeri"] == cc) & (df["masraf_yeri_tanimi"] == cc_name)]
            else:
                df = df[(df["masraf_yeri"] == self.muavin_cost_filter) | (df["masraf_yeri_tanimi"] == self.muavin_cost_filter)]

        if self.muavin_search_text:
            q = self.muavin_search_text
            mask = df_text_search_mask(
                df,
                ["ana_hesap","ana_hesap_adi","referans","belge_numarasi","belge_turu","karsi_hesap","karsi_hesap_tanimi","metin","kullanici","vergi_gostergesi","masraf_yeri","masraf_yeri_tanimi","doc_relation_status"],
                q,
                getattr(self, "muavin_search_mode", "contains"),
            )
            df = df[mask]

        if self.muavin_risk_only and "risk_flag" in df.columns:
            df = df[df["risk_flag"] == True]

        return df.copy()
    def populate_simple_table(self, table: QTableWidget, headers: List[str], rows: List[List[str]], badge_matrix: Optional[List[List[Optional[str]]]] = None):
        table.clear()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                sval = str(val)
                align = Qt.AlignRight | Qt.AlignVCenter if c > 0 and parse_sort_value(sval) is not None else Qt.AlignLeft | Qt.AlignVCenter
                badge = None
                if badge_matrix and r < len(badge_matrix) and c < len(badge_matrix[r]):
                    badge = badge_matrix[r][c]
                if badge is None:
                    badge = risk_badge_from_text(sval)
                if badge is None and headers[c] in ["Risk Skoru", "Risk", "Yoğunluk", "Anormallik"] and parse_sort_value(sval) is not None:
                    badge = numeric_risk_badge(parse_sort_value(sval))
                set_table_item(table, r, c, sval, align, badge)
        header = table.horizontalHeader()
        for i in range(len(headers)):
            header.setSectionResizeMode(i, QHeaderView.Interactive)
            table.resizeColumnToContents(i)
            width = table.columnWidth(i)
            table.setColumnWidth(i, min(max(width + 20, 90), 360))
        self.enable_manual_column_resize(table)
        self.apply_muavin_row_heights()

    def populate_muavin_cost_table(self, headers: List[str], rows: List[List[str]], badge_matrix: List[List[Optional[str]]]):
        self.populate_simple_table(self.muavin_tables["cost"], headers, rows, badge_matrix=badge_matrix)

    def refresh_muavin_tables(self):
        if not hasattr(self, "muavin_tables"):
            return
        if not self.muavin_analysis_ready and (self.muavin_raw_df is None or self.muavin_raw_df is None):
            return

        df = self.get_filtered_muavin_df()
        df = ensure_muavin_derived_columns(df)
        total_rows = len(df)
        info_txt = f"{total_rows} satır / filtreli görünüm" if self.language == "tr" else f"{total_rows} rows / filtered view"
        if self.muavin_risk_only:
            info_txt += " • sadece riskli" if self.language == "tr" else " • risky only"
        if self.muavin_selected_contra_name:
            info_txt += f" • satıcı/cari: {self.muavin_selected_contra_name}" if self.language == "tr" else f" • vendor/account: {self.muavin_selected_contra_name}"
        if getattr(self, "muavin_selected_document", ""):
            info_txt += f" • belge: {self.muavin_selected_document}" if self.language == "tr" else f" • document: {self.muavin_selected_document}"
        if getattr(self, "muavin_selected_reference", ""):
            info_txt += f" • referans: {self.muavin_selected_reference}" if self.language == "tr" else f" • reference: {self.muavin_selected_reference}"
        self.muavin_info_pill.setText(info_txt)

        if df.empty:
            empty_text = "Muavin verisi yok" if self.language == "tr" else "No subledger data"
            for table in self.muavin_tables.values():
                self.populate_simple_table(table, ["Bilgi" if self.language == "tr" else "Info"], [[empty_text]])
            if hasattr(self, "muavin_result_info"):
                self.muavin_result_info.setText(empty_text)
            for card, title in zip(self.muavin_metric_cards, [
                ("Seçili Hesap" if self.language == "tr" else "Selected Account"),
                ("Finansal Tablo" if self.language == "tr" else "Financial Statement"),
                ("Belge Sayısı" if self.language == "tr" else "Document Count"),
                ("Sonuç Görünümü" if self.language == "tr" else "Result View"),
            ]):
                card["title"].setText(title)
                card["value"].setText("-")
                card["sub"].setText("-")
            return

        statement_label = safe_unique_join(df["finansal_tablo"].unique(), limit=3)
        belge_sayisi = int(df["belge_anahtar"].replace("", pd.NA).dropna().nunique())
        ters_belge = int(df.loc[df["has_ters_kayit"] == True, "belge_anahtar"].replace("", pd.NA).dropna().nunique()) if "has_ters_kayit" in df.columns else 0
        secili_hesap = self.muavin_selected_account if self.muavin_selected_account != "Tümü" else ("Tümü" if self.language == "tr" else "All")
        hesap_adlari = safe_unique_join(df["ana_hesap_adi"].unique(), limit=2)
        view_labels = {
            "user_based": ("Kullanıcı Bazlı Analiz" if self.language == "tr" else "User-Based Analysis"),
            "tax_based": ("Vergisel Analiz" if self.language == "tr" else "Tax Analysis"),
            "account_content": ("Hesap İçerik Analiz" if self.language == "tr" else "Account Content Analysis"),
        }
        card_values = [
            (("Seçili Hesap" if self.language == "tr" else "Selected Account"), secili_hesap, hesap_adlari or "-"),
            (("Finansal Tablo" if self.language == "tr" else "Financial Statement"), statement_label or "-", safe_unique_join(df["ana_hesap"].unique(), limit=4) or "-"),
            (("Belge Sayısı" if self.language == "tr" else "Document Count"), f"{belge_sayisi}", (f"Ters kayıt belge: {ters_belge}" if self.language == "tr" else f"Reversal documents: {ters_belge}")),
            (("Sonuç Görünümü" if self.language == "tr" else "Result View"), view_labels.get(getattr(self, "muavin_current_view", "user_based"), "-"), info_txt),
        ]
        for card, vals in zip(self.muavin_metric_cards, card_values):
            card["title"].setText(vals[0]); card["value"].setText(vals[1]); card["sub"].setText(vals[2])

        user_headers, user_rows, user_badges = self.build_muavin_user_based_result(df)
        tax_results = self.build_muavin_tax_based_result(df)
        account_results = self.build_muavin_account_content_result(df)

        self.populate_simple_table(self.muavin_tables["user_based"], user_headers, user_rows, badge_matrix=user_badges)
        self.populate_simple_table(self.muavin_tables["tax_dup"], tax_results["dup"][0], tax_results["dup"][1], badge_matrix=tax_results["dup"][2])
        self.populate_simple_table(self.muavin_tables["tax_projection"], tax_results["projection"][0], tax_results["projection"][1], badge_matrix=tax_results["projection"][2])
        self.populate_simple_table(self.muavin_tables["tax_indicator"], tax_results["indicator"][0], tax_results["indicator"][1], badge_matrix=tax_results["indicator"][2])
        self.populate_simple_table(self.muavin_tables["account_text"], account_results["text"][0], account_results["text"][1], badge_matrix=account_results["text"][2])
        self.populate_simple_table(self.muavin_tables["account_relation"], account_results["relation"][0], account_results["relation"][1], badge_matrix=account_results["relation"][2])
        self.populate_simple_table(self.muavin_tables["account_cost"], account_results["cost"][0], account_results["cost"][1], badge_matrix=account_results["cost"][2])

        if hasattr(self, "muavin_result_info"):
            self.muavin_result_info.setText(
                (
                    "Vergisel Analiz ve Hesap İçerik Analiz alanları artık 3 alt başlık halinde ayrı ayrı gösterilir."
                    if self.language == "tr"
                    else "Tax Analysis and Account Content Analysis are now displayed as three separate sub-sections."
                )
            )
        self.set_muavin_result_view(getattr(self, "muavin_current_view", "user_based"))

    def build_muavin_user_based_result(self, df: pd.DataFrame):
        base = df.copy()
        grp = base.groupby(["kullanici", "donem"]).agg(
            belge_turleri=("belge_turu", lambda s: safe_unique_join(s.unique(), limit=12)),
            vergi_gostergeleri=("vergi_gostergesi", lambda s: safe_unique_join(s.unique(), limit=12)),
            satici_seti=("karsi_hesap", lambda s: safe_unique_join(s.unique(), limit=12)),
            satici_unvan_seti=("karsi_hesap_tanimi", lambda s: safe_unique_join(s.unique(), limit=12)),
            borc_tutar=("up_tutar", lambda s: float(pd.Series(s)[pd.Series(s) > 0].sum())),
            alacak_tutar=("up_tutar", lambda s: float(pd.Series(s)[pd.Series(s) < 0].abs().sum())),
            belge_adedi=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
        ).reset_index()

        ters = base[base["has_ters_kayit"] == True].groupby(["kullanici", "donem"]).agg(
            ters_belge_adedi=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
            ters_borc_tutar=("up_tutar", lambda s: float(pd.Series(s)[pd.Series(s) > 0].sum())),
            ters_alacak_tutar=("up_tutar", lambda s: float(pd.Series(s)[pd.Series(s) < 0].abs().sum())),
        ).reset_index()

        belge_late = base[base["belge_tarihi_dt"].notna() & base["giris_tarihi_dt"].notna()].copy()
        belge_late["belge_to_giris_days"] = (belge_late["giris_tarihi_dt"] - belge_late["belge_tarihi_dt"]).dt.days
        belge_late = belge_late[belge_late["belge_to_giris_days"] > 7]
        belge_late_grp = belge_late.groupby(["kullanici", "donem"]).agg(
            belge_giris_7gun_belge=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
            belge_giris_7gun_tutar=("up_tutar", lambda s: float(pd.Series(s).abs().sum())),
        ).reset_index()

        kayit_late = base[base["kayit_tarihi_dt"].notna() & base["giris_tarihi_dt"].notna()].copy()
        kayit_late["kayit_to_giris_days"] = (kayit_late["giris_tarihi_dt"] - kayit_late["kayit_tarihi_dt"]).dt.days
        kayit_late = kayit_late[kayit_late["kayit_to_giris_days"] > 7]
        kayit_late_grp = kayit_late.groupby(["kullanici", "donem"]).agg(
            kayit_giris_7gun_belge=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
            kayit_giris_7gun_tutar=("up_tutar", lambda s: float(pd.Series(s).abs().sum())),
        ).reset_index()

        result = grp.merge(ters, on=["kullanici", "donem"], how="left").merge(belge_late_grp, on=["kullanici", "donem"], how="left").merge(kayit_late_grp, on=["kullanici", "donem"], how="left")
        for col in ["ters_belge_adedi", "ters_borc_tutar", "ters_alacak_tutar", "belge_giris_7gun_belge", "belge_giris_7gun_tutar", "kayit_giris_7gun_belge", "kayit_giris_7gun_tutar"]:
            if col not in result.columns:
                result[col] = 0
            result[col] = result[col].fillna(0)

        result = result.sort_values(["donem", "belge_adedi", "borc_tutar", "alacak_tutar"], ascending=[True, False, False, False])
        headers = [
            "Kullanıcı", "Dönem", "Belge Türleri", "Vergi Göstergeleri", "Satıcı Kodları", "Satıcı Unvanları",
            "Borç Tutar", "Alacak Tutar", "Belge Adedi", "Ters Kayıt Belge Adedi", "Ters Kayıt Borç", "Ters Kayıt Alacak",
            "Belge Tarihi→Giriş 7+ Belge", "Belge Tarihi→Giriş 7+ Tutar", "Kayıt Tarihi→Giriş 7+ Belge", "Kayıt Tarihi→Giriş 7+ Tutar"
        ] if self.language == "tr" else [
            "User", "Period", "Document Types", "Tax Indicators", "Vendor Codes", "Vendor Names",
            "Debit Amount", "Credit Amount", "Document Count", "Reversal Document Count", "Reversal Debit", "Reversal Credit",
            "Doc Date→Entry 7+ Docs", "Doc Date→Entry 7+ Amount", "Posting Date→Entry 7+ Docs", "Posting Date→Entry 7+ Amount"
        ]
        rows=[]; badges=[]
        for _, r in result.iterrows():
            row=[
                r["kullanici"] or "-", r["donem"] or "-", r["belge_turleri"] or "-", r["vergi_gostergeleri"] or "-",
                r["satici_seti"] or "-", r["satici_unvan_seti"] or "-",
                format_number(r["borc_tutar"]), format_number(r["alacak_tutar"]), int(r["belge_adedi"]), int(r["ters_belge_adedi"]),
                format_number(r["ters_borc_tutar"]), format_number(r["ters_alacak_tutar"]), int(r["belge_giris_7gun_belge"]), format_number(r["belge_giris_7gun_tutar"]),
                int(r["kayit_giris_7gun_belge"]), format_number(r["kayit_giris_7gun_tutar"])
            ]
            warn = (safe_float(r["ters_belge_adedi"]) > 0) or (safe_float(r["belge_giris_7gun_belge"]) > 0) or (safe_float(r["kayit_giris_7gun_belge"]) > 0)
            badge_row=[None,None,None,None,None,None,None,None,None, "warn" if safe_float(r["ters_belge_adedi"])>0 else None, None, None, "danger" if safe_float(r["belge_giris_7gun_belge"])>0 else None, "danger" if safe_float(r["belge_giris_7gun_tutar"])>0 else None, "danger" if safe_float(r["kayit_giris_7gun_belge"])>0 else None, "danger" if safe_float(r["kayit_giris_7gun_tutar"])>0 else None]
            if warn:
                badge_row[0] = "warn"
            rows.append(row); badges.append(badge_row)
        return headers, rows or [["-" for _ in headers]], badges or [[None for _ in headers]]

    def build_muavin_tax_based_result(self, df: pd.DataFrame):
        base = df.copy()
        tax_accounts = {"360", "191", "190"}

        dup_headers = [
            "Satıcı Kodu/Karşı Kayıt Hesabı", "Unvan", "Referans", "Kayıt Adedi", "Dönemler", "Kullanıcılar", "Belge Türleri"
        ] if self.language == "tr" else [
            "Vendor/Contra Account", "Name", "Reference", "Entry Count", "Periods", "Users", "Document Types"
        ]
        dup_rows = []
        dup_badges = []
        dup = base[(base["referans"].astype(str).str.strip() != "")].groupby(["karsi_hesap", "karsi_hesap_tanimi", "referans"]).agg(
            kayit_adedi=("belge_anahtar", "size"),
            donemler=("donem", lambda s: safe_unique_join(sorted(set(s)), limit=20)),
            kullanicilar=("kullanici", lambda s: safe_unique_join(s.unique(), limit=20)),
            belge_turleri=("belge_turu", lambda s: safe_unique_join(s.unique(), limit=20)),
        ).reset_index()
        dup = dup[dup["kayit_adedi"] > 1].sort_values(["kayit_adedi", "karsi_hesap", "referans"], ascending=[False, True, True])
        for _, r in dup.iterrows():
            dup_rows.append([
                r["karsi_hesap"] or "-", r["karsi_hesap_tanimi"] or "-", r["referans"] or "-", int(r["kayit_adedi"]),
                r["donemler"] or "-", r["kullanicilar"] or "-", r["belge_turleri"] or "-"
            ])
            dup_badges.append(["warn", None, None, "danger", None, None, None])
        if not dup_rows:
            dup_rows = [["-" for _ in dup_headers]]
            dup_badges = [[None for _ in dup_headers]]

        proj_headers = [
            "Satıcı Kodu/Karşı Kayıt Hesabı", "Unvan", "Dönem", "Ana Hesaplar", "Ana Hesap Adları", "Vergi Göstergeleri", "Projeksiyon Özeti"
        ] if self.language == "tr" else [
            "Vendor/Contra Account", "Name", "Period", "Main Accounts", "Main Account Names", "Tax Indicators", "Projection Summary"
        ]
        proj_rows = []
        proj_badges = []
        proj = base[base["hesap_prefix"].isin(tax_accounts)].copy()
        if not proj.empty:
            proj_grp = proj.groupby(["karsi_hesap", "karsi_hesap_tanimi", "donem"]).agg(
                ana_hesaplar=("ana_hesap", lambda s: safe_unique_join(s.unique(), limit=20)),
                ana_hesap_adlari=("ana_hesap_adi", lambda s: safe_unique_join(s.unique(), limit=20)),
                vergi_gostergeleri=("vergi_gostergesi", lambda s: safe_unique_join([x for x in s.unique() if str(x).strip()], limit=20) or "-"),
            ).reset_index()
            proj_grp["projeksiyon_ozeti"] = proj_grp.apply(
                lambda x: f"{x['donem']}: {x['ana_hesaplar']} / {x['vergi_gostergeleri']}", axis=1
            )
            proj_grp = proj_grp.sort_values(["karsi_hesap", "donem"])
            for _, r in proj_grp.iterrows():
                proj_rows.append([
                    r["karsi_hesap"] or "-", r["karsi_hesap_tanimi"] or "-", r["donem"] or "-", r["ana_hesaplar"] or "-",
                    r["ana_hesap_adlari"] or "-", r["vergi_gostergeleri"] or "-", r["projeksiyon_ozeti"] or "-"
                ])
                proj_badges.append([None, None, None, "warn", None, "warn", None])
        if not proj_rows:
            proj_rows = [["-" for _ in proj_headers]]
            proj_badges = [[None for _ in proj_headers]]

        ind_headers = [
            "Vergi Göstergesi", "Dönemler", "Toplam Tutar", "Belge Adedi", "Satıcı Kodları/Karşı Kayıtlar", "Satıcı Unvanları"
        ] if self.language == "tr" else [
            "Tax Indicator", "Periods", "Total Amount", "Document Count", "Vendor/Contra Accounts", "Vendor Names"
        ]
        ind_rows = []
        ind_badges = []
        ind = base[base["has_ters_kayit"] != True].copy()
        ind["vergi_gostergesi"] = ind["vergi_gostergesi"].replace("", "-")
        ind_grp = ind.groupby(["vergi_gostergesi", "donem"]).agg(
            toplam_tutar=("up_tutar", lambda s: float(pd.Series(s).sum())),
            belge_adedi=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
            satici_kodlari=("karsi_hesap", lambda s: safe_unique_join(s.unique(), limit=20)),
            satici_unvanlari=("karsi_hesap_tanimi", lambda s: safe_unique_join(s.unique(), limit=20)),
        ).reset_index()
        if not ind_grp.empty:
            summary = ind_grp.groupby("vergi_gostergesi").agg(
                donemler=("donem", lambda s: safe_unique_join(sorted(set(s)), limit=20)),
                toplam_tutar=("toplam_tutar", "sum"),
                belge_adedi=("belge_adedi", "sum"),
                satici_kodlari=("satici_kodlari", lambda s: safe_unique_join(s.tolist(), limit=20)),
                satici_unvanlari=("satici_unvanlari", lambda s: safe_unique_join(s.tolist(), limit=20)),
            ).reset_index().sort_values(["vergi_gostergesi"])
            for _, r in summary.iterrows():
                ind_rows.append([
                    r["vergi_gostergesi"] or "-", r["donemler"] or "-", format_number(r["toplam_tutar"]), int(r["belge_adedi"]),
                    r["satici_kodlari"] or "-", r["satici_unvanlari"] or "-"
                ])
                ind_badges.append(["warn" if (r["vergi_gostergesi"] or "-") != "-" else None, None, None, None, None, None])
        if not ind_rows:
            ind_rows = [["-" for _ in ind_headers]]
            ind_badges = [[None for _ in ind_headers]]

        return {
            "dup": (dup_headers, dup_rows, dup_badges),
            "projection": (proj_headers, proj_rows, proj_badges),
            "indicator": (ind_headers, ind_rows, ind_badges),
        }

    def build_muavin_account_content_result(self, df: pd.DataFrame):
        base = df.copy()
        base["metin_kisa"] = base["metin"].fillna("").astype(str).str.strip().replace("", "-")
        text_pairs = base["metin_kisa"].fillna("").apply(build_text_group_nlp)
        text_pairs_df = pd.DataFrame(text_pairs.tolist(), index=base.index, columns=["metin_grubu_nlp", "frekans_skoru"])
        base["metin_grubu_nlp"] = text_pairs_df["metin_grubu_nlp"].fillna("-")

        text_headers = [
            "Ana Hesap", "Ana Hesap Adı", "Metin Grubu", "Dönemler", "Belge Adedi", "Toplam Tutar"
        ] if self.language == "tr" else [
            "Main Account", "Main Account Name", "Text Group", "Periods", "Document Count", "Total Amount"
        ]
        text_rows = []
        text_badges = []
        text_grp = base.groupby(["ana_hesap", "ana_hesap_adi", "metin_grubu_nlp"]).agg(
            donemler=("donem", lambda s: safe_unique_join(sorted(set(s)), limit=20)),
            belge_adedi=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
            toplam_tutar=("up_tutar", lambda s: float(pd.Series(s).sum())),
        ).reset_index().sort_values(["ana_hesap", "belge_adedi", "toplam_tutar"], ascending=[True, False, False])
        for _, r in text_grp.iterrows():
            text_rows.append([
                r["ana_hesap"] or "-", r["ana_hesap_adi"] or "-", r["metin_grubu_nlp"] or "-", r["donemler"] or "-",
                int(r["belge_adedi"]), format_number(r["toplam_tutar"])
            ])
            text_badges.append([None, None, None, None, None, None])
        if not text_rows:
            text_rows = [["-" for _ in text_headers]]
            text_badges = [[None for _ in text_headers]]

        relation_headers = [
            "Ana Hesap", "Ana Hesap Adı", "Belge Numarası", "Karşı Kayıt Hesapları", "Karşı Kayıt Hesap Adları", "Belge Adedi", "Toplam Tutar"
        ] if self.language == "tr" else [
            "Main Account", "Main Account Name", "Document Number", "Contra Accounts", "Contra Account Names", "Document Count", "Total Amount"
        ]
        relation_rows = []
        relation_badges = []
        relation_grp = base.groupby(["ana_hesap", "ana_hesap_adi", "belge_numarasi"]).agg(
            karsi_hesaplar=("karsi_hesap", lambda s: safe_unique_join(s.unique(), limit=20)),
            karsi_hesap_adlari=("karsi_hesap_tanimi", lambda s: safe_unique_join(s.unique(), limit=20)),
            belge_adedi=("belge_anahtar", lambda s: s.replace("", pd.NA).dropna().nunique()),
            toplam_tutar=("up_tutar", lambda s: float(pd.Series(s).sum())),
        ).reset_index()
        relation_grp = relation_grp[relation_grp["belge_numarasi"].astype(str).str.strip() != ""].sort_values(["ana_hesap", "belge_numarasi"])
        for _, r in relation_grp.iterrows():
            relation_rows.append([
                r["ana_hesap"] or "-", r["ana_hesap_adi"] or "-", r["belge_numarasi"] or "-", r["karsi_hesaplar"] or "-",
                r["karsi_hesap_adlari"] or "-", int(r["belge_adedi"]), format_number(r["toplam_tutar"])
            ])
            relation_badges.append([None, None, None, "warn", None, None, None])
        if not relation_rows:
            relation_rows = [["-" for _ in relation_headers]]
            relation_badges = [[None for _ in relation_headers]]

        cost_headers = [
            "Ana Hesap", "Ana Hesap Adı", "Masraf Yeri", "Masraf Yeri Adı", "Dönemler", "Toplam Tutar"
        ] if self.language == "tr" else [
            "Main Account", "Main Account Name", "Cost Center", "Cost Center Name", "Periods", "Total Amount"
        ]
        cost_rows = []
        cost_badges = []
        cc_base = base[base["has_ters_kayit"] != True].copy()
        cc_grp = cc_base.groupby(["ana_hesap", "ana_hesap_adi", "masraf_yeri", "masraf_yeri_tanimi"]).agg(
            donemler=("donem", lambda s: safe_unique_join(sorted(set(s)), limit=20)),
            toplam_tutar=("up_tutar", lambda s: float(pd.Series(s).sum())),
        ).reset_index().sort_values(["ana_hesap", "masraf_yeri"])
        for _, r in cc_grp.iterrows():
            cost_rows.append([
                r["ana_hesap"] or "-", r["ana_hesap_adi"] or "-", r["masraf_yeri"] or "-", r["masraf_yeri_tanimi"] or "-",
                r["donemler"] or "-", format_number(r["toplam_tutar"])
            ])
            cost_badges.append([None, None, None, None, None, None])
        if not cost_rows:
            cost_rows = [["-" for _ in cost_headers]]
            cost_badges = [[None for _ in cost_headers]]

        return {
            "text": (text_headers, text_rows, text_badges),
            "relation": (relation_headers, relation_rows, relation_badges),
            "cost": (cost_headers, cost_rows, cost_badges),
        }

    def fill_period_combos(self):
        self.cmb_current.blockSignals(True)
        self.cmb_previous.blockSignals(True)

        self.cmb_current.clear()
        self.cmb_previous.clear()

        selectable_periods = [p for p in self.available_periods if p != OPENING_PERIOD_TR]

        for tr_val in selectable_periods:
            self.cmb_current.addItem(self.period_display_label(tr_val), tr_val)

        prev_options = [p for p in selectable_periods if p != self.current_period]
        for tr_val in prev_options:
            self.cmb_previous.addItem(self.period_display_label(tr_val), tr_val)

        curr_idx = self.cmb_current.findData(self.current_period)
        if curr_idx >= 0:
            self.cmb_current.setCurrentIndex(curr_idx)

        if self.previous_period == self.current_period and prev_options:
            self.previous_period = prev_options[-1]

        prev_idx = self.cmb_previous.findData(self.previous_period)
        if prev_idx < 0 and prev_options:
            self.previous_period = prev_options[-1]
            prev_idx = self.cmb_previous.findData(self.previous_period)

        if prev_idx >= 0:
            self.cmb_previous.setCurrentIndex(prev_idx)

        self.cmb_current.blockSignals(False)
        self.cmb_previous.blockSignals(False)

    def update_progress(self):
        percent = 100 if self.analysis_has_run else (50 if (self.tb_file_path or self.plcc_file_path) else 0)
        self.progress.setValue(percent)
        self.proc_value.setText(f"{percent}%")

    
    def populate_tb_table(self):
        periods = self.available_periods[:]

        if self.tb_financial_filter == "balance":
            periods = normalize_balance_periods(self.available_periods[:])
            summary_rows = build_balance_sheet_summary(
                self.tb_rows_cache,
                periods,
                current_period=self.current_period,
                previous_period=self.previous_period
            )
            headers = (
                ["Bilanço Kalemi"]
                + [self.period_display_label(p) for p in periods]
                + [self.t("total"),
                   f"{self.t('compareChange')} ({self.period_display_label(self.current_period)} / {self.period_display_label(self.previous_period)})",
                   self.t("status")]
            )

            self.tb_table.clear()
            self.tb_table.setColumnCount(len(headers))
            self.tb_table.setHorizontalHeaderLabels(headers)
            self.tb_table.setRowCount(len(summary_rows))

            for r, row in enumerate(summary_rows):
                data = (
                    [row.get("label", "")]
                    + [row.get("values", {}).get(p, "0") for p in periods]
                    + [
                        row.get("values", {}).get("TOPLAM", "0"),
                        row.get("degisim", "-"),
                    ]
                )

                row_type = row.get("rowType", "group")
                status_text = row.get("varianceStatus", "Normal")
                if row.get("label") == "AKTİF - PASİF FARKI":
                    has_diff = any(abs(safe_float(v)) > 1e-6 for k, v in row.get("valuesNumeric", {}).items() if k != "TOPLAM")
                    status_text = self.t("reverseState") if has_diff else row.get("varianceStatus", "Normal")
                data.append(status_text)

                first_numeric_col = 1
                degisim_col = len(data) - 2

                for c, val in enumerate(data):
                    align = Qt.AlignRight | Qt.AlignVCenter if first_numeric_col <= c <= degisim_col else Qt.AlignLeft | Qt.AlignVCenter
                    badge = None
                    if row_type == "subtotal" and c < len(data) - 1:
                        badge = "subtotal"
                    elif row_type == "danger":
                        badge = "danger"
                    elif row_type == "success":
                        badge = "success"
                    elif c == len(data) - 1:
                        if row.get("label") == "AKTİF - PASİF FARKI":
                            badge = "danger" if status_text == self.t("reverseState") else ("danger" if row.get("criticalVariance") else "success")
                        else:
                            badge = "danger" if row.get("criticalVariance") else "success"
                    elif c == degisim_col and row.get("criticalVariance"):
                        badge = "danger"
                    set_table_item(self.tb_table, r, c, str(val), align, badge)

            widths = [280] + [120] * len(periods) + [120, 150, 240]
            for i, w in enumerate(widths[:self.tb_table.columnCount()]):
                self.tb_table.setColumnWidth(i, w)
            self.enable_manual_column_resize(self.tb_table)
            return

        if self.tb_financial_filter == "income":
            periods = self.available_periods[:]
            summary_rows = build_income_statement_summary(
                self.tb_rows_cache,
                periods,
                current_period=self.current_period,
                previous_period=self.previous_period
            )
            headers = [
                "Gelir Tablosu Satırı",
                *[self.period_display_label(p) for p in periods],
                self.t("total"),
                f"{self.t('compareChange')} ({self.period_display_label(self.current_period)} / {self.period_display_label(self.previous_period)})",
                self.t("status"),
            ]

            self.tb_table.clear()
            self.tb_table.setColumnCount(len(headers))
            self.tb_table.setHorizontalHeaderLabels(headers)
            self.tb_table.setRowCount(len(summary_rows))

            for r, row in enumerate(summary_rows):
                data = [
                    row.get("label", ""),
                    *[row.get("values", {}).get(p, "0") for p in periods],
                    row.get("values", {}).get("TOPLAM", "0"),
                    row.get("degisim", "-"),
                    row.get("varianceStatus", "Normal"),
                ]
                degisim_col = len(data) - 2

                for c, val in enumerate(data):
                    align = Qt.AlignRight | Qt.AlignVCenter if 1 <= c <= degisim_col else Qt.AlignLeft | Qt.AlignVCenter
                    badge = None
                    if row.get("rowType") in ["subtotal", "group"] and c < len(data) - 1:
                        badge = "subtotal"
                    if c == degisim_col and row.get("criticalVariance"):
                        badge = "danger"
                    if c == len(data) - 1:
                        badge = "danger" if row.get("criticalVariance") else "success"
                    set_table_item(self.tb_table, r, c, str(val), align, badge)

            widths = [360] + [120] * len(periods) + [120, 150, 150]
            for i, w in enumerate(widths[:self.tb_table.columnCount()]):
                self.tb_table.setColumnWidth(i, w)
            return

        periods_no_opening = [p for p in periods if p != OPENING_PERIOD_TR]

        headers = [
            "Ana Hesap",
            self.t("account"),
            "Sınıf",
            self.t("accountName"),
            self.t("opening"),
            *[self.period_display_label(p) for p in periods_no_opening],
            self.t("total"),
            f"{self.t('compareChange')} ({self.period_display_label(self.current_period)} / {self.period_display_label(self.previous_period)})",
            "Beklenen Yön",
            "Toplam Yön",
            self.t("responsible"),
            self.t("note"),
            self.t("status"),
        ]

        self.tb_table.clear()
        self.tb_table.setColumnCount(len(headers))
        self.tb_table.setHorizontalHeaderLabels(headers)

        rows = self.tb_rows_cache[:]

        if self.analysis_filter == "findings":
            rows = [r for r in rows if r.get("not") or r.get("kontrol") == "Ters duruyor"]

        q = self.search_input.text().strip().lower()
        if q:
            rows = [
                r for r in rows
                if row_text_search_match([
                    r.get("anaHesapTam", r.get("anaHesap", "")),
                    r.get("hesap", ""),
                    r.get("sinif", ""),
                    r.get("hesapAdi", ""),
                    r.get("opening", ""),
                    r.get("toplam", ""),
                    r.get("degisim", ""),
                    r.get("beklenenYon", ""),
                    r.get("toplamYon", ""),
                    r.get("sorumlu", ""),
                    r.get("not", ""),
                    " ".join(str(v) for v in r.get("values", {}).values()),
                ], q, getattr(self, "analysis_search_mode", "contains"))
            ]

        self.tb_table.setRowCount(len(rows))

        for r, row in enumerate(rows):
            data = [
                row.get("anaHesapTam", row.get("anaHesap", "")),
                row.get("hesap", ""),
                row.get("sinif", ""),
                row.get("hesapAdiEn", "") if self.language == "en" else row.get("hesapAdi", ""),
                row.get("opening", "0"),
                *[row.get("values", {}).get(p, "0") for p in periods_no_opening],
                row.get("toplam", "0"),
                row.get("degisim", "-"),
                row.get("beklenenYon", "-"),
                row.get("toplamYon", "-"),
                row.get("sorumlu", "") or "-",
                row.get("noteEn", "") if self.language == "en" else (row.get("not", "") or "-"),
                self.t("reverseState") if row.get("kontrol") == "Ters duruyor" else self.t("normalState"),
            ]

            last_status_col = len(data) - 1
            first_numeric_col = 4
            last_numeric_col = 4 + len(periods_no_opening) + 1

            for c, val in enumerate(data):
                align = Qt.AlignRight | Qt.AlignVCenter if first_numeric_col <= c <= last_numeric_col else Qt.AlignLeft | Qt.AlignVCenter
                badge = None
                if c == last_status_col:
                    badge = "danger" if row.get("kontrol") == "Ters duruyor" else "success"
                set_table_item(self.tb_table, r, c, str(val), align, badge)

        widths = [140, 80, 60, 220, 110] + [100] * len(periods_no_opening) + [120, 140, 120, 120, 160, 220, 130]
        for i, w in enumerate(widths[:self.tb_table.columnCount()]):
            self.tb_table.setColumnWidth(i, w)

    def populate_plcc_table(self):
        periods = self.available_periods[:]
        headers = [
            self.t("costCenter"),
            self.t("description"),
            self.t("account"),
            self.t("accountName"),
            *[self.period_display_label(p) for p in periods],
            self.t("total"),
            f"{self.t('compareChange')} ({self.period_display_label(self.current_period)} / {self.period_display_label(self.previous_period)})",
            self.t("responsible"),
            self.t("note"),
            self.t("status"),
        ]

        self.plcc_table.clear()
        self.plcc_table.setColumnCount(len(headers))
        self.plcc_table.setHorizontalHeaderLabels(headers)

        detail_rows = self.plcc_detail_cache[:]
        subtotals = self.plcc_subtotal_cache[:]

        if self.analysis_filter == "findings":
            detail_rows = [r for r in detail_rows if r.get("not") or r.get("kontrol") == "Ters duruyor"]

        q = self.search_input.text().strip().lower()
        if q:
            detail_rows = [
                r for r in detail_rows
                if row_text_search_match([
                    r.get("masrafYeri", ""),
                    r.get("ad", ""),
                    r.get("hesap", ""),
                    r.get("hesapAdi", ""),
                    r.get("toplam", ""),
                    r.get("degisim", ""),
                    r.get("sorumlu", ""),
                    r.get("not", ""),
                    " ".join(str(v) for v in r.get("values", {}).values()),
                ], q, getattr(self, "analysis_search_mode", "contains"))
            ]

        subtotal_map = {s["masrafYeri"]: s for s in subtotals}
        grouped = {}
        for row in detail_rows:
            grouped.setdefault(row["masrafYeri"], []).append(row)

        all_masraf_keys = sorted(
            set(list(grouped.keys()) + ([] if self.analysis_filter == "findings" else list(subtotal_map.keys()))),
            key=masraf_sort_key
        )

        display_rows = []
        for masraf in all_masraf_keys:
            details = grouped.get(masraf, [])
            for row in sorted(details, key=lambda x: (int(x.get("hesap")) if str(x.get("hesap")).isdigit() else str(x.get("hesap")))):
                display_rows.append(("detail", row))
            if self.analysis_filter != "findings" and masraf in subtotal_map:
                display_rows.append(("subtotal", subtotal_map[masraf]))

        self.plcc_table.setRowCount(len(display_rows))

        for r, (row_type, row) in enumerate(display_rows):
            if row_type == "detail":
                data = [
                    row.get("masrafYeri", "") or "",
                    row.get("adEn", "") if self.language == "en" else row.get("ad", ""),
                    row.get("hesap", ""),
                    row.get("hesapAdiEn", "") if self.language == "en" else row.get("hesapAdi", ""),
                    *[row.get("values", {}).get(p, "0") for p in periods],
                    row.get("toplam", "0"),
                    row.get("degisim", "-"),
                    row.get("sorumlu", "") or "-",
                    row.get("noteEn", "") if self.language == "en" else (row.get("not", "") or "-"),
                    self.t("reverseState") if row.get("kontrol") == "Ters duruyor" else self.t("normalState"),
                ]
                status_col = len(data) - 1
                first_numeric_col = 4
                last_numeric_col = 4 + len(periods)

                for c, val in enumerate(data):
                    align = Qt.AlignRight | Qt.AlignVCenter if first_numeric_col <= c <= last_numeric_col else Qt.AlignLeft | Qt.AlignVCenter
                    badge = None
                    if c == status_col:
                        badge = "danger" if row.get("kontrol") == "Ters duruyor" else "success"
                    set_table_item(self.plcc_table, r, c, str(val), align, badge)
            else:
                subtotal_label = f'{row.get("adEn", "") if self.language == "en" else row.get("ad", "")} {self.t("subtotal")}'.strip()
                data = [
                    row.get("masrafYeri", "") or "",
                    subtotal_label,
                    "-",
                    "-",
                    *[row.get("values", {}).get(p, "0") for p in periods],
                    row.get("toplam", "0"),
                    row.get("degisim", "-"),
                    "-",
                    "-",
                    "-",
                ]
                first_numeric_col = 4
                last_numeric_col = 4 + len(periods)

                for c, val in enumerate(data):
                    align = Qt.AlignRight | Qt.AlignVCenter if first_numeric_col <= c <= last_numeric_col else Qt.AlignLeft | Qt.AlignVCenter
                    set_table_item(self.plcc_table, r, c, str(val), align, "subtotal")

        widths = [110, 180, 140, 220] + [100] * len(periods) + [120, 150, 160, 220, 130]
        for i, w in enumerate(widths[:self.plcc_table.columnCount()]):
            self.plcc_table.setColumnWidth(i, w)

    def apply_row_heights(self):
        height = 30 if self.row_density == "compact" else 48 if self.row_density == "wide" else 38
        for t in [self.tb_table, self.plcc_table]:
            for r in range(t.rowCount()):
                t.setRowHeight(r, height)

    def apply_notes_row_heights(self):
        height = 30 if self.notes_row_density == "compact" else 48 if self.notes_row_density == "wide" else 38
        for r in range(self.notes_table.rowCount()):
            self.notes_table.setRowHeight(r, height)

    def set_language(self, lang: str):
        self.language = lang
        self.refresh_all()

    def set_active_view(self, view: str):
        perms = getattr(self, "current_user_permissions", normalize_user_permissions(None))
        if view == "tb" and not perms.get("analysis_tb", True):
            return
        if view == "plcc" and not perms.get("analysis_plcc", True):
            return
        self.active_view = view
        self.refresh_analysis_ui()
        self.refresh_analysis_tables()

    def set_analysis_filter(self, value: str):
        self.analysis_filter = value
        self.refresh_analysis_ui()
        self.refresh_analysis_tables()

    def set_tb_financial_filter(self, value: str):
        perms = getattr(self, "current_user_permissions", normalize_user_permissions(None))
        if value == "balance" and not perms.get("analysis_balance", True):
            return
        if value == "income" and not perms.get("analysis_income", True):
            return
        self.tb_financial_filter = value
        self.refresh_analysis_ui()
        self.refresh_analysis_tables()

    def on_density_changed(self):
        idx = self.cmb_density.currentIndex()
        self.row_density = {0: "compact", 1: "normal", 2: "wide"}.get(idx, "normal")
        self.apply_row_heights()

    def on_notes_density_changed(self):
        idx = self.cmb_notes_density.currentIndex()
        self.notes_row_density = {0: "compact", 1: "normal", 2: "wide"}.get(idx, "normal")
        self.apply_notes_row_heights()

    def on_period_changed(self):
        current = self.cmb_current.currentData()
        previous = self.cmb_previous.currentData()

        if current:
            self.current_period = current
        if previous:
            self.previous_period = previous

        selectable_periods = [p for p in self.available_periods if p != OPENING_PERIOD_TR]
        valid_previous = [p for p in selectable_periods if p != self.current_period]
        if self.previous_period not in valid_previous:
            self.previous_period = valid_previous[-1] if valid_previous else self.current_period

        if self.analysis_has_run:
            self.rebuild_caches()
        self.refresh_all()

    def load_tb_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "TB Excel Seç", BASE_DIR, "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        self.begin_busy_state("TB dosyası okunuyor..." if self.language == "tr" else "Reading TB file...", 5)
        started = self.start_background_worker(
            load_single_excel_payload,
            self.on_tb_file_loaded_async,
            self.on_tb_file_failed_async,
            path,
            "TB",
            self.language,
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def load_tb_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "TB Excel Seç", BASE_DIR, "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        self.begin_busy_state("TB dosyası okunuyor..." if self.language == "tr" else "Reading TB file...", 5)
        started = self.start_background_worker(
            load_single_excel_payload,
            self.on_tb_file_loaded_async,
            self.on_tb_file_failed_async,
            path,
            "TB",
            self.language,
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def load_plcc_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "PL-CC Excel Seç", BASE_DIR, "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        self.begin_busy_state("PL-CC dosyası okunuyor..." if self.language == "tr" else "Reading PL-CC file...", 5)
        started = self.start_background_worker(
            load_single_excel_payload,
            self.on_plcc_file_loaded_async,
            self.on_plcc_file_failed_async,
            path,
            "PL-CC",
            self.language,
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def start_analysis(self):
        if not self.tb_file_path and not self.plcc_file_path:
            QMessageBox.warning(self, "Uyarı" if self.language == "tr" else "Warning", "Lütfen önce en az bir TB veya PL-CC dosyası yükleyin." if self.language == "tr" else "Load at least one TB or PL-CC file first.")
            return

        self.begin_busy_state("Analiz hazırlanıyor..." if self.language == "tr" else "Preparing analysis...", 5)
        started = self.start_background_worker(
            self.background_build_main_analysis,
            self.on_main_analysis_loaded,
            self.on_main_analysis_failed,
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def save_note_from_form(self):
        return


    def save_notes_from_paste(self):
        parsed = []
        for row in range(self.input_note_paste.rowCount()):
            hesap_item = self.input_note_paste.item(row, 0)
            ana_hesap_item = self.input_note_paste.item(row, 1)
            masraf_item = self.input_note_paste.item(row, 2)
            note_item = self.input_note_paste.item(row, 3)

            hesap_raw = hesap_item.text().strip() if hesap_item else ""
            ana_hesap = ana_hesap_item.text().strip() if ana_hesap_item else ""
            masraf = masraf_item.text().strip() if masraf_item else ""
            note_text = note_item.text().strip() if note_item else ""

            if not hesap_raw and not ana_hesap and not masraf and not note_text:
                continue

            hesap = normalize_hesap_prefix(hesap_raw if hesap_raw else ana_hesap)

            if not (hesap or ana_hesap) or not note_text:
                continue

            parsed.append({
                "hesap": hesap,
                "anaHesap": ana_hesap,
                "masrafYeri": masraf,
                "not": note_text,
                "noteEn": note_text,
            })

        if not parsed:
            QMessageBox.warning(self, "Uyarı", "Lütfen Excel verisini dört kolonda yapıştırın.")
            return

        merged = {}
        for item in self.notes:
            hesap = normalize_hesap_prefix(item.get("hesap", ""))
            ana_hesap = str(item.get("anaHesap", "")).strip()
            masraf = str(item.get("masrafYeri", "")).strip()
            note_text = str(item.get("not", "")).strip()
            if (hesap or ana_hesap) and note_text:
                merged[(hesap, ana_hesap, masraf, note_text)] = {
                    "hesap": hesap,
                    "anaHesap": ana_hesap,
                    "masrafYeri": masraf,
                    "not": note_text,
                    "noteEn": str(item.get("noteEn", note_text)).strip() or note_text,
                }

        for item in parsed:
            merged[(item["hesap"], item["anaHesap"], item["masrafYeri"], item["not"])] = item

        self.notes = sorted(
            merged.values(),
            key=lambda x: (
                str(x.get("anaHesap", "")),
                normalize_hesap_prefix(x.get("hesap", "")),
                str(x.get("masrafYeri", "")).lower(),
                str(x.get("not", "")).lower()
            )
        )
        save_notes(self.notes)
        self.input_note_paste.clear_data()

        if self.analysis_has_run:
            self.rebuild_caches()
        self.refresh_all()
        QMessageBox.information(self, "Bilgi", f"{len(parsed)} {self.t('rowsLoaded')}.")

    def save_responsibles_from_paste(self):
        parsed = []
        for row in range(self.input_resp_paste.rowCount()):
            hesap_item = self.input_resp_paste.item(row, 0)
            ana_hesap_item = self.input_resp_paste.item(row, 1)
            sorumlu_item = self.input_resp_paste.item(row, 2)

            hesap_raw = hesap_item.text().strip() if hesap_item else ""
            ana_hesap = ana_hesap_item.text().strip() if ana_hesap_item else ""
            sorumlu = sorumlu_item.text().strip() if sorumlu_item else ""

            if not hesap_raw and not ana_hesap and not sorumlu:
                continue

            hesap = normalize_hesap_prefix(hesap_raw if hesap_raw else ana_hesap)

            if not sorumlu:
                continue
            if not hesap and not ana_hesap:
                continue

            parsed.append({
                "hesap": hesap,
                "anaHesap": ana_hesap,
                "sorumlu": sorumlu
            })

        if not parsed:
            QMessageBox.warning(self, "Uyarı", "Lütfen Excel verisini üç kolonda yapıştırın.")
            return

        merged = {}
        for item in self.responsibles:
            hesap = normalize_hesap_prefix(item.get("hesap", ""))
            ana_hesap = str(item.get("anaHesap", "")).strip()
            sorumlu = str(item.get("sorumlu", "")).strip()
            if (hesap or ana_hesap) and sorumlu:
                merged[(hesap, ana_hesap, sorumlu)] = {
                    "hesap": hesap,
                    "anaHesap": ana_hesap,
                    "sorumlu": sorumlu
                }

        for item in parsed:
            merged[(item["hesap"], item["anaHesap"], item["sorumlu"])] = item

        self.responsibles = sorted(
            merged.values(),
            key=lambda x: (
                str(x.get("anaHesap", "")),
                normalize_hesap_prefix(x.get("hesap", "")),
                str(x.get("sorumlu", "")).lower()
            )
        )
        save_responsibles(self.responsibles)
        self.input_resp_paste.clear_data()

        if self.analysis_has_run:
            self.rebuild_caches()
        self.refresh_all()
        QMessageBox.information(self, "Bilgi", f"{len(parsed)} {self.t('rowsLoaded')}.")
    def save_notes_table_changes(self):
        updated = []
        for row in range(self.notes_table.rowCount()):
            hesap_item = self.notes_table.item(row, 0)
            ana_hesap_item = self.notes_table.item(row, 1)
            masraf_item = self.notes_table.item(row, 2)
            note_item = self.notes_table.item(row, 3)

            hesap = normalize_hesap_prefix(hesap_item.text().strip() if hesap_item else "")
            ana_hesap = (ana_hesap_item.text().strip() if ana_hesap_item else "")
            masraf = (masraf_item.text().strip() if masraf_item else "")
            note_text = (note_item.text().strip() if note_item else "")

            if masraf == self.t("general"):
                masraf = ""

            if not hesap and not ana_hesap and not masraf and not note_text:
                continue
            if not note_text:
                continue
            if not hesap and not ana_hesap:
                continue

            updated.append({
                "hesap": hesap,
                "anaHesap": ana_hesap,
                "masrafYeri": masraf,
                "not": note_text,
                "noteEn": note_text,
            })

        self.notes = updated
        save_notes(self.notes)
        if self.analysis_has_run:
            self.rebuild_caches()
        self.refresh_all()
        QMessageBox.information(self, "Bilgi", "Kayıtlı notlar güncellendi.")
    def delete_selected_note(self):
        row = self.notes_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek için bir not seçin.")
            return
        self.notes_table.removeRow(row)
        self.save_notes_table_changes()


    def save_responsibles_table_changes(self):
        updated = []
        for row in range(self.responsibles_table.rowCount()):
            hesap_item = self.responsibles_table.item(row, 0)
            ana_hesap_item = self.responsibles_table.item(row, 1)
            sorumlu_item = self.responsibles_table.item(row, 2)

            hesap = normalize_hesap_prefix(hesap_item.text().strip() if hesap_item else "")
            ana_hesap = (ana_hesap_item.text().strip() if ana_hesap_item else "")
            sorumlu = (sorumlu_item.text().strip() if sorumlu_item else "")

            if not hesap and not ana_hesap and not sorumlu:
                continue
            if not sorumlu:
                continue
            if not hesap and not ana_hesap:
                continue

            updated.append({
                "hesap": hesap,
                "anaHesap": ana_hesap,
                "sorumlu": sorumlu
            })

        uniq = {}
        for item in updated:
            uniq[(item["hesap"], item["anaHesap"], item["sorumlu"])] = item

        self.responsibles = sorted(
            uniq.values(),
            key=lambda x: (
                str(x.get("anaHesap", "")),
                normalize_hesap_prefix(x.get("hesap", "")),
                str(x.get("sorumlu", "")).lower()
            )
        )
        save_responsibles(self.responsibles)
        if self.analysis_has_run:
            self.rebuild_caches()
        self.refresh_all()
        QMessageBox.information(self, "Bilgi", "Kayıtlı sorumlular güncellendi.")
    def delete_selected_responsible(self):
        row = self.responsibles_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek için bir sorumlu seçin.")
            return
        self.responsibles_table.removeRow(row)
        self.save_responsibles_table_changes()

    def auto_fit_active_table(self):
        table = self.tb_table if self.active_view == "tb" else self.plcc_table
        if not table.isVisible() or table.columnCount() == 0:
            return

        table.setUpdatesEnabled(False)
        header = table.horizontalHeader()

        for i in range(table.columnCount()):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)
            table.resizeColumnToContents(i)
            width = table.columnWidth(i)
            table.setColumnWidth(i, min(max(width + 18, 90), 340))

        table.setUpdatesEnabled(True)
        self.enable_manual_column_resize(table)


    def rebuild_caches(self):
        self.sync_period_selection()
        self.tb_rows_cache = self.build_tb_rows() if self.tb_file_path else []
        if self.plcc_file_path:
            self.plcc_detail_cache, self.plcc_subtotal_cache = self.build_plcc_rows()
        else:
            self.plcc_detail_cache, self.plcc_subtotal_cache = [], []

    def jump_to_tb_findings(self):
        self.tabs.setCurrentIndex(1)
        self.set_active_view("tb")
        self.set_analysis_filter("findings")

    def jump_to_plcc_findings(self):
        self.tabs.setCurrentIndex(1)
        self.set_active_view("plcc")
        self.set_analysis_filter("findings")

    def export_muavin_analysis(self):
        df = self.get_filtered_muavin_df()
        if df.empty:
            QMessageBox.warning(self, "Uyarı" if self.language == "tr" else "Warning", "Dışa aktarılacak muavin analizi yok." if self.language == "tr" else "There is no subledger analysis to export.")
            return

        default_path = build_versioned_filename("Muavin_Analiz")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Muavin Analiz Excel Çıktısı Kaydet" if self.language == "tr" else "Save Subledger Analysis Excel Output",
            default_path,
            "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not str(path).lower().endswith('.xlsx'):
            path = f"{path}.xlsx"

        export_cols = [
            "donem", "ana_hesap", "ana_hesap_adi", "referans", "belge_numarasi", "belge_turu",
            "karsi_hesap", "karsi_hesap_tanimi", "islem_kodu", "up_tutar", "belge_pb_tutar", "belge_pb",
            "denklestirme", "metin", "ters_kayit", "kullanici", "vergi_gostergesi", "masraf_yeri", "masraf_yeri_tanimi",
            "finansal_tablo"
        ]
        sheets = [{
            "sheet_name": "Muavin_Ham_Veri" if self.language == "tr" else "Subledger_Raw_Data",
            "df": df[[c for c in export_cols if c in df.columns]].copy()
        }]
        for key, sheet in [("finding", "Bulgu_Metinleri"), ("text", "Metin_Aciklama_Analizi"), ("period", "Donem_Ozeti"), ("doctype", "Belge_Turu_Ozeti"), ("risk_user", "Kullanici_Risk_Heatmap"), ("user", "Hesap_Davranis_Sapmasi"), ("contra", "Karsit_Hesap_Anormallik"), ("drilldown", "Kullanici_TersKayit"), ("document_lines", "Belge_Hesap_Sikligi"), ("account_doc_relation", "Hesap_Belge_Iliski"), ("late7", "Satici_7Gun_Ihlal"), ("doccheck", "Belge_Iliski_Kontrol"), ("dupref", "Mukerrer_Referanslar"), ("duprefdetail", "Referans_Detay_Listesi"), ("taxref", "Vergisel_Kontrol_Ref"), ("taxvendor", "Vergisel_Kontrol_Satici"), ("tax", "Vergi_Gostergesi"), ("cost", "Masraf_Yeri_Trend"), ("docflow", "Belge_Akis_Ozeti"), ("docmatrix", "Belge_Hesap_Matrisi"), ("userflow", "Kullanici_Yon_Zaman"), ("vendor_doc_relation", "Satici_Belge_Iliski")]:
            table = self.muavin_tables[key]
            rows = []
            for r in range(table.rowCount()):
                rows.append([table.item(r, c).text() if table.item(r, c) else "" for c in range(table.columnCount())])
            table_df = pd.DataFrame(rows, columns=[table.horizontalHeaderItem(i).text() for i in range(table.columnCount())])
            sheets.append({"sheet_name": sheet, "df": table_df})

        self.begin_busy_state("Muavin Excel çıktısı hazırlanıyor..." if self.language == "tr" else "Preparing subledger Excel export...", 10)
        started = self.start_background_worker(
            write_multi_sheet_excel,
            self.on_background_export_success,
            self.on_background_export_failed,
            path,
            sheets,
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)

    def export_current_view(self):
        if self.active_view == "tb" and self.tb_financial_filter == "balance":
            periods = normalize_balance_periods(self.available_periods[:])
            summary_rows = build_balance_sheet_summary(
                self.tb_rows_cache,
                periods,
                current_period=self.current_period,
                previous_period=self.previous_period
            )
            headers = (
                ["Bilanço Kalemi"]
                + [self.period_display_label(p) for p in periods]
                + [self.t("total"),
                   f"{self.t('compareChange')} ({self.period_display_label(self.current_period)} / {self.period_display_label(self.previous_period)})",
                   self.t("status")]
            )
            rows = []
            subtotal_indexes = []
            for idx, row in enumerate(summary_rows):
                status_text = row.get("varianceStatus", "Normal")
                if row.get("label") == "AKTİF - PASİF FARKI":
                    has_diff = any(abs(safe_float(v)) > 1e-6 for k, v in row.get("valuesNumeric", {}).items() if k != "TOPLAM")
                    status_text = self.t("reverseState") if has_diff else row.get("varianceStatus", "Normal")
                rows.append(
                    [row.get("label", "")]
                    + [row.get("values", {}).get(p, "0") for p in periods]
                    + [row.get("values", {}).get("TOPLAM", "0"), row.get("degisim", "-"), status_text]
                )
                if row.get("rowType") in ["subtotal", "danger", "success"]:
                    subtotal_indexes.append(idx)

            path, _ = QFileDialog.getSaveFileName(self, "Bilanço Excel Çıktısı Kaydet", build_versioned_filename("TB_Bilanco_Analiz"), "Excel Files (*.xlsx)")
            if not path:
                return
            if not str(path).lower().endswith('.xlsx'):
                path = f"{path}.xlsx"
            self.begin_busy_state("Bilanço Excel çıktısı hazırlanıyor..." if self.language == "tr" else "Preparing balance sheet Excel export...", 20)
            started = self.start_background_worker(
                write_financial_statement_excel_payload,
                self.on_background_export_success,
                self.on_background_export_failed,
                path,
                "VUK Bilanço Özeti",
                headers,
                rows,
                subtotal_indexes,
            )
            if not started:
                self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)
            return

        if self.active_view == "tb" and self.tb_financial_filter == "income":
            periods = self.available_periods[:]
            summary_rows = build_income_statement_summary(
                self.tb_rows_cache,
                periods,
                current_period=self.current_period,
                previous_period=self.previous_period
            )
            headers = [
                "Gelir Tablosu Satırı",
                *[self.period_display_label(p) for p in periods],
                self.t("total"),
                f"{self.t('compareChange')} ({self.period_display_label(self.current_period)} / {self.period_display_label(self.previous_period)})",
                self.t("status"),
            ]
            rows = []
            subtotal_indexes = []
            for idx, row in enumerate(summary_rows):
                rows.append([
                    row.get("label", ""),
                    *[row.get("values", {}).get(p, "0") for p in periods],
                    row.get("values", {}).get("TOPLAM", "0"),
                    row.get("degisim", "-"),
                    row.get("varianceStatus", "Normal"),
                ])
                if row.get("rowType") in ["subtotal", "group"]:
                    subtotal_indexes.append(idx)

            path, _ = QFileDialog.getSaveFileName(self, "Gelir Tablosu Excel Çıktısı Kaydet", build_versioned_filename("TB_Gelir_Tablosu_Analiz"), "Excel Files (*.xlsx)")
            if not path:
                return
            if not str(path).lower().endswith('.xlsx'):
                path = f"{path}.xlsx"
            self.begin_busy_state("Gelir tablosu Excel çıktısı hazırlanıyor..." if self.language == "tr" else "Preparing income statement Excel export...", 20)
            started = self.start_background_worker(
                write_financial_statement_excel_payload,
                self.on_background_export_success,
                self.on_background_export_failed,
                path,
                "VUK Gelir Tablosu Detay",
                headers,
                rows,
                subtotal_indexes,
            )
            if not started:
                self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)
            return

        if self.active_view == "tb":
            table = self.tb_table
            default_name = build_versioned_filename("TB_Analiz")
        else:
            table = self.plcc_table
            default_name = build_versioned_filename("PLCC_Analiz")

        if table.rowCount() == 0:
            QMessageBox.warning(self, self.t("warning"), self.t("noExportData"))
            return

        path, _ = QFileDialog.getSaveFileName(self, "Excel Çıktısı Kaydet", default_name, "Excel Files (*.xlsx)")
        if not path:
            return
        if not str(path).lower().endswith('.xlsx'):
            path = f"{path}.xlsx"

        headers = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
        rows = []
        for r in range(table.rowCount()):
            row_data = []
            for c in range(table.columnCount()):
                item = table.item(r, c)
                row_data.append(item.text() if item else "")
            rows.append(row_data)

        self.begin_busy_state("Excel çıktısı hazırlanıyor..." if self.language == "tr" else "Preparing Excel export...", 20)
        started = self.start_background_worker(
            write_single_sheet_excel,
            self.on_background_export_success,
            self.on_background_export_failed,
            path,
            headers,
            rows,
            "Rapor" if self.language == "tr" else "Report",
        )
        if not started:
            self.end_busy_state("Önce mevcut işlemin bitmesini bekleyin." if self.language == "tr" else "Wait for the current operation to finish.", 0)


def authenticate_user(username: str, password: str) -> bool:
    uname = str(username).strip().lower()
    pwd_hash = hash_password(password)
    for item in load_users():
        if str(item.get("username", "")).strip().lower() == uname and str(item.get("password_hash", "")).strip() == pwd_hash:
            return True
    return False


class LoginWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.language = "tr"
        self.username = ""
        self.setWindowTitle("TB & PL-CC Control Login")
        self.setModal(True)
        self.resize(760, 500)
        self.setMinimumSize(700, 460)
        self.setStyleSheet("""
        QDialog { background: #F3F7FC; }
        QWidget { font-family: 'Segoe UI'; font-size: 13px; color: #0F172A; }
        #loginCard { background: qlineargradient(x1:0,y1:0,x2:1,y2:1, stop:0 #0F172A, stop:0.55 #102A5C, stop:1 #1D4ED8); border-radius: 28px; }
        #whitePanel { background: rgba(255,255,255,0.97); border-radius: 24px; }
        #loginTitle { color: white; font-size: 30px; font-weight: 700; }
        #loginDesc { color: #DCE7FF; font-size: 13px; }
        #softBadge { background: rgba(255,255,255,0.14); color: white; border: 1px solid rgba(255,255,255,0.18); border-radius: 12px; padding: 6px 12px; font-weight: 600; }
        #langButton, #langButtonActive { border-radius: 12px; padding: 8px 14px; font-weight: 700; min-width: 54px; }
        #langButton { background: transparent; border: 1px solid rgba(255,255,255,0.15); color: white; }
        #langButtonActive { background: white; border: 1px solid white; color: #0F172A; }
        #fieldLabel { color: #334155; font-size: 12px; font-weight: 600; }
        QLineEdit { background: white; border: 1px solid #CBD5E1; border-radius: 14px; padding: 10px 12px; min-height: 42px; }
        QPushButton#primaryButton { background: #1D4ED8; color: white; border: 1px solid #1D4ED8; border-radius: 14px; padding: 10px 16px; font-weight: 700; min-height: 42px; }
        QPushButton#secondaryButton { background: white; color: #0F172A; border: 1px solid #CBD5E1; border-radius: 14px; padding: 10px 16px; font-weight: 700; min-height: 42px; }
        """)
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 24, 24, 24)
        card = QFrame()
        card.setObjectName("loginCard")
        root.addWidget(card)
        card_layout = QHBoxLayout(card)
        card_layout.setContentsMargins(26, 26, 26, 26)
        card_layout.setSpacing(18)

        left = QVBoxLayout()
        card_layout.addLayout(left, 1)
        self.login_badge = QLabel()
        self.login_badge.setObjectName("softBadge")
        left.addWidget(self.login_badge, 0, Qt.AlignLeft)
        self.login_title = QLabel()
        self.login_title.setObjectName("loginTitle")
        left.addWidget(self.login_title)
        self.login_desc = QLabel()
        self.login_desc.setObjectName("loginDesc")
        self.login_desc.setWordWrap(True)
        left.addWidget(self.login_desc)
        self.default_info = QLabel()
        self.default_info.setObjectName("softBadge")
        left.addWidget(self.default_info, 0, Qt.AlignLeft)
        left.addStretch()

        lang_row = QHBoxLayout()
        self.login_btn_tr = QPushButton("TR")
        self.login_btn_en = QPushButton("EN")
        self.login_btn_tr.clicked.connect(lambda: self.set_language("tr"))
        self.login_btn_en.clicked.connect(lambda: self.set_language("en"))
        lang_row.addWidget(self.login_btn_tr)
        lang_row.addWidget(self.login_btn_en)
        lang_row.addStretch()
        left.addLayout(lang_row)

        right_panel = QFrame()
        right_panel.setObjectName("whitePanel")
        card_layout.addWidget(right_panel, 0)
        form = QVBoxLayout(right_panel)
        form.setContentsMargins(22, 22, 22, 22)
        form.setSpacing(10)

        self.form_title = QLabel()
        self.form_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #0F172A;")
        form.addWidget(self.form_title)
        self.form_desc = QLabel()
        self.form_desc.setObjectName("fieldLabel")
        self.form_desc.setWordWrap(True)
        form.addWidget(self.form_desc)

        self.lbl_login_username = QLabel()
        self.lbl_login_username.setObjectName("fieldLabel")
        form.addWidget(self.lbl_login_username)
        self.input_login_username = QLineEdit()
        self.input_login_username.returnPressed.connect(self.try_login)
        form.addWidget(self.input_login_username)

        self.lbl_login_password = QLabel()
        self.lbl_login_password.setObjectName("fieldLabel")
        form.addWidget(self.lbl_login_password)
        login_pwd_row = QHBoxLayout()
        login_pwd_row.setSpacing(8)
        self.input_login_password = QLineEdit()
        self.input_login_password.setEchoMode(QLineEdit.Password)
        self.input_login_password.returnPressed.connect(self.try_login)
        login_pwd_row.addWidget(self.input_login_password, 1)
        self.btn_toggle_login_password = QPushButton("Göster")
        self.btn_toggle_login_password.setObjectName("secondaryButton")
        self.btn_toggle_login_password.setMinimumHeight(40)
        self.btn_toggle_login_password.setMinimumWidth(96)
        self.btn_toggle_login_password.clicked.connect(lambda: self.toggle_password_visibility("login_password"))
        login_pwd_row.addWidget(self.btn_toggle_login_password)
        form.addLayout(login_pwd_row)

        self.login_status = QLabel("")
        self.login_status.setStyleSheet("color: #B91C1C; font-size: 12px;")
        self.login_status.setWordWrap(True)
        form.addWidget(self.login_status)

        btn_row = QHBoxLayout()
        self.btn_login = QPushButton()
        self.btn_login.setObjectName("primaryButton")
        self.btn_login.clicked.connect(self.try_login)
        btn_row.addWidget(self.btn_login)
        self.btn_login_clear = QPushButton()
        self.btn_login_clear.setObjectName("secondaryButton")
        self.btn_login_clear.clicked.connect(self.clear_form)
        btn_row.addWidget(self.btn_login_clear)
        form.addLayout(btn_row)
        form.addStretch()
        self.refresh_texts()


    def _set_password_echo_mode(self, line_edit, button, visible: bool):
        line_edit.setEchoMode(QLineEdit.Normal if visible else QLineEdit.Password)
        tr = self.language == "tr"
        button.setText(("Gizle" if tr else "Hide") if visible else ("Göster" if tr else "Show"))

    def toggle_password_visibility(self, target: str):
        if target == "login_password":
            visible = self.input_login_password.echoMode() != QLineEdit.Normal
            self._set_password_echo_mode(self.input_login_password, self.btn_toggle_login_password, visible)

    def set_language(self, language: str):
        self.language = language
        self.refresh_texts()

    def refresh_texts(self):
        tr = self.language == "tr"
        self.login_badge.setText("Güvenli Giriş" if tr else "Secure Login")
        self.login_title.setText("TB & PL-CC Control")
        self.login_desc.setText("Programa erişmek için kullanıcı adı ve şifre ile giriş yapın." if tr else "Sign in with your username and password to access the program.")
        self.default_info.setText("Varsayılan giriş: admin / admin" if tr else "Default login: admin / admin")
        self.form_title.setText("Kullanıcı Girişi" if tr else "User Sign In")
        self.form_desc.setText("Dil seçimi burada yapılır; ana ekran ve Excel çıktıları aynı dilde devam eder." if tr else "Choose the language here; the main screen and Excel outputs continue in the same language.")
        self.lbl_login_username.setText("Kullanıcı Adı" if tr else "Username")
        self.lbl_login_password.setText("Şifre" if tr else "Password")
        self._set_password_echo_mode(self.input_login_password, self.btn_toggle_login_password, self.input_login_password.echoMode() == QLineEdit.Normal)
        self.btn_login.setText("Giriş Yap" if tr else "Sign In")
        self.btn_login_clear.setText("Temizle" if tr else "Clear")
        self.input_login_username.setPlaceholderText("Kullanıcı adı" if tr else "Username")
        self.input_login_password.setPlaceholderText("Şifre" if tr else "Password")
        self.login_btn_tr.setObjectName("langButtonActive" if tr else "langButton")
        self.login_btn_en.setObjectName("langButtonActive" if not tr else "langButton")
        self.login_btn_tr.style().unpolish(self.login_btn_tr); self.login_btn_tr.style().polish(self.login_btn_tr)
        self.login_btn_en.style().unpolish(self.login_btn_en); self.login_btn_en.style().polish(self.login_btn_en)

    def clear_form(self):
        self.input_login_username.clear()
        self.input_login_password.clear()
        self._set_password_echo_mode(self.input_login_password, self.btn_toggle_login_password, False)
        self.login_status.setText("")
        self.input_login_username.setFocus()

    def try_login(self):
        username = self.input_login_username.text().strip()
        password = self.input_login_password.text()
        if not username or not password:
            self.login_status.setText("Kullanıcı adı ve şifre boş olamaz." if self.language == "tr" else "Username and password cannot be empty.")
            return
        if authenticate_user(username, password):
            self.username = username
            self.accept()
        else:
            self.login_status.setText("Kullanıcı adı veya şifre hatalı." if self.language == "tr" else "Invalid username or password.")


def main():
    ensure_base_dir()
    app = QApplication(sys.argv)
    try:
        app_font = app.font()
        if app_font.pointSize() <= 0:
            app_font.setPointSize(10)
            app.setFont(app_font)
    except Exception:
        try:
            fallback_font = QFont("Segoe UI", 10)
            app.setFont(fallback_font)
        except Exception:
            pass
    app.setApplicationName("TB & PL-CC Control")
    login = LoginWindow()
    if login.exec() != QDialog.Accepted:
        sys.exit(0)
    window = TbPlCcControlWindow(current_user=login.username or "admin", language=login.language)
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()